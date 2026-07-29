from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections.abc import Mapping, Sequence
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlsplit, urlunsplit

import httpx
from openpyxl import load_workbook

DEFAULT_API_BASE_URL = "http://127.0.0.1:8010"
DEFAULT_SPEC_FILE = Path("data/examples/stock_spec_server_basic.json")
EXPECTED_OUTPUT_MODE = "single_best_cost_valid"
EXPECTED_PRIMARY_STATUS = "valid"
EXPECTED_EXCEL_SHEETS = ("AI-рекомендации", "Матрица компонентов")
SECRET_ENV_MARKERS = ("KEY", "TOKEN", "SECRET", "PASSWORD", "PROXY")
FORBIDDEN_USER_SURFACE_FRAGMENTS = (
    "component_candidate_id",
    "source_candidate_id",
    "llm_rec_",
    "Llm_rec_",
    "raw JSON",
    "raw_json",
    "raw prompt",
    "raw_prompt",
    "raw response",
    "raw_response",
    "web evidence not found",
    "keep engineer",
    "debug",
    "provider internals",
)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run a sanitized production smoke for the single-best server MVP."
    )
    parser.add_argument(
        "--base-url",
        default=os.environ.get("STOCK_API_BASE_URL") or DEFAULT_API_BASE_URL,
        help="stock-api base URL. Defaults to STOCK_API_BASE_URL or local VPS port.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=1200,
        help="HTTP timeout in seconds for the API smoke.",
    )
    source = parser.add_mutually_exclusive_group()
    source.add_argument(
        "--spec-file",
        type=Path,
        help="UTF-8 JSON Stock Spec file. Defaults to data/examples/stock_spec_server_basic.json.",
    )
    source.add_argument("--text", help="Free-form request text to post to stock-api.")
    source.add_argument("--match-run-id", type=int, help="Existing match_run_id to verify.")
    return parser.parse_args(argv)


def run(
    argv: Sequence[str] | None = None,
    *,
    client: httpx.Client | None = None,
    stdout: Any = None,
) -> int:
    args = parse_args(argv)
    owns_client = client is None
    http_client = client or httpx.Client(
        base_url=str(args.base_url).rstrip("/"),
        timeout=args.timeout,
        follow_redirects=False,
    )
    try:
        summary = _run_smoke(args, client=http_client)
    finally:
        if owns_client:
            http_client.close()

    _print_summary(summary, stdout=stdout)
    return 0 if summary["ok"] else 1


def _run_smoke(args: argparse.Namespace, *, client: httpx.Client) -> dict[str, Any]:
    errors: list[dict[str, Any]] = []
    checks: dict[str, bool] = {}

    payload: dict[str, Any] = {}
    match_run_id = args.match_run_id
    if match_run_id is None:
        response = _create_match(args, client=client)
        payload = _safe_json_response(response, check_name="create_match", errors=errors)
        match_run_id = _int_or_none(payload.get("match_run_id"))
        if match_run_id is None:
            errors.append({"check": "match_run_id_present"})
    else:
        response = client.get(f"/api/v1/match/{match_run_id}")
        payload = _safe_json_response(response, check_name="get_match", errors=errors)

    _validate_single_best_payload(payload, checks=checks, errors=errors)

    sheet_names: list[str] = []
    if match_run_id is not None:
        excel_response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")
        sheet_names = _validate_excel_response(
            excel_response,
            checks=checks,
            errors=errors,
        )
    else:
        checks["excel_two_sheets"] = False
        checks["excel_user_surfaces_clean"] = False

    summary = {
        "ok": not errors and all(checks.values()),
        "base_url": _safe_url(str(args.base_url)),
        "match_run_id": match_run_id,
        "checks": checks,
        "excel_sheet_names": sheet_names,
        "errors": errors,
    }
    return _redact_obj(summary, secrets=_env_secret_values())


def _create_match(args: argparse.Namespace, *, client: httpx.Client) -> httpx.Response:
    if args.text is not None:
        return client.post("/api/v1/match", json={"text": args.text})

    spec_file = args.spec_file or DEFAULT_SPEC_FILE
    spec = json.loads(spec_file.read_text(encoding="utf-8"))
    return client.post("/api/v1/match", json={"spec": spec})


def _safe_json_response(
    response: httpx.Response,
    *,
    check_name: str,
    errors: list[dict[str, Any]],
) -> dict[str, Any]:
    if response.status_code >= 400:
        errors.append(
            {
                "check": check_name,
                "http_status": response.status_code,
            }
        )
        return {}
    try:
        payload = response.json()
    except json.JSONDecodeError:
        errors.append({"check": check_name, "parse_status": "invalid_json"})
        return {}
    return payload if isinstance(payload, dict) else {}


def _validate_single_best_payload(
    payload: Mapping[str, Any],
    *,
    checks: dict[str, bool],
    errors: list[dict[str, Any]],
) -> None:
    checks["single_best_output_mode"] = payload.get("output_mode") == EXPECTED_OUTPUT_MODE
    checks["primary_recommendation_valid"] = (
        payload.get("primary_recommendation_status") == EXPECTED_PRIMARY_STATUS
    )
    commercial_summary = payload.get("commercial_summary")
    checks["commercial_summary_present"] = bool(commercial_summary)
    checks["commercial_summary_user_surface_clean"] = not _contains_forbidden_fragments(
        commercial_summary
    )
    _append_failed_checks(
        errors,
        checks,
        [
            "single_best_output_mode",
            "primary_recommendation_valid",
            "commercial_summary_present",
            "commercial_summary_user_surface_clean",
        ],
    )


def _validate_excel_response(
    response: httpx.Response,
    *,
    checks: dict[str, bool],
    errors: list[dict[str, Any]],
) -> list[str]:
    if response.status_code >= 400:
        checks["excel_two_sheets"] = False
        checks["excel_user_surfaces_clean"] = False
        errors.append({"check": "excel_report", "http_status": response.status_code})
        return []

    try:
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    except Exception:
        checks["excel_two_sheets"] = False
        checks["excel_user_surfaces_clean"] = False
        errors.append({"check": "excel_report", "parse_status": "invalid_xlsx"})
        return []

    try:
        sheet_names = list(workbook.sheetnames)
        checks["excel_two_sheets"] = tuple(sheet_names) == EXPECTED_EXCEL_SHEETS
        checks["excel_user_surfaces_clean"] = not _contains_forbidden_fragments(
            _workbook_text(workbook)
        )
    finally:
        workbook.close()

    _append_failed_checks(
        errors,
        checks,
        ["excel_two_sheets", "excel_user_surfaces_clean"],
    )
    return sheet_names


def _workbook_text(workbook: Any) -> str:
    values: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values.extend(str(value) for value in row if value is not None)
    return "\n".join(values)


def _append_failed_checks(
    errors: list[dict[str, Any]],
    checks: Mapping[str, bool],
    names: Sequence[str],
) -> None:
    for name in names:
        if not checks.get(name):
            errors.append({"check": name})


def _contains_forbidden_fragments(value: Any) -> bool:
    text = json.dumps(value, ensure_ascii=False) if not isinstance(value, str) else value
    normalized = _normalize_for_fragment_check(text)
    return any(
        _normalize_for_fragment_check(fragment) in normalized
        for fragment in FORBIDDEN_USER_SURFACE_FRAGMENTS
    )


def _normalize_for_fragment_check(value: str) -> str:
    return re.sub(r"\s+", " ", value).casefold()


def _print_summary(summary: Mapping[str, Any], *, stdout: Any = None) -> None:
    stream = stdout or sys.stdout
    safe_summary = _redact_obj(summary, secrets=_env_secret_values())
    print(json.dumps(safe_summary, ensure_ascii=False, indent=2, sort_keys=True), file=stream)


def _env_secret_values() -> set[str]:
    secrets: set[str] = set()
    for key, value in os.environ.items():
        if not value or len(value) < 4:
            continue
        if any(marker in key.upper() for marker in SECRET_ENV_MARKERS):
            secrets.add(value)
            encoded = quote(value, safe="")
            if encoded != value:
                secrets.add(encoded)
    return secrets


def _redact_obj(value: Any, *, secrets: set[str]) -> Any:
    if isinstance(value, str):
        return _redact_text(value, secrets=secrets)
    if isinstance(value, Mapping):
        return {key: _redact_obj(item, secrets=secrets) for key, item in value.items()}
    if isinstance(value, list):
        return [_redact_obj(item, secrets=secrets) for item in value]
    return value


def _redact_text(value: str, *, secrets: set[str]) -> str:
    redacted = value
    for secret in sorted(secrets, key=len, reverse=True):
        redacted = redacted.replace(secret, "[redacted]")
    return redacted


def _safe_url(value: str) -> str:
    parsed = urlsplit(value)
    host = parsed.hostname or ""
    if parsed.port:
        host = f"{host}:{parsed.port}"
    return urlunsplit((parsed.scheme, host, parsed.path.rstrip("/"), "", ""))


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    raise SystemExit(run())


if __name__ == "__main__":
    main()
