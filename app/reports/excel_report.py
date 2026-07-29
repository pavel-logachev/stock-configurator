from __future__ import annotations

from collections.abc import Mapping
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.db.models import MatchCandidate, MatchRun
from app.reports.commercial_summary import (
    build_grouped_commercial_summary,
    build_primary_commercial_summary,
    commercial_component_name,
    commercial_engineering_confidence,
    commercial_reason_for_option,
    commercial_safe_russian_text,
    grouped_commercial_excel_rows,
    primary_commercial_excel_rows,
)
from app.reports.match_text import (
    as_string_list,
    candidate_article,
    candidate_comment,
    candidate_display_name,
    candidate_outcome,
    human_match_status,
    humanized_checks,
    recommended_action,
    short_conclusion,
    yes_no,
)
from app.reports.recommendation_titles import humanized_recommendation_title
from app.reports.v3_full_category_report import (
    build_v3_full_category_excel_report,
    is_v3_full_category_report,
)
from app.user_facing_text import grouped_engineer_check_summary, sanitize_user_facing_text

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
AI_RECOMMENDATIONS_SHEET = "AI-рекомендации"
COMPONENT_MATRIX_SHEET = "Матрица компонентов"
SUMMARY_SHEET = "Сводка"
CANDIDATES_SHEET = "Варианты со склада"
BUILDS_SHEET = "Сборки из комплектующих"
COMPONENT_CANDIDATES_SHEET = "Варианты компонентов"
CHECKS_SHEET = "Что проверить"
REQUIREMENTS_SHEET = "Извлеченные требования"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
FULL_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
MONEY_NUMBER_FORMAT = "# ##0"
LLM_RECOMMENDATIONS_SHEET = "LLM-рекомендации"
LLM_ENGINEERING_REVIEW_NOTE = (
    "AI-рекомендации сформированы автоматически по складским данным. AI может ошибаться. "
    "Перед использованием в КП обязательна инженерная проверка совместимости и комплектации."
)

SERVER_PRODUCT_GROUP = "server"
NETWORK_PRODUCT_GROUP = "network"
STORAGE_PRODUCT_GROUP = "storage"

SERVER_MATRIX_COVERAGE_ROLES = (
    "ready_server",
    "platform",
    "cpu",
    "ram",
    "ssd",
    "hdd",
    "storage_controller",
    "network_adapter",
)
SERVER_MATRIX_ROW_ROLES = (
    "ready_server",
    "server_platform",
    "cpu",
    "ram",
    "ssd",
    "hdd",
    "storage_controller",
    "network_adapter",
)
NETWORK_MATRIX_ROLES = (
    "switch",
    "router",
    "firewall",
    "access_point",
    "transceiver",
    "dac_cable",
    "cable",
    "license",
    "support",
    "power_supply",
    "stacking_module",
    "other_accessory",
)
STORAGE_MATRIX_ROLES = (
    "storage_system",
    "controller",
    "controller_module",
    "disk_shelf",
    "drive",
    "ssd",
    "hdd",
    "cache",
    "host_port",
    "protocol_module",
    "transceiver",
    "cable",
    "license",
    "support",
    "power_supply",
    "rail_kit",
    "other_accessory",
)
MATRIX_CANDIDATE_ROLE_BY_KEY = {
    "ready_server_candidates": "ready_server",
    "platform_candidates": "server_platform",
    "server_platform_candidates": "server_platform",
    "cpu_candidates": "cpu",
    "ram_candidates": "ram",
    "drive_candidates": "drive",
    "ssd_candidates": "ssd",
    "hdd_candidates": "hdd",
    "storage_controller_candidates": "storage_controller",
    "network_adapter_candidates": "network_adapter",
    "switch_candidates": "switch",
    "router_candidates": "router",
    "firewall_candidates": "firewall",
    "access_point_candidates": "access_point",
    "storage_system_candidates": "storage_system",
    "controller_candidates": "controller",
    "controller_module_candidates": "controller_module",
    "disk_shelf_candidates": "disk_shelf",
    "cache_candidates": "cache",
    "host_port_candidates": "host_port",
    "protocol_module_candidates": "protocol_module",
    "transceiver_candidates": "transceiver",
    "dac_cable_candidates": "dac_cable",
    "cable_candidates": "cable",
    "license_candidates": "license",
    "support_candidates": "support",
    "power_supply_candidates": "power_supply",
    "rail_kit_candidates": "rail_kit",
    "stacking_module_candidates": "stacking_module",
    "other_accessory_candidates": "other_accessory",
}


def build_match_excel_report(match_run: MatchRun) -> bytes:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    if is_v3_full_category_report(report_json):
        return build_v3_full_category_excel_report(match_run)

    workbook = Workbook()
    ai_sheet = workbook.active
    ai_sheet.title = AI_RECOMMENDATIONS_SHEET

    _fill_ai_recommendations_sheet(ai_sheet, match_run)
    _fill_component_matrix_sheet(workbook.create_sheet(COMPONENT_MATRIX_SHEET), match_run)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fill_ai_recommendations_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    sheet.sheet_view.showGridLines = False

    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    configuration_groups = _configuration_groups(match_run)
    recommendations = _ai_recommendations(match_run)
    primary_recommendation = _primary_recommendation(match_run)
    normalized_requirements = _first_report_requirements(match_run)
    current_row = 1
    freeze_row = current_row
    if primary_recommendation:
        commercial = build_primary_commercial_summary(
            report_json,
            primary_recommendation,
            match_run_id=match_run.id,
        )
        if commercial is not None:
            current_row = _write_primary_commercial_summary_block(
                sheet,
                current_row,
                commercial=commercial,
            )
            current_row = max(current_row + 1, 23)
            current_row = _write_ai_service_blocks(sheet, current_row, match_run)
            current_row += 1
            freeze_row = current_row
    elif configuration_groups:
        current_row = _write_ai_recommendations_header(sheet)
        current_row = _write_ai_service_blocks(sheet, current_row + 1, match_run)
        current_row += 1
        commercial = build_grouped_commercial_summary(
            report_json,
            configuration_groups,
            match_run_id=match_run.id,
        )
        if commercial is not None:
            current_row = _write_grouped_commercial_summary_block(
                sheet,
                current_row,
                commercial=commercial,
            )
            current_row += 1
            freeze_row = current_row
        current_row = _write_configuration_groups_block(
            sheet,
            current_row,
            configuration_groups=configuration_groups,
        )
    elif not recommendations:
        current_row = _write_ai_recommendations_header(sheet)
        current_row = _write_ai_service_blocks(sheet, current_row + 1, match_run)
        current_row += 1
        freeze_row = current_row
        message = _llm_recommendations_empty_message(match_run)
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row + 2,
            end_column=8,
        )
        cell = sheet.cell(row=current_row, column=1, value=message)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _style_merged_range(
            sheet,
            f"A{current_row}:H{current_row + 2}",
            border=FULL_BORDER,
        )
        sheet.row_dimensions[current_row].height = 70
        current_row += 4
    else:
        for index, recommendation in enumerate(recommendations, start=1):
            title = humanized_recommendation_title(
                recommendation,
                recommendations,
                index=index,
                default_title=f"Рекомендация {index}",
            )
            if index == 1:
                current_row = _write_ai_recommendations_header(sheet)
                current_row = _write_ai_service_blocks(sheet, current_row + 1, match_run)
                current_row += 1
                freeze_row = current_row
            current_row = _write_recommendation_card(
                sheet,
                current_row,
                index=index,
                recommendation=recommendation,
                normalized_requirements=normalized_requirements,
                title=title,
            )
            current_row += 1

    _write_ai_validation_summary_block(sheet, current_row + 1, match_run)

    _set_widths(
        sheet,
        {
            1: 18,
            2: 18,
            3: 18,
            4: 18,
            5: 18,
            6: 18,
            7: 18,
            8: 18,
        },
    )
    sheet.freeze_panes = f"A{freeze_row}"


def _write_ai_recommendations_header(sheet: Worksheet) -> int:
    sheet["A1"] = "AI-рекомендации по складскому подбору"
    sheet["A1"].font = Font(size=16, bold=True, color="1F1F1F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="EAF3F8")
    sheet.merge_cells("A1:H1")
    for column_index in range(1, 9):
        sheet.cell(row=1, column=column_index).border = FULL_BORDER
        sheet.cell(row=1, column=column_index).fill = PatternFill("solid", fgColor="EAF3F8")
    sheet.row_dimensions[1].height = 32
    return 2


def _write_ai_service_blocks(
    sheet: Worksheet,
    start_row: int,
    match_run: MatchRun,
) -> int:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    evidence_mode = _report_evidence_mode_text(report_json)
    evidence_sources = _report_evidence_sources_count(report_json)
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        "Служебные данные",
        fill=PatternFill("solid", fgColor="F3F4F6"),
        font=Font(bold=True, color="666666"),
    )
    passport_cells = [
        (1, 2, "Номер подбора", True),
        (3, 4, match_run.id, False),
        (5, 6, "Дата/время", True),
        (7, 8, _format_datetime(match_run.created_at), False),
    ]
    _write_service_cells(sheet, current_row, passport_cells)
    sheet.row_dimensions[current_row].height = 24
    current_row += 1
    service_cells = [
        (1, 2, "Режим проверки", True),
        (3, 4, evidence_mode, False),
        (5, 6, "Источники", True),
        (7, 8, evidence_sources, False),
    ]
    _write_service_cells(sheet, current_row, service_cells)
    sheet.row_dimensions[current_row].height = 24
    current_row += 1
    current_row = _write_service_text_row(
        sheet,
        current_row,
        f"Исходный запрос:\n{match_run.source_text or ''}",
        height=62,
    )
    current_row = _write_service_text_row(
        sheet,
        current_row,
        f"Нормализованные требования:\n{_normalized_requirements_text(match_run)}",
        height=82,
    )
    current_row = _write_service_text_row(
        sheet,
        current_row,
        f"Важно: {LLM_ENGINEERING_REVIEW_NOTE}",
        height=46,
        font=Font(italic=True, color="666666"),
    )
    return current_row


def _write_service_cells(
    sheet: Worksheet,
    row_index: int,
    cells: list[tuple[int, int, Any, bool]],
) -> None:
    for start_column, end_column, value, is_label in cells:
        sheet.merge_cells(
            start_row=row_index,
            start_column=start_column,
            end_row=row_index,
            end_column=end_column,
        )
        cell = sheet.cell(row=row_index, column=start_column, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if is_label:
            cell.font = Font(bold=True, color="666666")
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
    for column_index in range(1, 9):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.border = FULL_BORDER
        if cell.fill.fill_type is None:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")


def _write_service_text_row(
    sheet: Worksheet,
    row_index: int,
    text: str,
    *,
    height: int,
    font: Font | None = None,
) -> int:
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
    cell = sheet.cell(row=row_index, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = font or Font(color="555555")
    _style_merged_range(sheet, f"A{row_index}:H{row_index}", border=FULL_BORDER)
    sheet.row_dimensions[row_index].height = height
    return row_index + 1


def _write_ai_validation_summary_block(
    sheet: Worksheet,
    start_row: int,
    match_run: MatchRun,
) -> None:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    proposal_count = _int_value(report_json.get("llm_proposals_count"))
    valid_count = _int_value(report_json.get("valid_proposals_count"))
    shown_count = _int_value(report_json.get("ai_recommendations_count"))
    if shown_count is None:
        shown_count = len(_ai_recommendations(match_run))
    if proposal_count is None:
        rejected_count = _int_value(report_json.get("rejected_ai_recommendations_count")) or 0
        proposal_count = shown_count + rejected_count
    if valid_count is None:
        validation_summary = report_json.get("ai_validation_summary")
        if isinstance(validation_summary, Mapping):
            valid_count = _int_value(validation_summary.get("accepted_after_validation"))
    if valid_count is None:
        valid_count = shown_count
    if proposal_count <= 0 and shown_count <= 0:
        return

    lines = [
        f"AI проверил {proposal_count} {_excel_variant_plural(proposal_count)}, "
        f"к показу выбрано {shown_count}."
    ]
    if report_json.get("llm_repair_success") is True or report_json.get("llm_repair_used") is True:
        lines.append(
            "AI перепроверил цены по матрице и выбрал более дешевую эквивалентную "
            "RAM/SSD, если она была доступна."
        )
    if valid_count > 0 and shown_count < proposal_count:
        lines.append(
            "Часть вариантов была скрыта как дубли или уступающие по цене/рискам."
        )
    lines.extend(_excel_rejected_summary_lines(report_json))

    sheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row + 1,
        end_column=8,
    )
    cell = sheet.cell(row=start_row, column=1, value="\n".join(lines))
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    _style_merged_range(sheet, f"A{start_row}:H{start_row + 1}", border=FULL_BORDER)
    sheet.row_dimensions[start_row].height = 38
    sheet.row_dimensions[start_row + 1].height = 22


def _excel_rejected_summary_lines(report_json: Mapping[str, Any]) -> list[str]:
    validation_summary = report_json.get("ai_validation_summary")
    if not isinstance(validation_summary, Mapping):
        return []
    compatibility = (
        (_int_value(validation_summary.get("rejected_fatal")) or 0)
        + (_int_value(validation_summary.get("rejected_role_mismatch")) or 0)
    )
    missing_required = _int_value(
        validation_summary.get("rejected_missing_required_role")
    )
    if missing_required is None:
        missing_required = _int_value(validation_summary.get("rejected_missing_required")) or 0
    stock_shortage = _int_value(validation_summary.get("rejected_stock_shortage"))
    if stock_shortage is None:
        stock_shortage = _int_value(validation_summary.get("rejected_stock")) or 0
    incomplete = missing_required + stock_shortage
    selection_duplicate = _selection_summary_value(
        validation_summary,
        "selection_skipped_duplicate",
        "rejected_duplicate",
    )
    selection_worse = (
        _selection_summary_value(
            validation_summary,
            "selection_skipped_dominated_by_cheaper_equivalent",
            "rejected_right_size",
        )
        + (_int_value(validation_summary.get("selection_skipped_worse_by_price")) or 0)
        + (
            _int_value(
                validation_summary.get(
                    "selection_skipped_same_platform_without_meaningful_difference"
                )
            )
            or 0
        )
        + (
            _int_value(validation_summary.get("selection_skipped_lower_ranked_alternative"))
            or 0
        )
    )
    validation_parts: list[str] = []
    if compatibility:
        validation_parts.append(f"{compatibility} из-за совместимости")
    if incomplete:
        validation_parts.append(f"{incomplete} из-за неполной комплектации")

    lines: list[str] = []
    if validation_parts:
        lines.append("Валидатор отклонил: " + ", ".join(validation_parts) + ".")

    selection_parts: list[str] = []
    if selection_duplicate:
        selection_parts.append(f"{selection_duplicate} как дубли")
    if selection_worse:
        selection_parts.append(f"{selection_worse} как уступающие по цене/рискам")
    if selection_parts:
        lines.append("Скрыто при выборе: " + ", ".join(selection_parts) + ".")
    top_reasons = [
        str(item.get("reason") or "").strip()
        for item in _component_rows(report_json.get("rejected_reasons_top"))
        if str(item.get("reason") or "").strip()
    ]
    if top_reasons:
        lines.append("Основные причины валидатора: " + ", ".join(top_reasons[:5]) + ".")
    return lines


def _selection_summary_value(
    validation_summary: Mapping[str, Any],
    key: str,
    legacy_key: str,
) -> int:
    value = _int_value(validation_summary.get(key))
    if value is not None:
        return value
    return _int_value(validation_summary.get(legacy_key)) or 0


def _excel_variant_plural(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return "вариант"
    if count % 10 in {2, 3, 4} and count % 100 not in {12, 13, 14}:
        return "варианта"
    return "вариантов"


def _excel_shown_recommendations_text(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return f"показана {count} безопасная рекомендация"
    if count % 10 in {2, 3, 4} and count % 100 not in {12, 13, 14}:
        return f"показаны {count} безопасные рекомендации"
    return f"показано {count} безопасных рекомендаций"


def _style_merged_range(
    sheet: Worksheet,
    range_name: str,
    *,
    border: Border,
    fill: PatternFill | None = None,
) -> None:
    for row in sheet[range_name]:
        for cell in row:
            cell.border = border
            if fill is not None:
                cell.fill = fill


def _write_recommendation_card(
    sheet: Worksheet,
    start_row: int,
    *,
    index: int,
    recommendation: Mapping[str, Any],
    normalized_requirements: Mapping[str, Any],
    title: str,
) -> int:
    title = _safe_user_text(title or recommendation.get("title") or f"Рекомендация {index}")
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    header = sheet.cell(row=start_row, column=1, value=f"{index}. {title}")
    header.fill = SUBHEADER_FILL
    header.font = Font(bold=True, size=12, color="1F1F1F")
    header.alignment = Alignment(wrap_text=True, vertical="center")
    for column_index in range(1, 9):
        cell = sheet.cell(row=start_row, column=column_index)
        cell.fill = SUBHEADER_FILL
        cell.border = FULL_BORDER
    sheet.row_dimensions[start_row].height = 28

    rows = [
        ("Тип", _recommendation_type_text(recommendation)),
        ("Платформа / позиция", _recommendation_display_text(recommendation)),
        ("Состав на 1 сервер", _composition_per_server_text(recommendation)),
        (
            "Всего к подбору / остатки",
            "\n".join(
                value
                for value in [
                    _total_quantities_text(recommendation),
                    _component_stocks_text(recommendation),
                ]
                if value
            ),
        ),
        (
            "Сумма",
            _recommendation_amount_text(recommendation),
        ),
        (
            "Подбор / right-size",
            _safe_user_text(recommendation.get("right_size_note") or ""),
        ),
        (
            "Уверенность",
            _confidence_display_text(recommendation),
        ),
        (
            "Почему выбрано",
            _why_selected_excel_text(recommendation),
        ),
        (
            "Что проверить",
            _checks_text(recommendation, normalized_requirements=normalized_requirements),
        ),
    ]
    optional_text = _optional_components_text(recommendation)
    if optional_text:
        rows.insert(
            4,
            ("Опционально / проверить инженеру", optional_text),
        )
    evidence_text = _recommendation_evidence_text(recommendation)
    if evidence_text:
        rows.append(("Доказательная проверка", evidence_text))
    rows.append(
        (
            "Статус",
            "\n".join(
                value
                for value in [
                    _recommendation_status_text(recommendation),
                    _missing_text(recommendation),
                    _work_action_text(recommendation),
                ]
                if value
            ),
        )
    )
    row_index = start_row + 1
    for label, value in rows:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
        sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=8)
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=3, value=value)
        label_cell.font = BOLD_FONT
        label_cell.fill = PatternFill("solid", fgColor="F3F7FA")
        for column_index in range(1, 9):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = _card_row_height(label, value)
        row_index += 1
    return row_index


def _write_grouped_commercial_summary_block(
    sheet: Worksheet,
    start_row: int,
    *,
    commercial: Any,
) -> int:
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        "Предварительная спецификация для КП",
    )
    for label, value in grouped_commercial_excel_rows(commercial):
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=8)
        label_cell = sheet.cell(row=current_row, column=1, value=label)
        value_cell = sheet.cell(row=current_row, column=3, value=value)
        label_cell.font = BOLD_FONT
        label_cell.fill = PatternFill("solid", fgColor="F3F7FA")
        for column_index in range(1, 9):
            cell = sheet.cell(row=current_row, column=column_index)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[current_row].height = _commercial_summary_row_height(value)
        current_row += 1
    return current_row


def _write_primary_commercial_summary_block(
    sheet: Worksheet,
    start_row: int,
    *,
    commercial: Mapping[str, Any],
) -> int:
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        str(commercial.get("title") or "Предварительная спецификация для КП"),
        fill=PatternFill("solid", fgColor="EAF3F8"),
        font=Font(size=16, bold=True, color="1F1F1F"),
    )
    current_row = _write_primary_summary_lines(sheet, current_row, commercial)
    current_row += 1
    current_row = _write_primary_bom_table(sheet, current_row, commercial)
    current_row += 1
    current_row = _write_primary_comment_block(sheet, current_row, commercial)
    current_row += 1
    current_row = _write_primary_checks_block(sheet, current_row, commercial)
    return current_row


def _write_primary_summary_lines(
    sheet: Worksheet,
    start_row: int,
    commercial: Mapping[str, Any],
) -> int:
    server_line = str(commercial.get("server_line") or "").strip()
    if not server_line:
        server_quantity = _int_value(commercial.get("server_quantity"))
        product_group = str(commercial.get("product_group") or "").strip()
        if product_group == NETWORK_PRODUCT_GROUP:
            label = "Сетевое оборудование"
        elif product_group == STORAGE_PRODUCT_GROUP:
            label = "СХД"
        else:
            label = "Сервер в сборе"
        server_line = f"{label} - {_quantity_units(server_quantity)}"
    price_line = str(commercial.get("price_line") or "").strip()
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    sheet.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=8)
    server_cell = sheet.cell(row=start_row, column=1, value=server_line)
    price_cell = sheet.cell(row=start_row, column=5, value=price_line)
    for cell in (server_cell, price_cell):
        cell.font = Font(size=12, bold=True, color="1F1F1F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for column_index in range(1, 9):
        cell = sheet.cell(row=start_row, column=column_index)
        cell.border = FULL_BORDER
        cell.fill = PatternFill("solid", fgColor="F8FBFD")
    sheet.row_dimensions[start_row].height = 30
    return start_row + 1


def _write_primary_bom_table(
    sheet: Worksheet,
    start_row: int,
    commercial: Mapping[str, Any],
) -> int:
    header_row = start_row
    _merge_table_row(sheet, header_row)
    product_group = str(commercial.get("product_group") or "").strip()
    headers = {
        1: "Позиция",
        4: "Количество" if product_group in {"network", "storage"} else "На 1 сервер",
        6: "Всего к заказу",
        7: "Остаток",
        8: "Примечание",
    }
    for column_index, value in headers.items():
        cell = sheet.cell(row=header_row, column=column_index, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for column_index in range(1, 9):
        cell = sheet.cell(row=header_row, column=column_index)
        cell.fill = HEADER_FILL
        cell.border = FULL_BORDER
    sheet.row_dimensions[header_row].height = 26

    current_row = header_row + 1
    rows = _commercial_bom_rows(commercial)
    for row in rows:
        _merge_table_row(sheet, current_row)
        values = {
            1: _bom_position_text(row),
            4: row.get("per_server") or "",
            6: row.get("total") or "",
            7: row.get("stock") or "",
            8: row.get("note") or "",
        }
        for column_index, value in values.items():
            cell = sheet.cell(row=current_row, column=column_index, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index in range(1, 9):
            cell = sheet.cell(row=current_row, column=column_index)
            cell.border = FULL_BORDER
        sheet.row_dimensions[current_row].height = 36
        current_row += 1
    return current_row


def _merge_table_row(sheet: Worksheet, row_index: int) -> None:
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
    sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index, end_column=5)


def _commercial_bom_rows(commercial: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    rows = commercial.get("bom_rows")
    if isinstance(rows, list):
        return [row for row in rows if isinstance(row, Mapping)]
    return []


def _bom_position_text(row: Mapping[str, Any]) -> str:
    label = str(row.get("label") or "").strip()
    name = str(row.get("name") or "").strip()
    if label and name:
        return f"{label}: {name}"
    return label or name


def _write_primary_comment_block(
    sheet: Worksheet,
    start_row: int,
    commercial: Mapping[str, Any],
) -> int:
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        "Комментарий",
        fill=PatternFill("solid", fgColor="F3F7FA"),
        font=Font(bold=True, color="1F1F1F"),
    )
    lines = [
        str(line).strip()
        for line in commercial.get("comment_lines", [])
        if str(line or "").strip()
    ]
    if not lines:
        rows = primary_commercial_excel_rows(commercial)
        lines = [value for label, value in rows if label == "Комментарий"]
    return _write_wide_text_row(sheet, current_row, "\n".join(lines), height=46)


def _write_primary_checks_block(
    sheet: Worksheet,
    start_row: int,
    commercial: Mapping[str, Any],
) -> int:
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        "Проверить перед КП",
        fill=PatternFill("solid", fgColor="F3F7FA"),
        font=Font(bold=True, color="1F1F1F"),
    )
    checks = [
        str(check).strip()
        for check in commercial.get("engineer_checks", [])
        if str(check or "").strip()
    ]
    text = "\n".join(f"- {check}" for check in checks)
    return _write_wide_text_row(sheet, current_row, text, height=96)


def _commercial_summary_row_height(value: Any) -> int:
    text = str(value or "")
    line_count = max(1, text.count("\n") + 1)
    return min(118, max(26, line_count * 18))


def _write_configuration_groups_block(
    sheet: Worksheet,
    start_row: int,
    *,
    configuration_groups: list[Mapping[str, Any]],
) -> int:
    current_row = start_row
    current_row = _write_wide_section_title(
        sheet,
        current_row,
        "Конфигурационные семейства",
    )
    group_count = len(configuration_groups)
    option_count = sum(
        len(_component_rows(group.get("platform_options")))
        for group in configuration_groups
    )
    summary = (
        f"Найдено конфигурационных семейств: {group_count}. "
        f"Вариантов платформ: {option_count}. "
        "Инженерная подтвержденность: предварительно, перед КП нужна проверка."
    )
    current_row = _write_wide_text_row(sheet, current_row, summary, height=42)

    for group_index, group in enumerate(configuration_groups, start=1):
        current_row += 1
        title = _safe_user_text(group.get("group_title") or f"Семейство {group_index}")
        current_row = _write_wide_section_title(
            sheet,
            current_row,
            f"Конфигурационная база {group_index}: {title}",
            fill=SUBHEADER_FILL,
            font=Font(bold=True, size=12, color="1F1F1F"),
        )
        architecture = _safe_user_text(group.get("architecture_summary") or "")
        why_group = _safe_user_text(group.get("why_group_matters") or "")
        if architecture or why_group:
            current_row = _write_wide_text_row(
                sheet,
                current_row,
                "\n".join(value for value in [architecture, why_group] if value),
                height=48,
            )

        current_row = _write_group_component_base(sheet, current_row, group)
        current_row = _write_group_platform_options(sheet, current_row, group)
        current_row = _write_group_optional_components(sheet, current_row, group)

    current_row += 1
    current_row = _write_wide_section_title(sheet, current_row, "Что проверить инженеру")
    current_row = _write_wide_text_row(
        sheet,
        current_row,
        "\n".join(f"- {check}" for check in _grouped_excel_checks(configuration_groups)),
        height=118,
    )
    return current_row


def _write_group_component_base(
    sheet: Worksheet,
    start_row: int,
    group: Mapping[str, Any],
) -> int:
    current_row = _write_wide_section_title(
        sheet,
        start_row,
        "Компонентная база",
        fill=PatternFill("solid", fgColor="F3F7FA"),
        font=BOLD_FONT,
    )
    headers = [
        "Роль",
        "Компонент",
        "На 1 сервер",
        "Всего к подбору",
        "Склад",
        "Цена",
        "Валюта",
        "Примечание",
    ]
    _write_header_at(sheet, current_row, headers)
    current_row += 1
    base = group.get("component_base")
    rows = _group_component_base_rows(base if isinstance(base, Mapping) else {})
    if not rows:
        _write_row(sheet, current_row, ["Компоненты", "не указаны", "", "", "", "", "", ""])
        return current_row + 1
    server_quantity = _group_server_quantity(group)
    for label, component in rows:
        per_server = _bom_per_server_line(component, server_quantity)
        total = _group_component_total_text(component)
        stock = _stock_count_text(component.get("available_quantity"))
        price = _money_value(component.get("price_value"))
        _write_row(
            sheet,
            current_row,
            [
                label,
                commercial_component_name(component),
                per_server,
                total,
                stock,
                price,
                component.get("price_currency") or "",
                "\n".join(_component_note_lines(component)),
            ],
        )
        sheet.row_dimensions[current_row].height = 58
        current_row += 1
    _set_number_format_from_row(sheet, [6], start_row=start_row)
    return current_row


def _write_group_platform_options(
    sheet: Worksheet,
    start_row: int,
    group: Mapping[str, Any],
) -> int:
    current_row = start_row + 1
    current_row = _write_wide_section_title(
        sheet,
        current_row,
        "Варианты платформ",
        fill=PatternFill("solid", fgColor="F3F7FA"),
        font=BOLD_FONT,
    )
    headers = [
        "Роль",
        "Платформа",
        "Склад",
        "Сумма за весь запрос",
        "Валюта",
        "Почему",
        "Компромисс",
        "Инженерный статус",
    ]
    _write_header_at(sheet, current_row, headers)
    current_row += 1
    options = _component_rows(group.get("platform_options"))
    if not options:
        _write_row(sheet, current_row, ["", "варианты платформ не указаны", "", "", "", "", "", ""])
        return current_row + 1
    for option in options:
        platform = option.get("platform") if isinstance(option.get("platform"), Mapping) else {}
        amount = _money_value(option.get("total_price_value"))
        tradeoff_text = "\n".join(
            commercial_safe_russian_text(value)
            for value in as_string_list(option.get("tradeoffs"))
            if commercial_safe_russian_text(value)
        )
        _write_row(
            sheet,
            current_row,
            [
                _platform_option_role_label(option),
                commercial_component_name(platform) if isinstance(platform, Mapping) else "",
                _safe_user_text(option.get("stock_status") or ""),
                amount,
                option.get("total_price_currency") or "",
                commercial_reason_for_option(option),
                tradeoff_text,
                commercial_engineering_confidence(option.get("engineering_confidence")),
            ],
        )
        sheet.row_dimensions[current_row].height = 66
        current_row += 1
    _set_number_format_from_row(sheet, [4], start_row=start_row)
    return current_row


def _write_group_optional_components(
    sheet: Worksheet,
    start_row: int,
    group: Mapping[str, Any],
) -> int:
    option_rows = _component_rows(group.get("platform_options"))
    optional_lines: list[str] = []
    for option in option_rows:
        platform = option.get("platform")
        platform_text = _component_article(platform) if isinstance(platform, Mapping) else ""
        for component in _component_rows(option.get("optional_components")):
            optional_lines.append(
                " - ".join(
                    part
                    for part in [
                        platform_text,
                        _bom_role_label(str(component.get("role") or "")),
                        _component_article(component),
                        _quantity_units(_int_value(component.get("quantity_required"))),
                        f"склад: {_stock_count_text(component.get('available_quantity'))}",
                    ]
                    if part
                )
            )
    if not optional_lines:
        return start_row
    current_row = start_row + 1
    current_row = _write_wide_section_title(
        sheet,
        current_row,
        "Опциональные компоненты отдельно",
        fill=PatternFill("solid", fgColor="F3F7FA"),
        font=BOLD_FONT,
    )
    return _write_wide_text_row(sheet, current_row, "\n".join(optional_lines), height=70)


def _write_wide_section_title(
    sheet: Worksheet,
    row_index: int,
    title: str,
    *,
    fill: PatternFill = HEADER_FILL,
    font: Font = HEADER_FONT,
) -> int:
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
    cell = sheet.cell(row=row_index, column=1, value=title)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    for column_index in range(1, 9):
        sheet.cell(row=row_index, column=column_index).fill = fill
        sheet.cell(row=row_index, column=column_index).border = FULL_BORDER
    sheet.row_dimensions[row_index].height = 28
    return row_index + 1


def _write_wide_text_row(
    sheet: Worksheet,
    row_index: int,
    text: str,
    *,
    height: int,
) -> int:
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
    cell = sheet.cell(row=row_index, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    _style_merged_range(sheet, f"A{row_index}:H{row_index}", border=FULL_BORDER)
    sheet.row_dimensions[row_index].height = height
    return row_index + 1


def _recommendation_amount_text(recommendation: Mapping[str, Any]) -> str:
    value = _money_value(recommendation.get("total_price_value"))
    currency = str(recommendation.get("total_price_currency") or "").strip()
    note = str(
        recommendation.get("total_price_note") or recommendation.get("price_note") or ""
    ).strip()
    if value is None:
        amount = "сумма не рассчитана"
    else:
        amount = f"{value:,}".replace(",", " ") if isinstance(value, int) else str(value)
    return " ".join(part for part in [amount, currency, note] if part)


def _confidence_display_text(recommendation: Mapping[str, Any]) -> str:
    explicit = _safe_user_text(recommendation.get("displayed_confidence") or "")
    if explicit:
        return explicit
    commercial = str(
        recommendation.get("commercial_fit_confidence")
        or recommendation.get("confidence")
        or ""
    ).strip()
    commercial_label = {
        "high": "высокое",
        "medium": "среднее",
        "low": "низкое",
    }.get(commercial, "")
    if not commercial_label:
        return ""
    engineering_label = "предварительно, требуется проверка"
    summary = recommendation.get("evidence_summary")
    if isinstance(summary, Mapping):
        source_count = _int_value(summary.get("sources_count")) or 0
        status = str(summary.get("status") or "").strip()
        missing = [
            *as_string_list(summary.get("missing")),
            *as_string_list(summary.get("not_confirmed")),
        ]
        fatal = as_string_list(summary.get("fatal_concerns"))
        if status == "mismatch" or fatal:
            engineering_label = "не подтверждено, требуется проверка"
        elif source_count > 0 and (status == "partially_confirmed" or missing):
            engineering_label = "частично проверено источниками, требуется проверка"
        elif source_count > 0 and status == "confirmed":
            engineering_label = "проверено по источникам, требуется финальная проверка"
    return (
        f"Коммерческое соответствие: {commercial_label}; "
        f"инженерная подтвержденность: {engineering_label}."
    )


def _card_row_height(label: str, value: Any) -> int:
    text = str(value or "")
    line_count = max(1, text.count("\n") + 1)
    if label in {"Состав на 1 сервер", "Всего к подбору / остатки"}:
        return min(110, max(70, line_count * 20))
    if label in {
        "Подбор / right-size",
        "Почему выбрано",
        "Что проверить",
        "Доказательная проверка",
    }:
        return min(100, max(55, line_count * 18))
    if label == "Статус" and line_count > 1:
        return min(80, max(55, line_count * 18))
    return min(54, max(26, line_count * 18))


def _fill_component_matrix_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    sheet.sheet_view.showGridLines = False
    product_group = _report_product_group(match_run)
    table_start_row = _write_matrix_coverage_block(sheet, match_run)
    headers = [
        "Роль",
        "Производитель",
        "Артикул",
        "Наименование",
        "Остаток",
        "Цена",
        "Валюта",
        "Извлеченные признаки",
        "Оценка соответствия",
        "Почему подходит",
        "Превышение требования",
        "Что проверить",
        "Доказательства",
        "Уверенность источников",
        "Источники",
        "Что подтверждено",
        "Что не подтверждено",
    ]
    _write_header_at(sheet, table_start_row, headers)

    rows = _unified_component_matrix_rows(match_run)
    normalized_requirements = _first_report_requirements(match_run)
    if not rows:
        sheet.cell(row=table_start_row + 1, column=1, value="Матрица компонентов не сформирована.")
        sheet.merge_cells(
            start_row=table_start_row + 1,
            start_column=1,
            end_row=table_start_row + 1,
            end_column=len(headers),
        )
    else:
        for row_index, row in enumerate(rows, start=table_start_row + 1):
            row_values = [
                _matrix_row_role_label(row, product_group),
                row.get("producer") or "",
                row.get("part_number") or "",
                row.get("name") or row.get("item_name") or "",
                row.get("available_quantity"),
                _money_value(row.get("price_value")),
                row.get("price_currency") or "",
                _facts_text(row.get("extracted_facts") or row.get("facts")),
                _matrix_fit_text(row),
                _safe_user_text(row.get("fit_reason") or ""),
                _safe_user_text(row.get("over_requirement") or ""),
                _matrix_checks_text(
                    row,
                    normalized_requirements=normalized_requirements,
                ),
                _matrix_evidence_status_text(row),
                _matrix_evidence_confidence_text(row),
                _matrix_evidence_sources_text(row),
                _matrix_evidence_confirmed_text(row),
                _matrix_evidence_missing_text(row),
            ]
            _write_row(sheet, row_index, row_values)
            sheet.row_dimensions[row_index].height = 60

    _set_number_format(sheet, [6])
    _set_widths(
        sheet,
        {
            1: 22,
            2: 28,
            3: 24,
            4: 55,
            5: 12,
            6: 16,
            7: 10,
            8: 45,
            9: 22,
            10: 45,
            11: 24,
            12: 50,
            13: 24,
            14: 22,
            15: 16,
            16: 42,
            17: 42,
        },
    )
    sheet.freeze_panes = f"A{table_start_row + 1}"
    sheet.auto_filter.ref = f"A{table_start_row}:Q{max(sheet.max_row, table_start_row + 1)}"
    sheet.row_dimensions[table_start_row].height = 36


def _write_matrix_coverage_block(sheet: Worksheet, match_run: MatchRun) -> int:
    coverage = _matrix_coverage_summary(match_run)
    product_group = _report_product_group(match_run)
    sheet["A1"] = "Охват матрицы"
    sheet["A1"].font = Font(size=14, bold=True, color="1F1F1F")
    sheet.merge_cells("A1:H1")

    headers = [
        "Роль",
        "В БД",
        "Допущено",
        "Передано в LLM",
        "Скрыто лимитом",
        "Лимит",
        "Buckets",
        "Стратегия",
    ]
    _write_header_at(sheet, 2, headers)

    roles = _matrix_coverage_roles(match_run, coverage, product_group)
    total_by_role = _coverage_mapping(coverage, "total_products_by_role")
    eligible_by_role = _coverage_mapping(coverage, "eligible_products_by_role")
    sent_by_role = _coverage_mapping(coverage, "sent_to_llm_by_role")
    omitted_by_role = _coverage_mapping(coverage, "omitted_by_role")
    bucket_by_role = _coverage_mapping(coverage, "bucket_summary_by_role")
    limit = coverage.get("limit_per_role") if isinstance(coverage, Mapping) else ""
    strategy = (
        str(coverage.get("selection_strategy") or "")
        if isinstance(coverage, Mapping)
        else ""
    )

    row_index = 3
    has_omitted = False
    if not roles:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
        cell = sheet.cell(
            row=row_index,
            column=1,
            value="Для этой группы продукта релевантные роли матрицы не найдены.",
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index in range(1, 9):
            sheet.cell(row=row_index, column=column_index).border = THIN_BORDER
        sheet.row_dimensions[row_index].height = 32
        row_index += 1

    for role in roles:
        omitted = _int_value(omitted_by_role.get(role)) or 0
        has_omitted = has_omitted or omitted > 0
        bucket_summary = bucket_by_role.get(role)
        sheet_values = [
            _coverage_role_label(role, product_group),
            total_by_role.get(role, 0),
            eligible_by_role.get(role, 0),
            sent_by_role.get(role, 0),
            omitted,
            limit,
            _bucket_summary_text(bucket_summary),
            strategy,
        ]
        _write_row(sheet, row_index, sheet_values)
        sheet.row_dimensions[row_index].height = 32
        row_index += 1

    if has_omitted:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
        cell = sheet.cell(
            row=row_index,
            column=1,
            value=(
                "Часть компонентов скрыта лимитом матрицы; при спорном подборе "
                "увеличьте LLM_COMPONENT_CANDIDATES_PER_ROLE."
            ),
        )
        cell.font = Font(italic=True, color="666666")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index in range(1, 9):
            sheet.cell(row=row_index, column=column_index).border = THIN_BORDER
        sheet.row_dimensions[row_index].height = 34
        row_index += 1

    row_index += 1
    return row_index


def _matrix_coverage_summary(match_run: MatchRun) -> Mapping[str, Any]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    coverage = report_json.get("component_matrix_coverage_summary")
    if isinstance(coverage, Mapping):
        return coverage
    matrix = report_json.get("component_candidate_matrix")
    if isinstance(matrix, Mapping) and isinstance(
        matrix.get("component_matrix_coverage_summary"),
        Mapping,
    ):
        return matrix["component_matrix_coverage_summary"]
    return {}


def _coverage_mapping(coverage: Mapping[str, Any], key: str) -> Mapping[str, Any]:
    value = coverage.get(key)
    return value if isinstance(value, Mapping) else {}


def _matrix_coverage_roles(
    match_run: MatchRun,
    coverage: Mapping[str, Any],
    product_group: str,
) -> list[str]:
    if product_group == SERVER_PRODUCT_GROUP:
        return list(SERVER_MATRIX_COVERAGE_ROLES)

    role_order = _matrix_role_order(product_group)
    allowed_roles = set(role_order)
    active_roles: set[str] = set()
    total_by_role = _coverage_mapping(coverage, "total_products_by_role")
    eligible_by_role = _coverage_mapping(coverage, "eligible_products_by_role")
    sent_by_role = _coverage_mapping(coverage, "sent_to_llm_by_role")
    omitted_by_role = _coverage_mapping(coverage, "omitted_by_role")
    bucket_by_role = _coverage_mapping(coverage, "bucket_summary_by_role")

    for role in allowed_roles:
        if (
            _int_value(total_by_role.get(role))
            or _int_value(eligible_by_role.get(role))
            or _int_value(sent_by_role.get(role))
            or _int_value(omitted_by_role.get(role))
            or bool(bucket_by_role.get(role))
        ):
            active_roles.add(role)

    for role in _report_required_roles(match_run):
        if role in allowed_roles:
            active_roles.add(role)

    for row in _component_candidate_rows(match_run, product_group=product_group):
        role = _matrix_candidate_role(row)
        if role in allowed_roles:
            active_roles.add(role)

    return [role for role in role_order if role in active_roles]


def _coverage_role_label(role: str, product_group: str = SERVER_PRODUCT_GROUP) -> str:
    network_labels = {
        "switch": "Коммутаторы",
        "router": "Маршрутизаторы",
        "firewall": "Межсетевые экраны",
        "access_point": "Точки доступа",
        "transceiver": "Трансиверы",
        "dac_cable": "Кабели/DAC",
        "cable": "Кабели/DAC",
        "license": "Поддержка/лицензии",
        "support": "Поддержка/лицензии",
        "power_supply": "Блоки питания",
        "stacking_module": "Модули стекирования",
        "other_accessory": "Аксессуары",
    }
    storage_labels = {
        "storage_system": "СХД",
        "controller": "Контроллеры СХД",
        "controller_module": "Контроллерные модули",
        "disk_shelf": "Дисковые полки",
        "drive": "Диски",
        "ssd": "SSD",
        "hdd": "HDD",
        "cache": "Кэш",
        "host_port": "Host ports",
        "protocol_module": "Протокольные модули",
        "transceiver": "Трансиверы",
        "cable": "Кабели",
        "license": "Лицензии",
        "support": "Поддержка",
        "power_supply": "Блоки питания",
        "rail_kit": "Рельсы",
        "other_accessory": "Аксессуары",
    }
    if product_group == NETWORK_PRODUCT_GROUP:
        return network_labels.get(role, role)
    if product_group == STORAGE_PRODUCT_GROUP:
        return storage_labels.get(role, role)
    labels = {
        "ready_server": "Готовые серверы",
        "platform": "Платформы",
        "switch": "Коммутаторы",
        "router": "Маршрутизаторы",
        "firewall": "Межсетевые экраны",
        "access_point": "Точки доступа",
        "cpu": "CPU",
        "ram": "RAM",
        "ssd": "SSD",
        "hdd": "HDD",
        "storage_controller": "Контроллеры",
        "network_adapter": "Сетевые адаптеры",
        "transceiver": "Трансиверы",
        "dac_cable": "DAC-кабели",
        "cable": "Кабели",
        "license": "Лицензии",
        "support": "Поддержка",
        "power_supply": "Блоки питания",
        "stacking_module": "Модули стекирования",
    }
    return labels.get(role, role)


def _bucket_summary_text(value: Any) -> str:
    if not isinstance(value, Mapping):
        return ""
    return "\n".join(f"{_matrix_bucket_label(str(key))}: {item}" for key, item in value.items())


def _matrix_bucket_label(value: str) -> str:
    labels = {
        "closest_to_min_cores": "Ближе к требованию по ядрам",
        "unknown_cores_good_stock": "CPU с неуказанными ядрами и хорошим остатком",
        "cpu_16_cores": "CPU 16 ядер",
        "cpu_20_cores": "CPU 20 ядер",
        "cpu_24_cores": "CPU 24 ядра",
        "cpu_32_cores": "CPU 32 ядра",
        "cpu_48_cores": "CPU 48 ядер",
        "cpu_64_cores": "CPU 64 ядра",
        "exact_or_close_fit": "Минимально подходит",
        "acceptable_overfit": "Выше требования, допустимо",
        "excessive_overfit": "Выше требования, нужна проверка альтернатив",
        "unknown_fit": "Нужна ручная проверка",
    }
    if value in labels:
        return labels[value]
    text = value.replace("_", " ")
    text = text.replace("cores", "ядра").replace("overfit", "выше требования")
    text = text.replace("fit label", "оценка соответствия")
    return text[:1].upper() + text[1:]


def _fill_summary_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    sheet["A1"] = "Сводка подбора"
    sheet["A1"].font = Font(size=14, bold=True)

    rows = [
        ("Номер подбора", match_run.id),
        ("Дата и время", _format_datetime(match_run.created_at)),
        ("Исходный запрос", match_run.source_text or ""),
        ("Итог", human_match_status(match_run.status)),
        ("Найдено вариантов", match_run.total_candidates),
        ("Полностью подходят", match_run.matched_items),
        ("Нужна проверка инженера", yes_no(match_run.engineer_review_required)),
        (
            "Краткий вывод",
            short_conclusion(
                status=match_run.status,
                engineer_review_required=match_run.engineer_review_required,
            ),
        ),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).font = BOLD_FONT
        sheet.cell(row=row_index, column=1).fill = SUBHEADER_FILL
        sheet.cell(row=row_index, column=1).border = THIN_BORDER
        sheet.cell(row=row_index, column=2).border = THIN_BORDER
        sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    sheet["A13"] = "Это предварительный подбор по складу, не финальная инженерная спецификация."
    sheet["A13"].font = Font(italic=True, color="666666")
    sheet.merge_cells("A13:B13")
    if _llm_recommended_builds(match_run):
        sheet["A14"] = LLM_ENGINEERING_REVIEW_NOTE
        sheet["A14"].font = Font(italic=True, color="666666")
        sheet["A14"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells("A14:B14")
    _set_widths(sheet, {1: 30, 2: 80})


def _fill_candidates_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    headers = [
        "№",
        "Дистрибьютор",
        "Производитель",
        "Артикул",
        "Наименование",
        "Категория",
        "Остаток, шт.",
        "Цена",
        "Валюта",
        "Оценка соответствия",
        "Итог по варианту",
        "Комментарий",
    ]
    _write_header(sheet, headers)

    candidates = _ready_candidates(match_run)
    if not candidates:
        sheet.cell(row=2, column=1, value="Складские варианты не найдены.")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sheet.cell(row=2, column=1).value = _llm_recommendations_empty_message(match_run)
    else:
        for row_index, candidate in enumerate(candidates, start=2):
            row_values = [
                row_index - 1,
                candidate.distributor_code,
                candidate.producer or "",
                candidate.part_number or "",
                candidate.item_name or "",
                candidate.category_id or "",
                candidate.available_quantity,
                _price_value(candidate),
                candidate.price_currency or "",
                f"{candidate.confidence_score} из 100",
                candidate_outcome(_candidate_mapping(candidate)),
                candidate_comment(_candidate_mapping(candidate)),
            ]
            _write_row(sheet, row_index, row_values)

    _set_number_format(sheet, [8])
    _set_widths(
        sheet,
        {
            1: 6,
            2: 16,
            3: 18,
            4: 24,
            5: 48,
            6: 16,
            7: 14,
            8: 16,
            9: 10,
            10: 22,
            11: 28,
            12: 60,
        },
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fill_builds_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    amount_header = _amount_header_for_request(match_run)
    headers = [
        "№",
        "Итог по сборке",
        "Комплектность",
        "Платформа",
        "CPU",
        "RAM",
        "Накопители",
        "Контроллеры",
        "Сетевые адаптеры",
        "Требуемое количество серверов",
        "Количество CPU на сервер",
        "Общее количество CPU",
        amount_header,
        "Валюта",
        "Что не входит в сумму",
        "Что проверить инженеру",
        "Score",
        "Причина ранжирования",
        "RAM всего к подбору",
    ]
    _write_header(sheet, headers)

    build_candidates = _build_candidates(match_run)
    if not build_candidates:
        sheet.cell(
            row=2,
            column=1,
            value=(
                "Сборка из комплектующих пока не предложена - нет достаточных складских "
                "данных по платформам/комплектующим."
            ),
        )
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    else:
        for row_index, candidate in enumerate(build_candidates, start=2):
            raw = _raw_mapping(candidate)
            components = _component_rows(raw.get("components"))
            checks = humanized_checks(
                risk_flags=[
                    *as_string_list(candidate.risk_flags_json),
                    *as_string_list(raw.get("compatibility_warnings")),
                ],
                missing_requirements=[
                    *as_string_list(candidate.missing_requirements_json),
                    *as_string_list(raw.get("missing_components")),
                ],
            )
            row_values = [
                row_index - 1,
                _build_result_text(raw),
                _build_completeness_text(raw),
                _components_text(components, "server_platform"),
                _cpu_components_text(candidate, components),
                _components_text(components, "ram"),
                _storage_components_text(components),
                _components_text(components, "storage_controller"),
                _components_text(components, "network_adapter"),
                raw.get("quantity_required") or _server_quantity_from_components(components),
                raw.get("cpu_per_server") or "",
                raw.get("total_cpu_required") or "",
                _money_value(raw.get("total_price_value") or candidate.price_value),
                raw.get("total_price_currency") or candidate.price_currency or "",
                _excluded_from_total_text(raw),
                "\n".join(checks),
                raw.get("score") or candidate.confidence_score,
                "\n".join(as_string_list(raw.get("rank_reason"))),
                _ram_order_text(components),
            ]
            _write_row(sheet, row_index, row_values)

    _set_number_format(sheet, [13])
    _set_widths(
        sheet,
        {
            1: 6,
            2: 24,
            3: 34,
            4: 42,
            5: 30,
            6: 38,
            7: 38,
            8: 34,
            9: 34,
            10: 24,
            11: 24,
            12: 20,
            13: 22,
            14: 10,
            15: 24,
            16: 70,
            17: 10,
            18: 70,
            19: 22,
        },
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fill_llm_recommendations_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    amount_header = _amount_header_for_request(match_run)
    headers = [
        "№",
        "Решение",
        "Платформа",
        "CPU",
        "RAM",
        "SSD/HDD",
        amount_header,
        "Валюта",
        "Почему выбрана",
        "Критичные риски",
        "Уверенность",
        "Нужна инженерная проверка",
        "Режим оптимизации",
        "Соответствие требованиям",
        "Избыточность CPU",
        "Избыточность накопителей",
        "Почему выбран компонент выше требования",
        "Что не входит в сумму",
    ]
    _write_header(sheet, headers)

    rows = _llm_recommended_builds(match_run)
    if not rows:
        sheet.cell(row=2, column=1, value="LLM-рекомендации не сформированы.")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    else:
        for row_index, candidate in enumerate(rows, start=2):
            components = _component_rows(candidate.get("components"))
            row_values = [
                row_index - 1,
                candidate.get("decision") or "",
                _components_text(components, "server_platform"),
                _components_text(components, "cpu"),
                _components_text(components, "ram"),
                _storage_components_text(components),
                _money_value(candidate.get("total_price_value")),
                candidate.get("total_price_currency") or "",
                candidate.get("why_selected") or "",
                "\n".join(as_string_list(candidate.get("critical_risks"))),
                _confidence_display_text(candidate),
                yes_no(candidate.get("engineer_review_required", True)),
                candidate.get("optimization_mode") or "",
                candidate.get("requirement_fit") or "",
                candidate.get("cpu_over_requirement") or "",
                candidate.get("storage_over_requirement") or "",
                candidate.get("overfit_reason") or "",
                _excluded_from_total_text(candidate),
            ]
            _write_row(sheet, row_index, row_values)

    _set_number_format(sheet, [7])
    _set_widths(
        sheet,
        {
            1: 6,
            2: 14,
            3: 42,
            4: 34,
            5: 38,
            6: 38,
            7: 22,
            8: 10,
            9: 60,
            10: 70,
            11: 16,
            12: 24,
            13: 22,
            14: 28,
            15: 18,
            16: 24,
            17: 60,
            18: 24,
        },
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fill_component_candidates_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    headers = [
        "№",
        "Роль",
        "Производитель",
        "Артикул",
        "Наименование",
        "Остаток",
        "Цена",
        "Валюта",
        "Извлеченные признаки",
        "Почему подходит",
        "Оценка соответствия",
        "Почему подходит",
        "Превышение требования",
        "Что проверить",
    ]
    _write_header(sheet, headers)

    rows = _component_candidate_rows(match_run)
    if not rows:
        sheet.cell(row=2, column=1, value="Варианты компонентов не сформированы.")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    else:
        for row_index, candidate in enumerate(rows, start=2):
            row_values = [
                row_index - 1,
                _component_role_label(str(candidate.get("role") or "")),
                candidate.get("producer") or "",
                candidate.get("part_number") or "",
                candidate.get("name") or "",
                candidate.get("available_quantity"),
                _money_value(candidate.get("price_value")),
                candidate.get("price_currency") or "",
                _facts_text(candidate.get("extracted_facts")),
                "\n".join(as_string_list(candidate.get("fit_reasons"))),
                candidate.get("fit_label") or "",
                candidate.get("fit_reason") or "",
                candidate.get("over_requirement") or "",
                "\n".join(as_string_list(candidate.get("eligibility_warnings"))),
            ]
            _write_row(sheet, row_index, row_values)

    _set_number_format(sheet, [7])
    _set_widths(
        sheet,
        {
            1: 6,
            2: 24,
            3: 18,
            4: 24,
            5: 52,
            6: 12,
            7: 16,
            8: 10,
            9: 52,
            10: 56,
            11: 22,
            12: 56,
            13: 18,
            14: 60,
        },
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fill_checks_sheet(sheet: Worksheet, match_run: MatchRun) -> None:
    headers = [
        "№",
        "Уровень",
        "Вариант / артикул",
        "Что проверить",
        "Детали",
        "Рекомендуемое действие",
    ]
    _write_header(sheet, headers)

    rows: list[list[Any]] = []
    for check in humanized_checks(
        risk_flags=match_run.risk_flags_json,
        missing_requirements=match_run.missing_requirements_json,
    ):
        rows.append(
            [
                len(rows) + 1,
                "Общее",
                "",
                check,
                check,
                recommended_action(check),
            ]
        )

    for candidate in _sorted_candidates(match_run):
        mapping = _candidate_mapping(candidate)
        raw = _raw_mapping(candidate)
        variant = candidate_display_name(mapping)
        for check in humanized_checks(
            risk_flags=[
                *as_string_list(candidate.risk_flags_json),
                *as_string_list(raw.get("compatibility_warnings")),
            ],
            missing_requirements=[
                *as_string_list(candidate.missing_requirements_json),
                *as_string_list(raw.get("missing_components")),
            ],
        ):
            rows.append(
                [
                    len(rows) + 1,
                    "Вариант",
                    variant,
                    check,
                    f"Артикул: {candidate_article(mapping)}",
                    recommended_action(check),
                ]
            )

    if not rows:
        rows.append([1, "Общее", "", "Дополнительные проверки не отмечены.", "", ""])

    for row_index, row_values in enumerate(rows, start=2):
        _write_row(sheet, row_index, row_values)

    _set_widths(sheet, {1: 6, 2: 14, 3: 28, 4: 52, 5: 36, 6: 44})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fill_requirements_sheet(sheet: Worksheet, spec_json: Mapping[str, Any]) -> None:
    headers = ["Параметр", "Требование", "Комментарий"]
    _write_header(sheet, headers)

    rows = _requirement_rows(spec_json)
    if not rows:
        rows = [["Требования", "Не извлечены", "Проверьте исходный запрос вручную."]]

    for row_index, row_values in enumerate(rows, start=2):
        _write_row(sheet, row_index, row_values)

    _set_widths(sheet, {1: 42, 2: 36, 3: 48})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _requirement_rows(spec_json: Mapping[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    source_text = spec_json.get("source_text")
    if source_text:
        rows.append(["Исходный текст", str(source_text), ""])

    shipment_city = spec_json.get("shipment_city")
    if shipment_city:
        rows.append(["Город отгрузки", str(shipment_city), ""])

    items = spec_json.get("items")
    if isinstance(items, list):
        for index, item in enumerate(items, start=1):
            if not isinstance(item, Mapping):
                continue
            prefix = f"Позиция {index}"
            rows.extend(_item_requirement_rows(prefix, item))

    requirements = spec_json.get("requirements")
    if isinstance(requirements, Mapping):
        rows.extend(_flatten_requirements("Общие требования", requirements))
    return rows


def _item_requirement_rows(prefix: str, item: Mapping[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    simple_fields = {
        "item_type": "Тип позиции",
        "quantity": "Количество",
        "name": "Название",
    }
    for key, label in simple_fields.items():
        if item.get(key) is not None:
            rows.append([f"{prefix} - {label}", _format_requirement_value(item[key]), ""])

    requirements = item.get("requirements")
    if isinstance(requirements, Mapping):
        rows.extend(_flatten_requirements(prefix, requirements))
    return rows


def _flatten_requirements(prefix: str, requirements: Mapping[str, Any]) -> list[list[str]]:
    labels = {
        "form_factor": "Форм-фактор",
        "cpu.sockets": "Процессорные сокеты",
        "ram.min_gb": "Оперативная память, минимум",
        "storage.type": "Тип накопителей",
        "power.psu_count": "Блоки питания",
        "power.redundant_psu": "Резервирование питания",
        "warranty": "Гарантия",
    }
    rows: list[list[str]] = []
    for path, value in _walk(requirements):
        label = labels.get(path, path)
        rows.append([f"{prefix} - {label}", _format_requirement_value(value), ""])
    return rows


def _walk(source: Mapping[str, Any], prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key, value in source.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            rows.extend(_walk(value, path))
        else:
            rows.append((path, value))
    return rows


def _write_header_at(sheet: Worksheet, row_index: int, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_row(sheet: Worksheet, row_index: int, values: list[Any]) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER


def _set_widths(sheet: Worksheet, widths: Mapping[int, int]) -> None:
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _set_number_format(sheet: Worksheet, columns: list[int]) -> None:
    for row_index in range(2, sheet.max_row + 1):
        for column_index in columns:
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, int | float):
                cell.number_format = MONEY_NUMBER_FORMAT


def _set_number_format_from_row(
    sheet: Worksheet,
    columns: list[int],
    *,
    start_row: int,
) -> None:
    for row_index in range(start_row, sheet.max_row + 1):
        for column_index in columns:
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, int | float):
                cell.number_format = MONEY_NUMBER_FORMAT


def _ai_recommendations(match_run: MatchRun) -> list[Mapping[str, Any]]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    for key in ("ai_recommendations", "llm_recommendations", "llm_recommended_build_candidates"):
        rows = _component_rows(report_json.get(key))
        if rows:
            return [row for row in rows if _is_user_visible_ai_recommendation(row)]
    return []


def _configuration_groups(match_run: MatchRun) -> list[Mapping[str, Any]]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    if report_json.get("grouped_presales_mode_used") is not True:
        return []
    return _component_rows(report_json.get("configuration_groups"))


def _primary_recommendation(match_run: MatchRun) -> Mapping[str, Any]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    if report_json.get("primary_recommendation_status") != "valid":
        return {}
    primary = report_json.get("primary_recommendation")
    return primary if isinstance(primary, Mapping) else {}


def _report_product_group(match_run: MatchRun) -> str:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    candidates: list[Any] = [report_json.get("product_group")]

    commercial = report_json.get("commercial_summary")
    if isinstance(commercial, Mapping):
        candidates.append(commercial.get("product_group"))

    primary = report_json.get("primary_recommendation")
    if isinstance(primary, Mapping):
        candidates.append(primary.get("product_group"))

    requirements = _first_report_requirements(match_run)
    candidates.append(requirements.get("product_group"))
    role_plan = requirements.get("role_plan")
    if isinstance(role_plan, Mapping):
        candidates.append(role_plan.get("product_group"))

    matrix = report_json.get("component_candidate_matrix")
    if isinstance(matrix, Mapping):
        candidates.append(matrix.get("product_group"))
        matrix_requirements = matrix.get("normalized_requirements")
        if isinstance(matrix_requirements, Mapping):
            candidates.append(matrix_requirements.get("product_group"))
        matrix_role_plan = matrix.get("role_plan")
        if isinstance(matrix_role_plan, Mapping):
            candidates.append(matrix_role_plan.get("product_group"))

    spec_json = match_run.spec_json if isinstance(match_run.spec_json, Mapping) else {}
    candidates.append(spec_json.get("product_group"))

    for candidate in candidates:
        product_group = str(candidate or "").strip().casefold()
        if product_group in {
            SERVER_PRODUCT_GROUP,
            NETWORK_PRODUCT_GROUP,
            STORAGE_PRODUCT_GROUP,
        }:
            return product_group
    return SERVER_PRODUCT_GROUP


def _report_required_roles(match_run: MatchRun) -> list[str]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    roles: list[str] = []
    roles.extend(as_string_list(report_json.get("required_roles")))

    requirements = _first_report_requirements(match_run)
    roles.extend(as_string_list(requirements.get("required_roles")))
    role_plan = requirements.get("role_plan")
    if isinstance(role_plan, Mapping):
        roles.extend(as_string_list(role_plan.get("required_roles")))

    matrix = report_json.get("component_candidate_matrix")
    if isinstance(matrix, Mapping):
        roles.extend(as_string_list(matrix.get("required_roles")))
        matrix_role_plan = matrix.get("role_plan")
        if isinstance(matrix_role_plan, Mapping):
            roles.extend(as_string_list(matrix_role_plan.get("required_roles")))

    return _unique_text([str(role).strip() for role in roles])


def _group_component_base_rows(
    component_base: Mapping[str, Any],
) -> list[tuple[str, Mapping[str, Any]]]:
    rows: list[tuple[str, Mapping[str, Any]]] = []
    for label, key in (("CPU", "cpu"), ("RAM", "ram"), ("SSD/HDD", "storage")):
        component = component_base.get(key)
        if isinstance(component, Mapping):
            rows.append((label, component))
    return rows


def _group_server_quantity(group: Mapping[str, Any]) -> int | None:
    for option in _component_rows(group.get("platform_options")):
        platform = option.get("platform")
        if isinstance(platform, Mapping):
            quantity = _int_value(platform.get("quantity_required"))
            if quantity is not None:
                return quantity
    base = group.get("component_base")
    if isinstance(base, Mapping):
        for component in base.values():
            if isinstance(component, Mapping):
                quantity = _int_value(component.get("server_quantity"))
                if quantity is not None:
                    return quantity
    return None


def _group_component_total_text(component: Mapping[str, Any]) -> str:
    role = str(component.get("role") or "").strip()
    quantity = _int_value(component.get("quantity_required"))
    if role == "ram":
        module_gb = _ram_module_capacity_gb(component)
        module_text = f", модули по {module_gb} ГБ" if module_gb is not None else ""
        return f"{_quantity_units(quantity)} всего{module_text}"
    return f"{_quantity_units(quantity)} всего"


def _component_note_lines(component: Mapping[str, Any]) -> list[str]:
    notes: list[str] = []
    role = str(component.get("role") or "").strip()
    if role == "ram":
        module_gb = _ram_module_capacity_gb(component)
        total_gb = _int_value(component.get("ram_total_gb_per_server"))
        if module_gb is not None:
            notes.append(f"модуль {module_gb} ГБ")
        if total_gb is not None:
            notes.append(f"{total_gb} ГБ на сервер")
    if role in {"ssd", "hdd"}:
        capacity = _storage_capacity_tb(component)
        if capacity is not None:
            notes.append(f"{_format_number(capacity)} ТБ")
        interface = _component_fact_text(component, "storage_interface")
        if interface:
            notes.append(interface)
    cores = _int_value(component.get("cpu_cores"))
    if role == "cpu" and cores is not None:
        notes.append(f"{cores} ядер")
    return notes


def _component_fact_text(component: Mapping[str, Any], key: str) -> str:
    facts = component.get("facts")
    if isinstance(facts, Mapping):
        value = facts.get(key)
        if value not in (None, ""):
            return str(value)
    value = component.get(key)
    if value not in (None, ""):
        return str(value)
    return ""


def _platform_option_role_label(option: Mapping[str, Any]) -> str:
    labels = {
        "cheapest_quote": "Самый дешевый для КП",
        "preferred_for_database": "Более спокойный под БД",
        "branded_safe": "Брендовый / инженерно понятный",
        "engineering_clear": "Инженерно понятный",
        "high_headroom": "С запасом",
        "alternative": "Альтернатива",
    }
    role = str(option.get("role") or option.get("option_role") or "").strip()
    return labels.get(role, "Альтернатива")


def _grouped_excel_checks(groups: list[Mapping[str, Any]]) -> list[str]:
    checks: list[str] = []
    product_group = "server"
    for group in groups:
        group_product = str(group.get("product_group") or "").strip()
        if group_product:
            product_group = group_product
        for check in as_string_list(group.get("engineer_checks")):
            text = _safe_user_text(check)
            if text:
                checks.append(text)
        for option in _component_rows(group.get("platform_options")):
            for check in as_string_list(option.get("engineer_checks")):
                text = _safe_user_text(check)
                if text:
                    checks.append(text)
    return grouped_engineer_check_summary(checks, product_group=product_group)


def _is_user_visible_ai_recommendation(recommendation: Mapping[str, Any]) -> bool:
    if str(recommendation.get("decision") or "").strip() == "do_not_use":
        return False
    return _fatal_warning_text(
        [
            *_checks_text(recommendation).splitlines(),
            *_missing_text(recommendation).splitlines(),
        ]
    ) is None


def _normalized_requirements_text(match_run: MatchRun) -> str:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    normalized = report_json.get("normalized_requirements")
    if isinstance(normalized, list) and normalized:
        rows = [row for row in normalized if isinstance(row, Mapping)]
        if rows:
            return "\n\n".join(_normalized_requirement_summary(row, match_run) for row in rows)
    if isinstance(normalized, Mapping) and normalized:
        return _normalized_requirement_summary(normalized, match_run)
    matrix = report_json.get("component_candidate_matrix")
    if isinstance(matrix, Mapping) and isinstance(matrix.get("normalized_requirements"), Mapping):
        return _normalized_requirement_summary(matrix["normalized_requirements"], match_run)
    return ""


def _first_report_requirements(match_run: MatchRun) -> Mapping[str, Any]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    normalized = report_json.get("normalized_requirements")
    if isinstance(normalized, Mapping):
        return normalized
    if isinstance(normalized, list):
        for row in normalized:
            if isinstance(row, Mapping):
                return row
    matrix = report_json.get("component_candidate_matrix")
    if isinstance(matrix, Mapping):
        matrix_requirements = matrix.get("normalized_requirements")
        if isinstance(matrix_requirements, Mapping):
            return matrix_requirements
    return {}


def _normalized_requirement_summary(
    requirements: Mapping[str, Any],
    match_run: MatchRun,
) -> str:
    product_group = str(
        requirements.get("product_group") or _report_product_group(match_run)
    ).strip()
    if product_group == NETWORK_PRODUCT_GROUP:
        return _network_normalized_requirement_summary(requirements, match_run)
    if product_group == STORAGE_PRODUCT_GROUP:
        return _storage_normalized_requirement_summary(requirements, match_run)

    spec_json = match_run.spec_json if isinstance(match_run.spec_json, Mapping) else {}
    location = requirements.get("location") or spec_json.get("shipment_city") or ""
    cpu_parts = [
        _label_value("на сервер", requirements.get("cpu_per_server")),
        _label_value("всего", requirements.get("total_cpu_required")),
        _label_value("vendor", _unknown_to_empty(requirements.get("cpu_vendor_preference"))),
        _label_value("family", _unknown_to_empty(requirements.get("cpu_family_preference"))),
        _label_value("минимум ядер", requirements.get("cpu_min_cores_per_cpu")),
    ]
    ram_text = " ".join(
        str(value)
        for value in [
            _label_value("объем", _gb_text(requirements.get("ram_gb_per_server"))),
            _label_value("тип", _unknown_to_empty(requirements.get("ram_type_preference"))),
        ]
        if value
    )
    storage_text = " ".join(
        str(value)
        for value in [
            _label_value("тип", _unknown_to_empty(requirements.get("storage_type_preference"))),
            _label_value(
                "интерфейс",
                _unknown_to_empty(requirements.get("storage_interface_preference")),
            ),
            _label_value("объем", requirements.get("storage_min_capacity")),
            _label_value("шт./сервер", requirements.get("storage_qty_per_server")),
        ]
        if value
    )
    rows = [
        ("Количество серверов", requirements.get("server_qty")),
        ("Форм-фактор", requirements.get("form_factor")),
        ("CPU", ", ".join(part for part in cpu_parts if part)),
        ("RAM", ram_text),
        ("Накопители", storage_text),
        ("PSU", requirements.get("psu_count_per_server")),
        ("Склад/город", location),
    ]
    return "\n".join(
        f"- {label}: {value}" for label, value in rows if value not in (None, "")
    )


def _network_normalized_requirement_summary(
    requirements: Mapping[str, Any],
    match_run: MatchRun,
) -> str:
    spec_json = match_run.spec_json if isinstance(match_run.spec_json, Mapping) else {}
    location = requirements.get("location") or spec_json.get("shipment_city") or ""
    role = (
        str(requirements.get("network_device_role") or "").strip()
        or _first_role_from_capabilities(requirements, NETWORK_MATRIX_ROLES)
        or _first_role_from_list(requirements.get("required_roles"), NETWORK_MATRIX_ROLES)
    )
    parsed = _merged_capability_requirements(requirements, role=role, hard=True)
    optional = _merged_capability_requirements(requirements, role=role, hard=False)
    device_qty = (
        requirements.get("device_qty")
        or requirements.get("device_quantity")
        or parsed.get("device_count")
        or parsed.get("count")
        or requirements.get("server_qty")
    )
    rows = [
        ("Количество сетевого оборудования", device_qty),
        ("Тип оборудования", _component_role_label(role, NETWORK_PRODUCT_GROUP) if role else ""),
        ("Склад/город", location),
        ("Порты", _network_ports_requirement_text(parsed)),
        ("Uplinks", _network_uplinks_requirement_text(parsed)),
        ("PoE", _network_poe_requirement_text(parsed)),
        ("L2/L3", _network_l2_l3_requirement_text(parsed)),
        ("Стекирование", _network_stacking_requirement_text(parsed, optional)),
    ]
    return "\n".join(
        [
            f"- product_group={NETWORK_PRODUCT_GROUP}",
            *[
                f"- {label}: {value}"
                for label, value in rows
                if value not in (None, "")
            ],
        ]
    )


def _storage_normalized_requirement_summary(
    requirements: Mapping[str, Any],
    match_run: MatchRun,
) -> str:
    spec_json = match_run.spec_json if isinstance(match_run.spec_json, Mapping) else {}
    location = requirements.get("location") or spec_json.get("shipment_city") or ""
    system_qty = (
        requirements.get("system_qty")
        or requirements.get("device_qty")
        or requirements.get("device_quantity")
        or requirements.get("server_qty")
    )
    rows = [
        ("Количество СХД", system_qty),
        ("Склад/город", location),
        ("Емкость", _storage_capacity_requirement_text(requirements)),
        ("Контроллеры", requirements.get("controller_count")),
        ("Диски", _storage_drive_requirement_text(requirements)),
        ("Протокол/порты", _storage_protocol_requirement_text(requirements)),
        ("Поддержка/лицензии", _storage_support_requirement_text(requirements)),
    ]
    return "\n".join(
        [
            f"- product_group={STORAGE_PRODUCT_GROUP}",
            *[
                f"- {label}: {value}"
                for label, value in rows
                if value not in (None, "")
            ],
        ]
    )


def _first_role_from_capabilities(
    requirements: Mapping[str, Any],
    role_order: tuple[str, ...],
) -> str:
    allowed = set(role_order)
    for key in ("required_capabilities", "optional_capabilities"):
        for capability in _component_rows(requirements.get(key)):
            role = str(capability.get("role") or "").strip()
            if role in allowed:
                return role
    return ""


def _first_role_from_list(value: Any, role_order: tuple[str, ...]) -> str:
    allowed = set(as_string_list(value))
    for role in role_order:
        if role in allowed:
            return role
    return ""


def _merged_capability_requirements(
    requirements: Mapping[str, Any],
    *,
    role: str,
    hard: bool,
) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    keys = ("required_capabilities",) if hard else ("optional_capabilities",)
    for key in keys:
        for capability in _component_rows(requirements.get(key)):
            capability_role = str(capability.get("role") or "").strip()
            if role and capability_role and capability_role != role:
                continue
            if hard and capability.get("hard") is False:
                continue
            parsed = capability.get("parsed_requirements")
            if isinstance(parsed, Mapping):
                merged.update(parsed)
    if hard:
        network_requirement = requirements.get("network_requirement")
        if isinstance(network_requirement, Mapping):
            merged.update({key: value for key, value in network_requirement.items() if value})
    return merged


def _network_ports_requirement_text(parsed: Mapping[str, Any]) -> str:
    count = parsed.get("port_count") or parsed.get("min_ports") or parsed.get("ports")
    speed = parsed.get("port_speed") or parsed.get("speed")
    media = parsed.get("port_media") or parsed.get("media")
    return _count_speed_media_text(count, speed, media)


def _network_uplinks_requirement_text(parsed: Mapping[str, Any]) -> str:
    count = parsed.get("uplink_count")
    speed = parsed.get("uplink_speed")
    media = parsed.get("uplink_media")
    return _count_speed_media_text(count, speed, media)


def _count_speed_media_text(count: Any, speed: Any, media: Any) -> str:
    parts = [str(value) for value in (speed, media) if value not in (None, "", "unknown")]
    if count not in (None, "", "unknown"):
        return f"{count} x {' '.join(parts)}".strip()
    return " ".join(parts)


def _network_poe_requirement_text(parsed: Mapping[str, Any]) -> str:
    if not parsed.get("poe_required") and parsed.get("poe_supported") is not True:
        return ""
    parts = ["требуется"]
    standard = parsed.get("poe_standard")
    budget = parsed.get("poe_budget_w") or parsed.get("poe_budget")
    if standard not in (None, "", "unknown"):
        parts.append(str(standard))
    if budget not in (None, "", "unknown"):
        parts.append(f"{budget} W")
    return " ".join(parts)


def _network_l2_l3_requirement_text(parsed: Mapping[str, Any]) -> str:
    values: list[str] = []
    if parsed.get("l2_required") or parsed.get("l2_supported") is True:
        values.append("L2")
    if parsed.get("l3_required") or parsed.get("l3_supported") is True:
        values.append("L3")
    return "/".join(values)


def _network_stacking_requirement_text(
    parsed: Mapping[str, Any],
    optional: Mapping[str, Any],
) -> str:
    value = parsed.get("stacking_required")
    if value is None:
        value = optional.get("stacking_required")
    if value is True:
        return "требуется или желательно"
    if value is False:
        return "не требуется"
    return ""


def _storage_capacity_requirement_text(requirements: Mapping[str, Any]) -> str:
    parts = []
    usable = requirements.get("usable_capacity_tb")
    raw = requirements.get("raw_capacity_tb")
    minimum = requirements.get("storage_min_capacity")
    if usable not in (None, "", "unknown"):
        parts.append(f"usable {_tb_text(usable)}")
    if raw not in (None, "", "unknown"):
        parts.append(f"raw {_tb_text(raw)}")
    if not parts and minimum not in (None, "", "unknown"):
        parts.append(str(minimum))
    redundancy = _unknown_to_empty(requirements.get("redundancy_level"))
    if redundancy:
        parts.append(str(redundancy))
    return "; ".join(parts)


def _storage_drive_requirement_text(requirements: Mapping[str, Any]) -> str:
    count = requirements.get("drive_count") or requirements.get("storage_qty_per_server")
    capacity = requirements.get("drive_capacity_tb")
    drive_type = _unknown_to_empty(requirements.get("drive_type"))
    interface = _unknown_to_empty(
        requirements.get("drive_interface") or requirements.get("storage_interface_preference")
    )
    parts = []
    if count not in (None, "", "unknown"):
        parts.append(f"{count} шт.")
    if capacity not in (None, "", "unknown"):
        parts.append(_tb_text(capacity))
    for value in (drive_type, interface):
        if value:
            parts.append(str(value))
    return " ".join(parts)


def _storage_protocol_requirement_text(requirements: Mapping[str, Any]) -> str:
    protocol = _unknown_to_empty(requirements.get("host_protocol"))
    count = requirements.get("host_port_count")
    speed = _unknown_to_empty(requirements.get("host_port_speed"))
    media = _unknown_to_empty(requirements.get("host_port_media"))
    parts = [str(value) for value in (protocol,) if value]
    port_text = _count_speed_media_text(count, speed, media)
    if port_text:
        parts.append(port_text)
    return "; ".join(parts)


def _storage_support_requirement_text(requirements: Mapping[str, Any]) -> str:
    parts = []
    if requirements.get("support_required") is True:
        parts.append("поддержка требуется")
    if requirements.get("license_required") is True:
        parts.append("лицензия требуется")
    warranty_months = requirements.get("warranty_months")
    if warranty_months not in (None, "", "unknown"):
        parts.append(f"{warranty_months} мес.")
    return "; ".join(parts)


def _tb_text(value: Any) -> str:
    if value in (None, "", "unknown"):
        return ""
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return str(value)
    if amount == amount.to_integral():
        return f"{int(amount)} ТБ"
    return f"{format(amount.normalize(), 'f')} ТБ"


def _label_value(label: str, value: Any) -> str:
    if value in (None, "", "unknown"):
        return ""
    text = _safe_user_text(value)
    return f"{label}: {text}" if text else ""


def _unknown_to_empty(value: Any) -> Any:
    return "" if value in (None, "", "unknown") else value


def _gb_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return f"{value} ГБ"


def _recommendation_type_text(recommendation: Mapping[str, Any]) -> str:
    labels = {
        "ready_server": "Готовый сервер",
        "build_from_parts": "Сборка из комплектующих",
        "partial_build": "Частичная сборка",
    }
    return labels.get(_recommendation_source_type(recommendation), "AI-рекомендация")


def _recommendation_source_type(recommendation: Mapping[str, Any]) -> str:
    source_type = str(
        recommendation.get("source_type") or recommendation.get("candidate_type") or ""
    )
    if source_type == "build_from_parts":
        missing = as_string_list(recommendation.get("missing_component_roles"))
        if recommendation.get("completeness_status") == "incomplete" or missing:
            return "partial_build"
    if source_type in {"ready_server", "build_from_parts", "partial_build"}:
        return source_type
    return "build_from_parts"


def _recommendation_status_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "partial_build":
        return "Частичная сборка"
    decision = str(recommendation.get("decision") or "").strip()
    if decision == "recommend":
        return "Рекомендовано"
    if decision == "recommend_with_checks":
        return "Рекомендовано с проверками"
    status = str(recommendation.get("completeness_status") or "").strip()
    if status == "incomplete":
        return "Частично"
    return "Предварительно"


def _recommendation_display_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "ready_server":
        summary = recommendation.get("component_summary")
        if isinstance(summary, Mapping):
            platform = str(summary.get("platform") or "").strip()
            if platform:
                return platform
        display = candidate_display_name(recommendation)
        if display != "артикул не указан":
            return display
        return str(recommendation.get("display_name") or "").strip()

    platform = _component_article_for_role(recommendation, "server_platform")
    if platform:
        return platform
    display = candidate_display_name(recommendation)
    return "" if display == "артикул не указан" else display


def _composition_per_server_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "ready_server":
        article = _ready_server_article(recommendation)
        return f"Готовая позиция: {article}" if article else "Готовая позиция"

    server_quantity = _server_quantity_from_report_candidate(recommendation)
    rows: list[str] = []
    for role in ("server_platform", "cpu", "ram", "ssd", "hdd"):
        component = _component_by_role(recommendation, role)
        if component is None:
            continue
        line = _bom_per_server_line(component, server_quantity)
        if line:
            rows.append(line)
    return "\n".join(rows)


def _total_quantities_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "ready_server":
        quantity = _server_quantity_from_report_candidate(recommendation)
        return f"Готовая позиция: {_quantity_units(quantity)}"

    rows: list[str] = []
    for role in ("server_platform", "cpu", "ram", "ssd", "hdd"):
        component = _component_by_role(recommendation, role)
        if component is None:
            continue
        quantity = _int_value(component.get("quantity_required"))
        rows.append(
            f"{_bom_role_label(role)}: {_quantity_units(quantity)} всего"
        )
    return "\n".join(rows)


def _component_stocks_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "ready_server":
        return f"Готовая позиция: {_stock_count_text(recommendation.get('available_quantity'))}"

    rows: list[str] = []
    for role in ("server_platform", "cpu", "ram", "ssd", "hdd"):
        component = _component_by_role(recommendation, role)
        if component is None:
            continue
        stock = _stock_count_text(component.get("available_quantity"))
        rows.append(f"{_bom_role_label(role)}: {stock}")
    return "\n".join(rows)


def _optional_components_text(recommendation: Mapping[str, Any]) -> str:
    rows: list[str] = []
    for component in _component_rows(recommendation.get("optional_components")):
        role = str(component.get("role") or "").strip()
        article = _component_article(component)
        quantity = _int_value(component.get("quantity_required"))
        stock = _stock_count_text(component.get("available_quantity"))
        if article:
            rows.append(
                f"{_bom_role_label(role)}: {article}; "
                f"{_quantity_units(quantity)} всего; склад {stock}"
            )
    optional_total = _money_value(recommendation.get("optional_total_price_value"))
    optional_currency = str(recommendation.get("optional_total_price_currency") or "").strip()
    if optional_total is not None:
        amount = (
            f"{optional_total:,}".replace(",", " ")
            if isinstance(optional_total, int)
            else str(optional_total)
        )
        rows.append(
            "Опциональная сумма отдельно: "
            + " ".join(part for part in [amount, optional_currency] if part)
        )
    return "\n".join(rows)


def _bom_per_server_line(
    component: Mapping[str, Any],
    server_quantity: int | None,
) -> str:
    role = str(component.get("role") or "").strip()
    article = commercial_component_name(component)
    if not article:
        return ""
    total_quantity = _int_value(component.get("quantity_required"))
    per_server_quantity = _int_value(component.get("per_server_quantity"))
    if per_server_quantity is None:
        per_server_quantity = _per_server_quantity(total_quantity, server_quantity, role)
    suffix = ""
    if role == "ram":
        module_gb = _ram_module_capacity_gb(component)
        ram_total_gb = _ram_total_gb_per_server(component, per_server_quantity)
        if module_gb is not None and per_server_quantity is not None:
            total_text = (
                f" = {_format_number(ram_total_gb)} ГБ"
                if ram_total_gb is not None
                else ""
            )
            return (
                f"{_bom_role_label(role)}: {article} - "
                f"{per_server_quantity} x {module_gb} ГБ на сервер{total_text}"
            )
    elif role in {"ssd", "hdd"}:
        capacity = _storage_capacity_tb(component)
        if capacity is not None:
            suffix = f" x {_format_number(capacity)} ТБ на сервер"
    return f"{_bom_role_label(role)}: {article} - {_quantity_units(per_server_quantity)}{suffix}"


def _component_by_role(
    recommendation: Mapping[str, Any],
    role: str,
) -> Mapping[str, Any] | None:
    for component in _component_rows(recommendation.get("components")):
        if component.get("role") == role:
            return component
    return None


def _component_article_for_role(recommendation: Mapping[str, Any], role: str) -> str:
    component = _component_by_role(recommendation, role)
    return _component_article(component) if component is not None else ""


def _component_article(component: Mapping[str, Any]) -> str:
    parts = [
        str(component.get("producer") or "").strip(),
        str(component.get("part_number") or "").strip(),
    ]
    display = " ".join(part for part in parts if part)
    if display:
        return display
    return str(
        component.get("item_name")
        or component.get("name")
        or component.get("item_id")
        or ""
    ).strip()


def _ready_server_article(recommendation: Mapping[str, Any]) -> str:
    article = str(recommendation.get("part_number") or "").strip() or candidate_article(
        recommendation
    )
    if article:
        return article
    display = candidate_display_name(recommendation)
    if display != "артикул не указан":
        return display
    return str(recommendation.get("display_name") or "").strip()


def _bom_role_label(role: str) -> str:
    labels = {
        "ready_server": "готовый сервер",
        "server_platform": "Платформа",
        "storage_system": "СХД",
        "controller": "Контроллер СХД",
        "controller_module": "Контроллерный модуль",
        "disk_shelf": "Дисковая полка",
        "drive": "Диск",
        "cache": "Кэш",
        "host_port": "Host-порт",
        "protocol_module": "Протокольный модуль",
        "switch": "Коммутатор",
        "router": "Маршрутизатор",
        "firewall": "Межсетевой экран",
        "access_point": "Точка доступа",
        "cpu": "CPU",
        "ram": "RAM",
        "ssd": "SSD",
        "hdd": "HDD",
        "storage_controller": "Контроллер",
        "network_adapter": "Сетевой адаптер",
        "transceiver": "Трансивер",
        "dac_cable": "DAC-кабель",
        "cable": "Кабель",
        "license": "Лицензия",
        "support": "Поддержка",
        "power_supply": "Блок питания",
        "stacking_module": "Модуль стекирования",
        "other_accessory": "Аксессуар",
    }
    return labels.get(role, role.upper() if role else "Компонент")


def _per_server_quantity(
    total_quantity: int | None,
    server_quantity: int | None,
    role: str,
) -> int | None:
    if role == "server_platform":
        return 1
    if total_quantity is None:
        return None
    if server_quantity is not None and server_quantity > 0:
        return max(1, total_quantity // server_quantity)
    return total_quantity


def _quantity_units(value: int | None) -> str:
    if value is None:
        return "количество уточняется"
    return f"{value} шт."


def _stock_count_text(value: Any) -> str:
    if value in (None, ""):
        return "неизвестно"
    return str(value)


def _ram_total_gb_per_server(
    component: Mapping[str, Any],
    per_server_quantity: int | None,
) -> int | None:
    if per_server_quantity is None:
        return None
    explicit_total = _int_value(component.get("ram_total_gb_per_server"))
    if explicit_total is not None:
        return explicit_total
    module_gb = _ram_module_capacity_gb(component)
    if module_gb is None:
        return None
    return module_gb * per_server_quantity


def _ram_module_capacity_gb(component: Mapping[str, Any]) -> int | None:
    module_gb = _int_value(component.get("ram_module_capacity_gb"))
    facts = component.get("facts")
    if module_gb is None and isinstance(facts, Mapping):
        module_gb = _int_value(facts.get("ram_capacity_gb"))
    return module_gb


def _storage_capacity_tb(component: Mapping[str, Any]) -> Decimal | None:
    value = component.get("storage_capacity_tb")
    facts = component.get("facts")
    if value in (None, "") and isinstance(facts, Mapping):
        value = facts.get("storage_capacity_tb")
    if value in (None, ""):
        return None
    try:
        return value if isinstance(value, Decimal) else Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _format_number(value: int | Decimal) -> str:
    amount = value if isinstance(value, Decimal) else Decimal(value)
    if amount == amount.to_integral():
        return str(int(amount))
    return format(amount.normalize(), "f")


def _recommendation_component_text(recommendation: Mapping[str, Any], key: str) -> str:
    role = "server_platform" if key == "platform" else key
    component_text = _components_text(_component_rows(recommendation.get("components")), role)
    if component_text:
        return component_text
    summary = recommendation.get("component_summary")
    if isinstance(summary, Mapping):
        return str(summary.get(key) or "").strip()
    return ""


def _recommendation_storage_text(recommendation: Mapping[str, Any]) -> str:
    component_text = _storage_components_text(_component_rows(recommendation.get("components")))
    if component_text:
        return component_text
    summary = recommendation.get("component_summary")
    if isinstance(summary, Mapping):
        return str(summary.get("storage") or "").strip()
    return ""


def _availability_text(recommendation: Mapping[str, Any]) -> str:
    available = recommendation.get("available_quantity")
    if available is None or available == "":
        return ""
    return f"{available} шт."


def _missing_text(recommendation: Mapping[str, Any]) -> str:
    values = [
        *as_string_list(recommendation.get("what_is_missing")),
        *as_string_list(recommendation.get("missing_components")),
        *as_string_list(recommendation.get("missing_requirements")),
    ]
    return "\n".join(_unique_text([_safe_user_text(value) for value in values]))


def _checks_text(
    recommendation: Mapping[str, Any],
    *,
    normalized_requirements: Mapping[str, Any] | None = None,
) -> str:
    values = [
        *as_string_list(recommendation.get("critical_checks")),
        *as_string_list(recommendation.get("critical_risks")),
        *as_string_list(recommendation.get("compatibility_warnings")),
        *as_string_list(recommendation.get("risk_flags")),
    ]
    checks = humanized_checks(risk_flags=values, missing_requirements=[])
    return "\n".join(
        _filter_contradictory_checks(
            _unique_text([_safe_user_text(value) for value in checks]),
            normalized_requirements or {},
        )
    )


def _matrix_checks_text(
    row: Mapping[str, Any],
    *,
    normalized_requirements: Mapping[str, Any],
) -> str:
    values = [
        *as_string_list(row.get("eligibility_warnings")),
        *as_string_list(row.get("risks")),
        *as_string_list(row.get("gaps")),
    ]
    checks = humanized_checks(risk_flags=values, missing_requirements=[])
    return "\n".join(
        _filter_contradictory_checks(
            _unique_text([_safe_user_text(value) for value in checks]),
            normalized_requirements,
        )
    )


def _why_selected_excel_text(recommendation: Mapping[str, Any]) -> str:
    short = _safe_user_text(recommendation.get("why_selected_short") or "")
    if short:
        return short
    return _safe_user_text(recommendation.get("why_selected") or "")


def _work_action_text(recommendation: Mapping[str, Any]) -> str:
    if _recommendation_source_type(recommendation) == "partial_build":
        return "только после доукомплектования и инженерной проверки"
    if _missing_text(recommendation):
        return "только после доукомплектования и инженерной проверки"
    if str(recommendation.get("decision") or "") in {"recommend", "recommend_with_checks"}:
        return "после инженерной проверки"
    return "после инженерной проверки"


def _fatal_warning_text(values: list[str]) -> str | None:
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        lowered = text.casefold()
        if (
            "fatal" in lowered
            or "incompat" in lowered
            or "mismatch" in lowered
            or "несовмест" in lowered
        ):
            return text
    return None


def _unified_component_matrix_rows(match_run: MatchRun) -> list[Mapping[str, Any]]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    product_group = _report_product_group(match_run)
    rows: list[Mapping[str, Any]] = []
    evidence_by_id = _report_evidence_by_component_id(report_json)
    matrix_rows = _component_candidate_rows(match_run, product_group=product_group)
    if (
        product_group == SERVER_PRODUCT_GROUP
        and any(row.get("role") == "ready_server" for row in matrix_rows)
    ):
        rows.extend(_with_matrix_evidence(row, evidence_by_id) for row in matrix_rows)
        return rows
    if product_group == SERVER_PRODUCT_GROUP:
        for candidate in _component_rows(report_json.get("ready_stock_candidates")):
            candidate_id = str(
                candidate.get("candidate_id") or candidate.get("item_id") or ""
            ).strip()
            rows.append(
                _with_matrix_evidence(
                    {
                        "role": "ready_server",
                        "role_label": "готовый сервер",
                        "producer": candidate.get("producer"),
                        "part_number": candidate.get("part_number"),
                        "name": candidate.get("item_name") or candidate.get("name"),
                        "available_quantity": candidate.get("available_quantity"),
                        "price_value": candidate.get("price_value"),
                        "price_currency": candidate.get("price_currency"),
                        "extracted_facts": {},
                        "fit_label": "",
                        "fit_reason": "\n".join(
                            as_string_list(candidate.get("matched_requirements"))
                        ),
                        "over_requirement": "",
                        "gaps": candidate.get("missing_requirements"),
                        "risks": candidate.get("risk_flags"),
                        "component_candidate_id": candidate_id,
                    },
                    evidence_by_id,
                )
            )
    rows.extend(_with_matrix_evidence(row, evidence_by_id) for row in matrix_rows)
    return rows


def _unique_text(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in result:
            result.append(text)
    return result


def _component_candidate_rows(
    match_run: MatchRun,
    *,
    product_group: str | None = None,
) -> list[Mapping[str, Any]]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    matrix = report_json.get("component_candidate_matrix")
    if not isinstance(matrix, Mapping):
        return []

    product_group = product_group or _report_product_group(match_run)
    rows: list[Mapping[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for key in _component_candidate_keys(matrix, product_group):
        value = matrix.get(key)
        if not isinstance(value, list):
            continue
        fallback_role = MATRIX_CANDIDATE_ROLE_BY_KEY.get(key, key.removesuffix("_candidates"))
        for candidate in value:
            if not isinstance(candidate, Mapping):
                continue
            row = dict(candidate)
            row.setdefault("role", fallback_role)
            role = _matrix_candidate_role(row)
            if not _matrix_role_allowed(role, product_group):
                continue
            identity = _matrix_row_identity(row, role)
            if identity in seen:
                continue
            seen.add(identity)
            rows.append(row)
    return rows


def _component_candidate_keys(matrix: Mapping[str, Any], product_group: str) -> list[str]:
    if product_group == SERVER_PRODUCT_GROUP:
        return [
            "ready_server_candidates",
            "platform_candidates",
            "cpu_candidates",
            "ram_candidates",
            "ssd_candidates",
            "hdd_candidates",
            "storage_controller_candidates",
            "network_adapter_candidates",
        ]

    keys: list[str] = []
    for role in _matrix_role_order(product_group):
        role_keys = [
            key
            for key, mapped_role in MATRIX_CANDIDATE_ROLE_BY_KEY.items()
            if mapped_role == role
        ]
        role_keys.append(f"{role}_candidates")
        for key in role_keys:
            if key in matrix and key not in keys:
                keys.append(key)

    for key, value in matrix.items():
        if not key.endswith("_candidates") or key in keys or not isinstance(value, list):
            continue
        role = MATRIX_CANDIDATE_ROLE_BY_KEY.get(key, key.removesuffix("_candidates"))
        if _matrix_role_allowed(role, product_group):
            keys.append(key)
            continue
        if any(
            isinstance(candidate, Mapping)
            and _matrix_role_allowed(_matrix_candidate_role(candidate), product_group)
            for candidate in value
        ):
            keys.append(key)
    return keys


def _matrix_role_order(product_group: str) -> tuple[str, ...]:
    if product_group == NETWORK_PRODUCT_GROUP:
        return NETWORK_MATRIX_ROLES
    if product_group == STORAGE_PRODUCT_GROUP:
        return STORAGE_MATRIX_ROLES
    return SERVER_MATRIX_ROW_ROLES


def _matrix_role_allowed(role: str, product_group: str) -> bool:
    if product_group == NETWORK_PRODUCT_GROUP:
        return role in NETWORK_MATRIX_ROLES
    if product_group == STORAGE_PRODUCT_GROUP:
        return role in STORAGE_MATRIX_ROLES
    return role in SERVER_MATRIX_ROW_ROLES or role == "platform"


def _matrix_candidate_role(row: Mapping[str, Any]) -> str:
    role = str(row.get("role") or "").strip()
    if role == "platform":
        return "server_platform"
    return role


def _matrix_row_identity(row: Mapping[str, Any], role: str) -> tuple[str, str, str]:
    identifier = str(
        row.get("component_candidate_id")
        or row.get("candidate_id")
        or row.get("item_id")
        or ""
    ).strip()
    part_number = str(row.get("part_number") or "").strip()
    name = str(row.get("name") or row.get("item_name") or "").strip()
    return role, identifier or part_number, name


def _report_evidence_by_component_id(
    report_json: Mapping[str, Any],
) -> dict[str, Mapping[str, Any]]:
    pack = report_json.get("web_evidence_pack")
    if not isinstance(pack, Mapping):
        return {}
    evidence_by_id: dict[str, Mapping[str, Any]] = {}
    for component in _component_rows(pack.get("components")):
        component_id = str(component.get("component_candidate_id") or "").strip()
        if component_id:
            evidence_by_id[component_id] = component
    return evidence_by_id


def _report_evidence_mode_text(report_json: Mapping[str, Any]) -> str:
    mode = str(report_json.get("evidence_mode") or "").strip()
    if not mode:
        diagnostics = report_json.get("web_evidence_diagnostics")
        if isinstance(diagnostics, Mapping):
            mode = str(diagnostics.get("evidence_mode") or "").strip()
    if mode == "online_composer":
        return "online composer"
    if mode == "separate":
        return "separate"
    pack = report_json.get("web_evidence_pack")
    if isinstance(pack, Mapping) and pack.get("enabled"):
        return "separate"
    return "не использовалась"


def _report_evidence_sources_count(report_json: Mapping[str, Any]) -> int:
    value = _int_value(report_json.get("evidence_sources_count"))
    if value is not None:
        return value
    diagnostics = report_json.get("web_evidence_diagnostics")
    if isinstance(diagnostics, Mapping):
        value = _int_value(diagnostics.get("evidence_sources_count"))
        if value is not None:
            return value
    count = 0
    pack = report_json.get("web_evidence_pack")
    if isinstance(pack, Mapping):
        for component in _component_rows(pack.get("components")):
            count += len(_component_rows(component.get("sources")))
    return count


def _with_matrix_evidence(
    row: Mapping[str, Any],
    evidence_by_id: Mapping[str, Mapping[str, Any]],
) -> Mapping[str, Any]:
    component_id = str(
        row.get("component_candidate_id") or row.get("candidate_id") or row.get("item_id") or ""
    ).strip()
    evidence = evidence_by_id.get(component_id)
    if not evidence:
        return row
    merged = dict(row)
    merged["evidence"] = evidence
    merged["evidence_status"] = evidence.get("evidence_status")
    merged["evidence_confidence"] = evidence.get("confidence")
    merged["evidence_source_count"] = len(_component_rows(evidence.get("sources")))
    merged["evidence_confirmed"] = _evidence_confirmed_texts(evidence)
    merged["evidence_missing"] = _evidence_missing_texts(evidence)
    return merged


def _matrix_row_role_label(row: Mapping[str, Any], product_group: str) -> str:
    role = _matrix_candidate_role(row)
    if product_group == SERVER_PRODUCT_GROUP:
        return str(row.get("role_label") or _component_role_label(role, product_group))
    return _component_role_label(role, product_group)


def _component_role_label(role: str, product_group: str = SERVER_PRODUCT_GROUP) -> str:
    network_labels = {
        "switch": "Коммутатор",
        "router": "Маршрутизатор",
        "firewall": "Межсетевой экран",
        "access_point": "Точка доступа",
        "transceiver": "Трансивер",
        "dac_cable": "Кабель/DAC",
        "cable": "Кабель/DAC",
        "license": "Поддержка/лицензии",
        "support": "Поддержка/лицензии",
        "power_supply": "Блок питания",
        "stacking_module": "Модуль стекирования",
        "other_accessory": "Аксессуар",
    }
    storage_labels = {
        "storage_system": "СХД",
        "controller": "Контроллер СХД",
        "controller_module": "Контроллерный модуль",
        "disk_shelf": "Дисковая полка",
        "drive": "Диск",
        "ssd": "SSD",
        "hdd": "HDD",
        "cache": "Кэш",
        "host_port": "Host-порт",
        "protocol_module": "Протокольный модуль",
        "transceiver": "Трансивер",
        "cable": "Кабель",
        "license": "Лицензия",
        "support": "Поддержка",
        "power_supply": "Блок питания",
        "rail_kit": "Рельсы",
        "other_accessory": "Аксессуар",
    }
    if product_group == NETWORK_PRODUCT_GROUP:
        return network_labels.get(role, role)
    if product_group == STORAGE_PRODUCT_GROUP:
        return storage_labels.get(role, role)
    labels = {
        "ready_server": "готовый сервер",
        "server_platform": "Платформа",
        "switch": "Коммутатор",
        "router": "Маршрутизатор",
        "firewall": "Межсетевой экран",
        "access_point": "Точка доступа",
        "cpu": "CPU",
        "ram": "RAM",
        "ssd": "SSD",
        "hdd": "HDD",
        "storage_controller": "Контроллер",
        "network_adapter": "Сетевой адаптер",
        "transceiver": "Трансивер",
        "dac_cable": "DAC-кабель",
        "cable": "Кабель",
        "license": "Лицензия",
        "support": "Поддержка",
        "power_supply": "Блок питания",
        "stacking_module": "Модуль стекирования",
        "other_accessory": "Аксессуар",
    }
    return labels.get(role, role)


def _facts_text(value: Any) -> str:
    if not isinstance(value, Mapping):
        return ""
    rows: list[str] = []
    labels = {
        "normalized_vendor": "Производитель",
        "cpu_brand": "Производитель CPU",
        "cpu_family": "Семейство CPU",
        "cpu_cores": "Ядра CPU",
        "ram_capacity_gb": "Модуль RAM",
        "ram_type": "Тип RAM",
        "storage_capacity": "Объем накопителя",
        "storage_capacity_tb": "Объем накопителя, ТБ",
        "storage_interface": "Интерфейс накопителя",
        "raw_capacity_tb": "Raw, ТБ",
        "usable_capacity_tb": "Usable, ТБ",
        "redundancy_level": "RAID/защита",
        "controller_count": "Контроллеры",
        "drive_count": "Диски",
        "drive_capacity_tb": "Емкость диска, ТБ",
        "drive_type": "Тип диска",
        "drive_interface": "Интерфейс диска",
        "host_protocol": "Протокол",
        "host_port_count": "Host ports",
        "host_port_speed": "Скорость host ports",
        "host_port_media": "Среда host ports",
        "warranty_months": "Гарантия, мес.",
        "network_ports_count": "Сетевые порты",
        "network_speed": "Скорость сети",
        "network_media": "Среда сети",
        "network_interface": "Интерфейс сети",
        "port_count": "Порты",
        "port_speed": "Скорость портов",
        "port_media": "Среда портов",
        "uplink_count": "Uplinks",
        "uplink_speed": "Скорость uplink",
        "uplink_media": "Среда uplink",
        "poe_supported": "PoE",
        "poe_budget_w": "PoE budget, W",
        "poe_standard": "PoE стандарт",
        "l2_supported": "L2",
        "l3_supported": "L3",
        "stacking_supported": "Stacking",
        "airflow": "Airflow",
        "redundant_psu": "Redundant PSU",
        "transceiver_form_factor": "Форм-фактор трансивера",
        "form_factor_hints": "Форм-фактор",
    }
    for key, label in labels.items():
        fact = value.get(key)
        if fact in (None, "", "unknown", []):
            continue
        if isinstance(fact, list):
            fact_text = ", ".join(str(item) for item in fact if str(item))
        elif isinstance(fact, bool):
            fact_text = yes_no(fact)
        elif key == "ram_capacity_gb":
            fact_text = f"{fact} ГБ"
        else:
            fact_text = str(fact)
        if fact_text:
            rows.append(f"{label}: {fact_text}")
    return "\n".join(rows)


def _recommendation_evidence_text(recommendation: Mapping[str, Any]) -> str:
    summary = recommendation.get("evidence_summary")
    if not isinstance(summary, Mapping):
        return ""
    source_count = _int_value(summary.get("sources_count")) or 0
    if source_count <= 0:
        status = str(summary.get("status") or "").strip()
        if status == "disabled":
            return "Внешние источники не использовались"
        if status in {"not_found", "not_confirmed", "error"}:
            return "\n".join(
                [
                    "Источники: 0",
                    "Не подтверждено: внешние источники не дали подтверждения.",
                    "Финальная проверка инженером обязательна.",
                ]
            )
        return ""
    missing = [
        *as_string_list(summary.get("missing")),
        *as_string_list(summary.get("not_confirmed")),
    ]
    fatal = as_string_list(summary.get("fatal_concerns"))
    relation_status = str(summary.get("status") or "").strip()
    status_label = "Не подтверждено" if missing or fatal else "Подтверждено"
    if relation_status == "partially_confirmed":
        status_label = "Частично подтверждено"
    elif relation_status == "mismatch":
        status_label = "Не подтверждено"
    status_text = str(summary.get("status_text") or "").strip()
    if not status_text:
        status_text = (
            "часть совместимости не подтверждена источниками"
            if missing or fatal
            else "совместимость подтверждена найденными источниками"
        )
    lines = [
        f"{status_label}: {status_text}",
        f"Уверенность: {summary.get('confidence') or 'unknown'}",
        f"Источники: {source_count}",
    ]
    source_domains = as_string_list(summary.get("source_domains"))
    if source_domains:
        lines.append(
            "Домены источников: "
            + ", ".join(_safe_user_text(value) for value in source_domains[:4])
        )
    confirmed = as_string_list(summary.get("confirmed")) or as_string_list(
        summary.get("confirmed_facts")
    )
    if confirmed:
        lines.append(
            "Подтверждено: " + "; ".join(_safe_user_text(value) for value in confirmed[:4])
        )
    if missing:
        lines.append(
            "Не подтверждено: " + "; ".join(_safe_user_text(value) for value in missing[:4])
        )
    checks = as_string_list(summary.get("engineering_checks"))
    if checks:
        lines.append(
            "Проверить инженеру: " + "; ".join(_safe_user_text(value) for value in checks[:4])
        )
    return "\n".join(line for line in lines if line)


def _matrix_evidence_status_text(row: Mapping[str, Any]) -> str:
    evidence = row.get("evidence")
    if not isinstance(evidence, Mapping):
        return ""
    status = str(evidence.get("evidence_status") or row.get("evidence_status") or "").strip()
    if status == "found":
        return "найдено"
    if status == "not_found":
        return "не найдено"
    if status == "disabled":
        return "выключено"
    if status == "error":
        return "ошибка"
    return status


def _matrix_evidence_confidence_text(row: Mapping[str, Any]) -> str:
    return str(row.get("evidence_confidence") or "").strip()


def _matrix_evidence_sources_text(row: Mapping[str, Any]) -> str:
    evidence = row.get("evidence")
    if not isinstance(evidence, Mapping):
        return ""
    sources = _component_rows(evidence.get("sources"))
    if not sources:
        return ""
    domains = _unique_text(
        [str(source.get("domain") or "").strip() for source in sources if source.get("domain")]
    )
    count = len(sources)
    domain_text = ", ".join(domains[:3])
    return f"{count}: {domain_text}" if domain_text else str(count)


def _matrix_evidence_confirmed_text(row: Mapping[str, Any]) -> str:
    return "\n".join(as_string_list(row.get("evidence_confirmed")))


def _matrix_evidence_missing_text(row: Mapping[str, Any]) -> str:
    return "\n".join(as_string_list(row.get("evidence_missing")))


def _evidence_confirmed_texts(evidence: Mapping[str, Any]) -> list[str]:
    facts = evidence.get("facts")
    if not isinstance(facts, Mapping):
        return []
    labels = {
        "vendor": "производитель",
        "platform_family": "семейство платформы",
        "cpu_generation": "поколение CPU",
        "socket_family": "socket",
        "supported_cpu_generation": "поддерживаемое поколение CPU",
        "memory_type": "тип памяти",
        "dimm_slots": "DIMM-слоты",
        "drive_bays": "отсеки",
        "nvme_support": "NVMe",
        "form_factor": "форм-фактор",
        "psu_info": "PSU",
        "storage_interface": "интерфейс",
        "capacity": "емкость",
    }
    rows: list[str] = []
    for key, label in labels.items():
        value = facts.get(key)
        if value in (None, "", "unknown", []):
            continue
        rows.append(f"{label}: {value}")
    return rows


def _evidence_missing_texts(evidence: Mapping[str, Any]) -> list[str]:
    role = str(evidence.get("role") or "").strip()
    facts = evidence.get("facts")
    if not isinstance(facts, Mapping):
        return ["источник не найден"]
    expected = {
        "server_platform": ["supported_cpu_generation", "socket_family", "memory_type"],
        "platform": ["supported_cpu_generation", "socket_family", "memory_type"],
        "cpu": ["cpu_generation", "socket_family"],
        "ram": ["memory_type", "capacity"],
        "ssd": ["storage_interface", "capacity"],
        "hdd": ["storage_interface", "capacity"],
        "storage": ["storage_interface", "capacity"],
        "ready_server": ["platform_family"],
    }.get(role, [])
    labels = {
        "supported_cpu_generation": "поддержка CPU",
        "socket_family": "socket",
        "memory_type": "тип памяти",
        "cpu_generation": "поколение CPU",
        "capacity": "емкость",
        "storage_interface": "интерфейс",
        "platform_family": "семейство платформы",
    }
    return [labels.get(key, key) for key in expected if facts.get(key) in (None, "", "unknown", [])]


def _fit_label_text(value: Any) -> str:
    labels = {
        "exact_or_close_fit": "Минимально подходит / близко к требованию",
        "acceptable_overfit": "Выше требования, коммерчески допустимо",
        "excessive_overfit": "Существенно выше требования, нужна проверка альтернатив",
        "unknown_fit": "Требуется ручная проверка",
    }
    return labels.get(str(value or "").strip(), "")


def _matrix_fit_text(row: Mapping[str, Any]) -> str:
    return "\n".join(
        _unique_text(
            [
                _fit_label_text(row.get("fit_label")),
                _fit_tier_text(row.get("fit_tier")),
            ]
        )
    )


def _fit_tier_text(value: Any) -> str:
    labels = {
        "strong_fit": "Сильное соответствие",
        "possible_fit": "Возможное соответствие",
        "fallback_unknown": "Нужна проверка по данным товара",
        "explicit_mismatch": "Не закрывает явное требование",
        "wrong_role": "Не та роль",
    }
    return labels.get(str(value or "").strip(), "")


def _ram_order_text(components: list[Mapping[str, Any]]) -> str:
    for component in components:
        if component.get("role") != "ram":
            continue
        quantity = _int_value(component.get("quantity_required"))
        facts = component.get("facts")
        capacity = None
        if isinstance(facts, Mapping):
            capacity = _int_value(facts.get("ram_capacity_gb"))
        if quantity is not None and capacity is not None:
            return f"{quantity * capacity} ГБ"
        if quantity is not None:
            return f"{quantity} модулей"
    return ""


def _sorted_candidates(match_run: MatchRun) -> list[MatchCandidate]:
    return sorted(
        match_run.candidates,
        key=lambda candidate: (
            candidate.confidence_score,
            candidate.available_quantity or 0,
            candidate.reservable_locations,
        ),
        reverse=True,
    )


def _ready_candidates(match_run: MatchRun) -> list[MatchCandidate]:
    return [
        candidate
        for candidate in _sorted_candidates(match_run)
        if _candidate_type(candidate) == "ready_server"
    ]


def _build_candidates(match_run: MatchRun) -> list[MatchCandidate]:
    return [
        candidate
        for candidate in _sorted_candidates(match_run)
        if _candidate_type(candidate) == "build_from_parts"
    ]


def _llm_recommended_builds(match_run: MatchRun) -> list[Mapping[str, Any]]:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    return _component_rows(report_json.get("llm_recommended_build_candidates"))


def _llm_recommendations_empty_message(match_run: MatchRun) -> str:
    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    if not report_json.get("llm_configurator_enabled"):
        return "LLM Composer не использовался: слой выключен."
    reason = report_json.get("no_recommendation_reason")
    if isinstance(reason, Mapping) and reason:
        lines = ["Безопасную складскую рекомендацию дать нельзя."]
        partial_available_components = _component_rows(
            reason.get("partial_available_components")
        )
        failed_requirements = _excel_reason_items(reason.get("failed_requirements"))
        role_failures = _component_rows(reason.get("role_failures"))
        unverified_requirements = _excel_reason_items(
            reason.get("unverified_requirements")
        )
        hard_mismatch_risks = _excel_reason_items(reason.get("hard_mismatch_risks"))
        recommended_next_actions = as_string_list(
            reason.get("recommended_next_actions")
        )
        engineer_checks = as_string_list(
            reason.get("engineer_checks") or reason.get("engineering_checks")
        )
        if partial_available_components:
            lines.append(
                "Что можно закрыть со склада: "
                + "; ".join(
                    _excel_reason_item_text(item)
                    for item in partial_available_components[:5]
                )
                + "."
            )
        if failed_requirements:
            lines.append(
                "Что не закрывается: "
                + "; ".join(
                    _excel_reason_item_text(item) for item in failed_requirements[:6]
                )
                + "."
            )
        if role_failures:
            lines.append(
                "Проблемы по ролям: "
                + "; ".join(
                    _excel_role_failure_text(item) for item in role_failures[:6]
                )
                + "."
            )
        missing_roles = as_string_list(reason.get("missing_roles"))
        if missing_roles and not role_failures:
            lines.append("Не хватает ролей: " + ", ".join(missing_roles) + ".")
        stock_shortages = _component_rows(reason.get("stock_shortages"))
        if stock_shortages:
            lines.append("Есть нехватка склада по обязательным компонентам.")
        hard_incompatibility = as_string_list(reason.get("hard_incompatibility"))
        if hard_mismatch_risks:
            lines.append(
                "Почему BOM нельзя показать как КП-ready: "
                + "; ".join(
                    _excel_reason_item_text(item) for item in hard_mismatch_risks[:6]
                )
                + "."
            )
        elif hard_incompatibility:
            details = ", ".join(hard_incompatibility[:3])
            lines.append(f"Есть жесткая несовместимость: {details}.")
        if unverified_requirements:
            lines.append(
                "Что осталось неподтвержденным: "
                + "; ".join(
                    _excel_reason_item_text(item)
                    for item in unverified_requirements[:5]
                )
                + "."
            )
        if recommended_next_actions:
            lines.append(
                "Что нужно дозакупить / проверить: "
                + "; ".join(_safe_user_text(item) for item in recommended_next_actions[:5])
                + "."
            )
        if engineer_checks:
            lines.append(
                "Инженерно проверить: "
                + "; ".join(_safe_user_text(item) for item in engineer_checks[:5])
                + "."
            )
        coverage_line = _excel_no_recommendation_coverage_line(reason)
        if coverage_line:
            lines.append(coverage_line)
        lines.append("См. лист Матрица компонентов для инженерной проверки.")
        return "\n".join(lines)
    reason = str(report_json.get("llm_fallback_reason") or "").strip()
    if reason:
        return (
            "Безопасную складскую рекомендацию дать нельзя.\n"
            f"Причина: {reason}.\n"
            "См. лист Матрица компонентов для инженерной проверки."
        )
    return (
        "Безопасную складскую рекомендацию дать нельзя.\n"
        "См. лист Матрица компонентов для инженерной проверки."
    )


def _excel_reason_items(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, Mapping):
        return [value]
    if not isinstance(value, list):
        text = _safe_user_text(value)
        return [text] if text else []
    result: list[Any] = []
    for item in value:
        if isinstance(item, Mapping):
            result.append(item)
            continue
        text = _safe_user_text(item)
        if text:
            result.append(text)
    return result


def _excel_reason_item_text(item: Any) -> str:
    if not isinstance(item, Mapping):
        return _safe_user_text(item)
    role = _safe_user_text(item.get("role") or item.get("component_role"))
    text = _safe_user_text(
        item.get("source_text")
        or item.get("requirement_text")
        or item.get("requirement")
        or item.get("message")
        or item.get("reason")
        or item.get("user_message")
        or item.get("status")
        or item.get("type")
    )
    action = _safe_user_text(
        item.get("suggested_action") or item.get("recommended_action")
    )
    return ": ".join(part for part in (role, text, action) if part) or "требуется проверка"


def _excel_role_failure_text(item: Mapping[str, Any]) -> str:
    base = _excel_reason_item_text(item)
    coverage = item.get("candidate_coverage")
    if not isinstance(coverage, Mapping):
        return base
    considered = coverage.get("considered_count")
    total = coverage.get("candidate_count")
    if considered is None and isinstance(coverage.get("considered_candidate_ids"), list):
        considered = len(coverage["considered_candidate_ids"])
    if considered is None and total is None:
        return base
    if total is None:
        return f"{base} (рассмотрено кандидатов: {considered})"
    return f"{base} (coverage: {considered}/{total})"


def _excel_no_recommendation_coverage_line(reason: Mapping[str, Any]) -> str:
    coverage = reason.get("no_recommendation_coverage")
    if isinstance(coverage, Mapping):
        percent_by_role = coverage.get("coverage_percent_by_role")
        if isinstance(percent_by_role, Mapping) and percent_by_role:
            parts = [
                f"{role}: {round(float(value), 1)}%"
                for role, value in list(percent_by_role.items())[:4]
                if isinstance(value, (int, float))
            ]
            if parts:
                return "Coverage по матрице: " + ", ".join(parts) + "."
    role_coverage = reason.get("role_evaluation_coverage_by_role")
    if isinstance(role_coverage, Mapping) and role_coverage:
        covered = 0
        total = 0
        for value in role_coverage.values():
            if not isinstance(value, Mapping):
                continue
            total += 1
            if value.get("all_candidates_considered") is True:
                covered += 1
        if total:
            return f"Coverage role evaluation: {covered}/{total} ролей полностью рассмотрены."
    return ""


def _candidate_mapping(candidate: MatchCandidate) -> dict[str, Any]:
    raw = _raw_mapping(candidate)
    return {
        "candidate_type": _candidate_type(candidate),
        "distributor_code": candidate.distributor_code,
        "item_id": candidate.item_id,
        "product_key": candidate.product_key,
        "part_number": candidate.part_number,
        "producer": candidate.producer,
        "category_id": candidate.category_id,
        "item_name": candidate.item_name,
        "confidence_score": candidate.confidence_score,
        "price_value": candidate.price_value,
        "price_currency": candidate.price_currency,
        "available_quantity": candidate.available_quantity,
        "matched_requirements": candidate.matched_requirements_json,
        "missing_requirements": candidate.missing_requirements_json,
        "risk_flags": candidate.risk_flags_json,
        "components": raw.get("components") if isinstance(raw.get("components"), list) else [],
        "total_price_value": raw.get("total_price_value"),
        "total_price_currency": raw.get("total_price_currency"),
        "missing_components": raw.get("missing_components"),
        "compatibility_warnings": raw.get("compatibility_warnings"),
        "completeness_status": raw.get("completeness_status"),
        "completeness_label": raw.get("completeness_label"),
        "included_component_roles": raw.get("included_component_roles"),
        "missing_component_roles": raw.get("missing_component_roles"),
        "excluded_from_total_roles": raw.get("excluded_from_total_roles"),
        "cpu_per_server": raw.get("cpu_per_server"),
        "total_cpu_required": raw.get("total_cpu_required"),
        "total_price_note": raw.get("total_price_note"),
    }


def _candidate_type(candidate: MatchCandidate) -> str:
    raw = _raw_mapping(candidate)
    candidate_type = raw.get("candidate_type")
    if isinstance(candidate_type, str) and candidate_type:
        return candidate_type
    return "ready_server"


def _raw_mapping(candidate: MatchCandidate) -> Mapping[str, Any]:
    if isinstance(candidate.raw_json, Mapping):
        return candidate.raw_json
    return {}


def _component_rows(value: Any) -> list[Mapping[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, Mapping)]


def _components_text(components: list[Mapping[str, Any]], role: str) -> str:
    values: list[str] = []
    for component in components:
        if component.get("role") != role:
            continue
        values.append(_component_text(component))
    return "\n".join(values)


def _storage_components_text(components: list[Mapping[str, Any]]) -> str:
    return "\n".join(
        value
        for value in [
            _components_text(components, "ssd"),
            _components_text(components, "hdd"),
        ]
        if value
    )


def _component_text(component: Mapping[str, Any]) -> str:
    parts = [
        str(component.get("producer") or "").strip(),
        str(component.get("part_number") or "").strip(),
    ]
    display = " ".join(part for part in parts if part)
    if not display:
        display = str(component.get("item_name") or component.get("item_id") or "").strip()
    quantity = component.get("quantity_required")
    available = component.get("available_quantity")
    available_text = "не найден" if available is None else str(available)
    return f"{display}; требуется {quantity} шт.; остаток {available_text}"


def _stock_summary(components: list[Mapping[str, Any]]) -> str:
    rows: list[str] = []
    for component in components:
        role = component.get("role_ru") or component.get("role") or "Компонент"
        quantity = component.get("quantity_required")
        available = component.get("available_quantity")
        available_text = "не найден" if available is None else str(available)
        rows.append(f"{role}: требуется {quantity} шт., остаток {available_text}")
    return "\n".join(rows)


def _build_result_text(raw: Mapping[str, Any]) -> str:
    status = str(raw.get("completeness_status") or "").strip()
    if status == "incomplete":
        return "Неполная сборка"
    if status == "complete":
        return "Предварительная сборка"
    return "Сборка из комплектующих"


def _build_completeness_text(raw: Mapping[str, Any]) -> str:
    label = str(raw.get("completeness_label") or "").strip()
    if label:
        return label
    status = str(raw.get("completeness_status") or "").strip()
    if status == "incomplete":
        return "Неполная сборка"
    if status == "complete":
        return "Компоненты подобраны, требуется инженерная проверка"
    return "Требуется инженерная проверка"


def _cpu_components_text(candidate: MatchCandidate, components: list[Mapping[str, Any]]) -> str:
    cpu = _components_text(components, "cpu")
    if cpu:
        return cpu
    raw = _raw_mapping(candidate)
    if (
        "cpu" in as_string_list(raw.get("missing_component_roles"))
        or raw.get("cpu_per_server")
        or raw.get("total_cpu_required")
    ):
        return "не подобраны"
    return "не указано"


def _server_quantity_from_components(components: list[Mapping[str, Any]]) -> Any:
    for component in components:
        if component.get("role") == "server_platform":
            return component.get("quantity_required") or ""
    return ""


def _excluded_from_total_text(raw: Mapping[str, Any]) -> str:
    roles = as_string_list(raw.get("excluded_from_total_roles"))
    if roles:
        optional_roles = set(
            as_string_list(raw.get("optional_component_roles"))
            or as_string_list(raw.get("engineer_check_component_roles"))
        )
        required_roles = [role for role in roles if role not in optional_roles]
        parts: list[str] = []
        if required_roles:
            parts.append(", ".join(_excluded_role_label(role) for role in required_roles))
        if optional_roles:
            parts.append(
                "опционально: "
                + ", ".join(
                    _excluded_role_label(role)
                    for role in roles
                    if role in optional_roles
                )
            )
        return "; ".join(parts)

    note = str(raw.get("total_price_note") or "").strip()
    if note.casefold().startswith("без "):
        return note[4:]
    return note


def _excluded_role_label(role: str) -> str:
    labels = {
        "switch": "коммутаторы",
        "router": "маршрутизаторы",
        "firewall": "межсетевые экраны",
        "access_point": "точки доступа",
        "cpu": "CPU",
        "ram": "RAM",
        "ssd": "SSD",
        "hdd": "HDD",
        "storage_controller": "контроллеры",
        "network_adapter": "сетевые адаптеры",
        "transceiver": "трансиверы",
        "dac_cable": "DAC-кабели",
        "cable": "кабели",
        "license": "лицензии",
        "support": "поддержка",
        "power_supply": "блоки питания",
        "stacking_module": "модули стекирования",
    }
    return labels.get(role, role)


def _price_value(candidate: MatchCandidate) -> int | float | None:
    return _money_value(candidate.price_value)


def _money_value(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    try:
        amount = value if isinstance(value, Decimal) else Decimal(str(value).replace(" ", ""))
    except (InvalidOperation, ValueError):
        return None
    if amount == amount.to_integral():
        return int(amount)
    return float(amount)


def _int_value(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _amount_header_for_request(match_run: MatchRun) -> str:
    server_quantity = _server_quantity_for_request(match_run)
    if server_quantity is None:
        return "Ориентировочная сумма за весь запрос"
    return f"Ориентировочная сумма за {_server_quantity_text(server_quantity)}"


def _ai_amount_header_for_request(match_run: MatchRun) -> str:
    server_quantity = _server_quantity_for_request(match_run)
    if server_quantity is None:
        return "Ориентировочная сумма за весь запрос"
    return f"Ориентировочная сумма за {_server_quantity_text(server_quantity)} или за весь запрос"


def _server_quantity_for_request(match_run: MatchRun) -> int | None:
    spec_json = match_run.spec_json if isinstance(match_run.spec_json, Mapping) else {}
    quantity = _server_quantity_from_spec(spec_json)
    if quantity is not None:
        return quantity

    report_json = match_run.report_json if isinstance(match_run.report_json, Mapping) else {}
    for key in ("llm_recommended_build_candidates", "build_candidates"):
        for candidate in _component_rows(report_json.get(key)):
            quantity = _server_quantity_from_report_candidate(candidate)
            if quantity is not None:
                return quantity

    for candidate in _build_candidates(match_run):
        quantity = _server_quantity_from_report_candidate(_raw_mapping(candidate))
        if quantity is not None:
            return quantity
    return None


def _server_quantity_from_spec(spec_json: Mapping[str, Any]) -> int | None:
    items = spec_json.get("items")
    if not isinstance(items, list):
        return None

    quantities: list[int] = []
    for item in items:
        if not isinstance(item, Mapping):
            continue
        item_text = " ".join(
            str(item.get(key) or "").casefold()
            for key in ("item_type", "name")
        )
        if item_text and "server" not in item_text and "сервер" not in item_text:
            continue
        quantity = _int_value(item.get("quantity"))
        if quantity is not None and quantity > 0:
            quantities.append(quantity)
    return sum(quantities) if quantities else None


def _server_quantity_from_report_candidate(candidate: Mapping[str, Any]) -> int | None:
    quantity = _int_value(candidate.get("quantity_required"))
    if quantity is not None and quantity > 0:
        return quantity
    return _int_value(
        _server_quantity_from_components(_component_rows(candidate.get("components")))
    )


def _server_quantity_text(value: int) -> str:
    if value % 10 == 1 and value % 100 != 11:
        word = "сервер"
    elif value % 10 in {2, 3, 4} and value % 100 not in {12, 13, 14}:
        word = "сервера"
    else:
        word = "серверов"
    return f"{value} {word}"


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S %Z").strip()


def _safe_user_text(value: Any) -> str:
    return sanitize_user_facing_text(value)


def _filter_contradictory_checks(
    checks: list[str],
    requirements: Mapping[str, Any],
) -> list[str]:
    result: list[str] = []
    for check in checks:
        if _is_contradictory_check(check, requirements):
            continue
        if check not in result:
            result.append(check)
    return result


def _is_contradictory_check(
    check: str,
    requirements: Mapping[str, Any],
) -> bool:
    lowered = check.casefold()
    ram_type = str(requirements.get("ram_type_preference") or "").strip()
    if ram_type and ram_type != "unknown" and "тип оперативной памяти не указан" in lowered:
        return True
    if requirements.get("ram_gb_per_server") not in (None, "", "unknown") and (
        "объем оперативной памяти не указан" in lowered
        or "требуемый объем оперативной памяти" in lowered
    ):
        return True
    storage_interface = str(requirements.get("storage_interface_preference") or "").strip()
    if (
        storage_interface
        and storage_interface != "unknown"
        and "интерфейс накопителя не указан" in lowered
    ):
        return True
    cpu_vendor = str(requirements.get("cpu_vendor_preference") or "").strip()
    return bool(cpu_vendor and cpu_vendor != "unknown" and (
        "vendor cpu не указан" in lowered
        or "производитель cpu не указан" in lowered
    ))


def _format_requirement_value(value: Any) -> str:
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, list):
        return ", ".join(_format_requirement_value(item) for item in value)
    if isinstance(value, Mapping):
        return ", ".join(
            f"{key}: {_format_requirement_value(item)}" for key, item in value.items()
        )
    return str(value)
