from __future__ import annotations

import json
import re
from collections.abc import Mapping, Sequence
from copy import copy
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.db.models import MatchRun
from app.user_facing_text import sanitize_user_facing_text

V3_FULL_CATEGORY_PIPELINE_VERSION = "v3_full_category_matrix"
SIMPLE_STOCK_QUOTE_PIPELINE_VERSION = "simple_stock_quote"

TITLE_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
BOLD_FONT = Font(bold=True)
MUTED_FONT = Font(color="667085")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="EAF3FA")
REVIEW_FILL = PatternFill("solid", fgColor="FFF4D6")
NOTE_FILL = PatternFill("solid", fgColor="F6F8FA")
GRID_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(bottom=GRID_SIDE)
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
COMMERCIAL_COLUMN_COUNT = 8


def is_v3_full_category_report(report_json: Mapping[str, Any]) -> bool:
    return str(report_json.get("pipeline_version") or "") in {
        V3_FULL_CATEGORY_PIPELINE_VERSION,
        SIMPLE_STOCK_QUOTE_PIPELINE_VERSION,
    }


def build_v3_full_category_markdown_report(
    report_json: Mapping[str, Any],
    *,
    match_run_id: int | None = None,
) -> str:
    diagnostics = _mapping(report_json.get("diagnostics"))
    quote = _mapping(report_json.get("validated_quote"))
    no_recommendation = _mapping(report_json.get("no_recommendation_reason"))
    validation_failure = _mapping(report_json.get("validation_failure_reason"))
    result_state = report_json.get("v3_result_state") or report_json.get(
        "final_status_source"
    )
    lines = [
        "# КП draft",
        "",
        f"- ID результата: {match_run_id or report_json.get('match_run_id') or ''}".rstrip(),
        (
            "- Статус: "
            f"{_state_label(result_state)}"
        ),
        f"- Профиль: {_profile_label(report_json.get('v3_profile') or 'custom')}",
        f"- Склад: {_distributor_label(report_json.get('distributor_code'))}",
        f"- Категории: {_category_summary(report_json.get('category_ids'))}",
        f"- Строк матрицы: {diagnostics.get('matrix_row_count') or 0}",
        "",
    ]

    if quote:
        total_lines = _total_price_lines(quote)
        if total_lines:
            lines.extend(["## Итог", "", *total_lines, ""])
        _append_text_block(lines, "Класс подбора", _selection_summary(quote))
        _append_text_block(lines, "Тип предложения", quote.get("client_status_label"))
        _append_text_block(lines, "Кратко для менеджера", quote.get("client_summary"))
        _append_text_block(lines, "Покрытие ТЗ", quote.get("coverage_summary"))
        _append_list_block(lines, "Целевые объекты", quote.get("target_decisions"))
        quote_lines = _group_quote_lines(_list_of_mappings(quote.get("lines")))
        if quote_lines:
            lines.extend(["## Спецификация для КП", ""])
            for index, item in enumerate(quote_lines, start=1):
                line_total = _price_text(
                    item.get("line_total_value"),
                    item.get("line_total_currency"),
                )
                title = _line_title(item)
                lines.append(
                    f"{index}. {_role_label(item.get('role'))}: {title}, "
                    f"кол-во {item.get('quantity') or 1}, сумма {line_total}".rstrip()
                )
                comment = _line_comment(item)
                if comment:
                    lines.append(f"   - Примечание: {comment}")
            lines.append("")
        _append_text_block(lines, "Комментарий", quote.get("why_selected"))
        _append_list_block(lines, "Покрытие требований", quote.get("requirement_coverage"))
        _append_list_block(lines, "Ключевые отличия", quote.get("key_deviations"))
        _append_list_block(lines, "Отклонения от ТЗ", quote.get("deviation_notes"))
        _append_list_block(
            lines,
            "Что нужно добрать или согласовать",
            quote.get("procurement_gaps"),
        )
        _append_available_alternatives_block(
            lines,
            "Доступные варианты для согласования - не включены в итог",
            quote.get("available_alternatives"),
        )
        _append_list_block(lines, "Проверка цены", quote.get("price_audit"))
        _append_list_block(lines, "Допущения", quote.get("assumptions"))
        _append_stock_confirmation_block(
            lines,
            "Подтвердить количество по складу",
            quote,
        )
        compatibility = _mapping(quote.get("compatibility_check"))
        if compatibility:
            lines.extend(["## Проверка совместимости", ""])
            lines.append(f"- Статус: {_compatibility_status_label(compatibility.get('status'))}")
            _append_list_items(lines, "Подтверждено", compatibility.get("checked_facts"))
            _append_list_items(lines, "Блокер", compatibility.get("blocking_mismatches"))
            _append_list_items(lines, "Риск", compatibility.get("unresolved_risks"))
            lines.append("")
        _append_list_block(lines, "Перед отправкой проверить", quote.get("engineer_checks"))
    elif validation_failure:
        lines.extend(["## Ошибка проверки ответа модели", ""])
        _append_text_block(
            lines,
            "Причина",
            validation_failure.get("summary"),
        )
        _append_list_block(
            lines,
            "Что не сошлось",
            validation_failure.get("next_actions")
            or [
                "Повторить запрос после обновления склада.",
                "Проверить технические детали на листе диагностики.",
            ],
        )
    elif no_recommendation:
        lines.extend(["## Нет рекомендации", ""])
        _append_text_block(lines, "Причина", _no_recommendation_summary(no_recommendation))
        _append_list_block(
            lines,
            "Что не закрыто",
            _no_recommendation_failed_items(no_recommendation),
        )
        _append_list_block(
            lines,
            "Что сделать дальше",
            _no_recommendation_next_actions(no_recommendation),
        )

    validation_errors = _string_list(report_json.get("v3_validation_errors"))
    if validation_errors:
        lines.extend(["## Ошибки механической проверки", ""])
        lines.extend(f"- {item}" for item in validation_errors[:20])
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def build_v3_full_category_excel_report(match_run: MatchRun) -> bytes:
    report_json = _mapping(match_run.report_json)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "КП"
    review_sheet = workbook.create_sheet("Инженерная проверка")
    diagnostics_sheet = workbook.create_sheet("Склад и диагностика")

    _fill_v3_summary_sheet(summary_sheet, match_run, report_json)
    _fill_v3_review_sheet(review_sheet, match_run, report_json)
    _fill_v3_diagnostics_sheet(diagnostics_sheet, report_json)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fill_v3_summary_sheet(
    sheet: Worksheet,
    match_run: MatchRun,
    report_json: Mapping[str, Any],
) -> None:
    sheet.sheet_view.showGridLines = False
    diagnostics = _mapping(report_json.get("diagnostics"))
    quote = _mapping(report_json.get("validated_quote"))
    no_recommendation = _mapping(report_json.get("no_recommendation_reason"))
    validation_failure = _mapping(report_json.get("validation_failure_reason"))

    row = _write_commercial_header(sheet, match_run, report_json, diagnostics)

    if quote:
        row = _write_commercial_total(
            sheet,
            row,
            _total_price_lines(quote),
            match_run.engineer_review_required,
        )
        row = _write_commercial_text(sheet, row, "Класс подбора", _selection_summary(quote))
        row = _write_commercial_text(
            sheet,
            row,
            "Тип предложения",
            quote.get("client_status_label"),
        )
        row = _write_commercial_text(
            sheet,
            row,
            "Кратко для менеджера",
            quote.get("client_summary"),
        )
        row = _write_commercial_text(sheet, row, "Покрытие ТЗ", quote.get("coverage_summary"))
        row = _write_commercial_target_decisions(sheet, row, quote.get("target_decisions"))
        row = _write_commercial_quote_lines(
            sheet,
            row,
            _group_quote_lines(_list_of_mappings(quote.get("lines"))),
        )
        row = _write_commercial_text(sheet, row, "Комментарий", quote.get("why_selected"))
        row = _write_commercial_list(sheet, row, "Ключевые отличия", quote.get("key_deviations"))
        row = _write_commercial_list(sheet, row, "Отклонения от ТЗ", quote.get("deviation_notes"))
        row = _write_commercial_procurement_gaps(
            sheet,
            row,
            "Что нужно добрать или согласовать",
            quote.get("procurement_gaps"),
        )
        row = _write_commercial_available_alternatives(
            sheet,
            row,
            "Доступные варианты для согласования - не включены в итог",
            quote.get("available_alternatives"),
        )
        row = _write_commercial_list(sheet, row, "Проверка цены", quote.get("price_audit"))
        row = _write_commercial_list(sheet, row, "Допущения для КП", quote.get("assumptions"))
        row = _write_stock_confirmation_table(
            sheet,
            row,
            "Подтвердить количество по складу",
            quote,
            commercial=True,
        )
        row = _write_commercial_list(
            sheet,
            row,
            "Перед отправкой проверить",
            quote.get("engineer_checks"),
            fill=REVIEW_FILL,
        )
    elif validation_failure:
        row = _write_commercial_status_block(
            sheet,
            row,
            "КП не сформировано",
            validation_failure.get("summary"),
        )
        row = _write_commercial_list(
            sheet,
            row,
            "Что не сошлось",
            validation_failure.get("error_details")
            or report_json.get("v3_validation_error_details")
            or report_json.get("v3_validation_errors"),
        )
    elif no_recommendation:
        row = _write_commercial_status_block(
            sheet,
            row,
            "КП не сформировано",
            _no_recommendation_summary(no_recommendation)
            or _no_recommendation_details(no_recommendation),
        )
        row = _write_commercial_list(
            sheet,
            row,
            "Что не закрыто",
            _no_recommendation_failed_items(no_recommendation),
        )
        row = _write_commercial_list(
            sheet,
            row,
            "Что сделать дальше",
            _no_recommendation_next_actions(no_recommendation),
        )

    _write_footer_note(sheet, row)
    _finish_sheet(
        sheet,
        widths={1: 18, 2: 18, 3: 52, 4: 10, 5: 14, 6: 10, 7: 16, 8: 42},
        freeze_panes="A10",
    )
    sheet.print_title_rows = "1:9"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def _fill_v3_review_sheet(
    sheet: Worksheet,
    match_run: MatchRun,
    report_json: Mapping[str, Any],
) -> None:
    sheet.sheet_view.showGridLines = False
    quote = _mapping(report_json.get("validated_quote"))
    no_recommendation = _mapping(report_json.get("no_recommendation_reason"))
    validation_failure = _mapping(report_json.get("validation_failure_reason"))
    row = _write_title(sheet, 1, "Инженерная проверка перед КП")
    row = _write_key_value(
        sheet,
        row,
        "Статус",
        "нужна проверка" if match_run.engineer_review_required else "явных блокеров нет",
    )
    row += 1

    if quote:
        row = _write_list(sheet, row, "Покрытие требований", quote.get("requirement_coverage"))
        row = _write_procurement_gaps(
            sheet,
            row,
            "Что нужно добрать или согласовать",
            quote.get("procurement_gaps"),
        )
        row = _write_available_alternatives(
            sheet,
            row,
            "Доступные варианты для согласования - не включены в итог",
            quote.get("available_alternatives"),
        )
        row = _write_gap_considered_candidates(
            sheet,
            row,
            "Рассмотренные складские кандидаты для gaps",
            quote.get("procurement_gaps"),
        )
        row = _write_stock_confirmation_table(
            sheet,
            row,
            "Подтвердить количество по складу",
            quote,
        )
        row = _write_list(sheet, row, "Проверить инженеру", quote.get("engineer_checks"))
        row = _write_list(sheet, row, "Ключевые отличия", quote.get("key_deviations"))
        row = _write_list(sheet, row, "Отклонения от ТЗ", quote.get("deviation_notes"))
        row = _write_list(sheet, row, "Допущения", quote.get("assumptions"))
        row = _write_compatibility(sheet, row, _mapping(quote.get("compatibility_check")))
        row = _write_list(sheet, row, "Проверка цены", quote.get("price_audit"))
        row = _write_long_text(sheet, row, "Комментарий LLM", quote.get("why_selected"))
    elif validation_failure:
        row = _write_long_text(sheet, row, "Причина", validation_failure.get("summary"))
        row = _write_list(
            sheet,
            row,
            "Что не сошлось",
            validation_failure.get("error_details")
            or report_json.get("v3_validation_error_details")
            or report_json.get("v3_validation_errors"),
        )
    elif no_recommendation:
        row = _write_long_text(sheet, row, "Причина", _no_recommendation_summary(no_recommendation))
        row = _write_list(
            sheet,
            row,
            "Что не закрыто",
            _no_recommendation_failed_items(no_recommendation),
        )
        row = _write_list(
            sheet,
            row,
            "Что сделать дальше",
            _no_recommendation_next_actions(no_recommendation),
        )

    _finish_sheet(
        sheet,
        widths={1: 46, 2: 30, 3: 12, 4: 18, 5: 70},
        freeze_panes="A3",
    )


def _fill_v3_diagnostics_sheet(sheet: Worksheet, report_json: Mapping[str, Any]) -> None:
    sheet.sheet_view.showGridLines = False
    row = _write_title(sheet, 1, "Склад и диагностика")
    quote = _mapping(report_json.get("validated_quote"))
    row = _write_selected_stock_rows(sheet, row, _list_of_mappings(quote.get("lines")))
    row = _write_available_alternatives(
        sheet,
        row,
        "Доступные альтернативы, raw",
        quote.get("available_alternatives"),
    )
    row = _write_gap_considered_candidates(
        sheet,
        row,
        "Рассмотренные кандидаты для gaps, raw",
        quote.get("procurement_gaps"),
    )
    row = _write_quote_integrity_adjustments(
        sheet,
        row,
        _mapping(quote.get("quote_integrity")),
    )
    row = _write_list(sheet, row, "Dominance audit", quote.get("dominance_audit"))
    diagnostics = _mapping(report_json.get("diagnostics"))
    row = _write_section(sheet, row, "Диагностика")
    for key in sorted(diagnostics):
        row = _write_key_value(sheet, row, key, _json_preview(diagnostics.get(key)))
    row += 1
    row = _write_list(
        sheet,
        row,
        "Ошибки механической проверки",
        report_json.get("v3_validation_errors"),
    )
    row = _write_list(
        sheet,
        row,
        "Детали ошибок проверки",
        report_json.get("v3_validation_error_details"),
    )
    row = _write_list(sheet, row, "Предупреждения", report_json.get("v3_validation_warnings"))
    row = _write_long_text(sheet, row, "Разбор запроса", report_json.get("resolved_request"))
    _finish_sheet(sheet, widths={1: 32, 2: 100, 3: 36, 4: 36, 5: 18, 6: 36, 7: 36})


def _write_commercial_header(
    sheet: Worksheet,
    match_run: MatchRun,
    report_json: Mapping[str, Any],
    diagnostics: Mapping[str, Any],
) -> int:
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    title = sheet.cell(row=1, column=1, value="Коммерческое предложение - черновик")
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    subtitle = (
        "Сформировано по складской матрице. Перед отправкой клиенту нужна финальная "
        "проверка совместимости и условий поставки."
    )
    sheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    subtitle_cell = sheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = MUTED_FONT
    subtitle_cell.alignment = WRAP
    sheet.row_dimensions[2].height = 34

    distributor = _distributor_label(report_json.get("distributor_code"))
    status = report_json.get("v3_result_state") or match_run.status
    quote = _mapping(report_json.get("validated_quote"))
    metadata = [
        ("Склад", distributor),
        ("Статус", _state_label(status)),
        ("Тип", _status_badge_text(quote) if quote else "КП не сформировано"),
        (
            "Проверка",
            "нужна инженерная проверка"
            if match_run.engineer_review_required
            else "явных блокеров нет",
        ),
    ]
    row = 4
    for index, (label, value) in enumerate(metadata, start=1):
        cell = sheet.cell(row=row, column=index, value=label)
        cell.fill = NOTE_FILL
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = GRID_BORDER
        value_cell = sheet.cell(row=row + 1, column=index, value=_cell_value(value))
        value_cell.alignment = CENTER
        value_cell.border = GRID_BORDER
    sheet.row_dimensions[row].height = 22
    sheet.row_dimensions[row + 1].height = 42
    return row + 3


def _write_commercial_total(
    sheet: Worksheet,
    row: int,
    totals: Sequence[str],
    engineer_review_required: bool,
) -> int:
    total_lines = [total for total in totals if total] or [""]
    for index, total in enumerate(total_lines):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        label = sheet.cell(
            row=row,
            column=1,
            value="Итого по включённым складским позициям" if index == 0 else "",
        )
        label.fill = TOTAL_FILL
        label.font = Font(size=13, bold=True)
        label.border = GRID_BORDER
        sheet.merge_cells(
            start_row=row,
            start_column=5,
            end_row=row,
            end_column=COMMERCIAL_COLUMN_COUNT,
        )
        value = sheet.cell(row=row, column=5, value=total)
        value.fill = TOTAL_FILL
        value.font = Font(size=13, bold=True)
        value.alignment = Alignment(horizontal="right", vertical="center")
        value.border = GRID_BORDER
        sheet.row_dimensions[row].height = 28
        row += 1
    if engineer_review_required:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=COMMERCIAL_COLUMN_COUNT,
        )
        note = sheet.cell(
            row=row,
            column=1,
            value="Статус: черновик КП. Перед отправкой клиенту нужна инженерная проверка.",
        )
        note.fill = REVIEW_FILL
        note.font = BOLD_FONT
        note.alignment = WRAP
        note.border = GRID_BORDER
        sheet.row_dimensions[row].height = 30
        row += 1
    return row + 1


def _write_commercial_quote_lines(
    sheet: Worksheet,
    row: int,
    quote_lines: list[Mapping[str, Any]],
) -> int:
    if not quote_lines:
        return row
    row = _write_section(
        sheet,
        row,
        "Спецификация для КП",
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    headers = [
        "Артикул",
        "Производитель",
        "Наименование",
        "Кол-во",
        "Цена",
        "Валюта",
        "Сумма",
        "Комментарий",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = GRID_BORDER
    row += 1
    for item in quote_lines:
        values = [
            _commercial_part_number(item),
            _commercial_producer(item),
            _commercial_name(item),
            item.get("quantity"),
            _amount_text(item.get("unit_price_value")),
            _text(item.get("unit_price_currency")),
            _amount_text(item.get("line_total_value")),
            _line_comment(item),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value))
            cell.alignment = CENTER if column in {4, 5, 6, 7} else WRAP
            cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 42
        row += 1
    return row + 1


def _write_commercial_text(sheet: Worksheet, row: int, title: str, value: Any) -> int:
    text = _text(value)
    if not text:
        return row
    return _write_commercial_status_block(sheet, row, title, text)


def _write_commercial_list(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
    *,
    fill: PatternFill = NOTE_FILL,
) -> int:
    items = _string_list(value)
    if not items:
        return row
    row = _write_section(
        sheet,
        row,
        title,
        fill=fill,
        font=BOLD_FONT,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    for item in items:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=COMMERCIAL_COLUMN_COUNT,
        )
        cell = sheet.cell(row=row, column=1, value=f"- {item}")
        cell.alignment = WRAP
        cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 32
        row += 1
    return row + 1


def _write_commercial_target_decisions(sheet: Worksheet, row: int, value: Any) -> int:
    decisions = _list_of_mappings(value)
    if not decisions:
        return row
    row = _write_section(
        sheet,
        row,
        "Целевые объекты",
        fill=NOTE_FILL,
        font=BOLD_FONT,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    headers = ["Объект", "Статус", "Строка КП", "Комментарий"]
    widths = [3, 2, 1, 2]
    column = 1
    for header, span in zip(headers, widths, strict=True):
        sheet.merge_cells(
            start_row=row,
            start_column=column,
            end_row=row,
            end_column=column + span - 1,
        )
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = GRID_BORDER
        column += span
    row += 1

    for item in decisions:
        values = [
            _text(item.get("target_label") or item.get("label") or item.get("target_id")),
            _target_anchor_status_label(item.get("anchor_status")),
            _text(item.get("anchor_line_id")),
            _text(item.get("reason")),
        ]
        column = 1
        for value_text, span in zip(values, widths, strict=True):
            sheet.merge_cells(
                start_row=row,
                start_column=column,
                end_row=row,
                end_column=column + span - 1,
            )
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = WRAP
            cell.border = GRID_BORDER
            column += span
        sheet.row_dimensions[row].height = 36
        row += 1
    return row + 1


def _write_commercial_procurement_gaps(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
) -> int:
    gaps = _procurement_gap_rows(value)
    if not gaps:
        return row
    row = _write_section(
        sheet,
        row,
        title,
        fill=NOTE_FILL,
        font=BOLD_FONT,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    headers = ["Позиция", "Кол-во", "Причина"]
    spans = [3, 1, 4]
    column = 1
    for header, span in zip(headers, spans, strict=True):
        sheet.merge_cells(
            start_row=row,
            start_column=column,
            end_row=row,
            end_column=column + span - 1,
        )
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = GRID_BORDER
        column += span
    row += 1

    for item, quantity, reason in gaps:
        values = [item, quantity, reason]
        column = 1
        for index, (value_text, span) in enumerate(zip(values, spans, strict=True), start=1):
            sheet.merge_cells(
                start_row=row,
                start_column=column,
                end_row=row,
                end_column=column + span - 1,
            )
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if index == 2 else WRAP
            cell.border = GRID_BORDER
            column += span
        sheet.row_dimensions[row].height = 36
        row += 1
    return row + 1


def _write_commercial_available_alternatives(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
) -> int:
    alternatives = _commercial_available_alternative_rows(value)
    if not alternatives:
        return row
    row = _write_section(
        sheet,
        row,
        title,
        fill=NOTE_FILL,
        font=BOLD_FONT,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    headers = [
        "Требование",
        "Артикул",
        "Производитель",
        "Наименование",
        "Доступно",
        "Цена",
        "Валюта",
        "Причина",
    ]
    spans = [1] * COMMERCIAL_COLUMN_COUNT
    column = 1
    for header, span in zip(headers, spans, strict=True):
        sheet.merge_cells(
            start_row=row,
            start_column=column,
            end_row=row,
            end_column=column + span - 1,
        )
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = GRID_BORDER
        column += span
    row += 1

    for alternative in alternatives:
        values = list(alternative)
        column = 1
        for index, (value_text, span) in enumerate(zip(values, spans, strict=True), start=1):
            sheet.merge_cells(
                start_row=row,
                start_column=column,
                end_row=row,
                end_column=column + span - 1,
            )
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if index in {5, 6, 7} else WRAP
            cell.border = GRID_BORDER
            column += span
        sheet.row_dimensions[row].height = 42
        row += 1
    return row + 1


def _write_commercial_status_block(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
) -> int:
    row = _write_section(
        sheet,
        row,
        title,
        fill=REVIEW_FILL,
        font=BOLD_FONT,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    cell = sheet.cell(row=row, column=1, value=_cell_value(value))
    cell.alignment = WRAP
    cell.border = GRID_BORDER
    sheet.row_dimensions[row].height = 42
    return row + 2


def _write_footer_note(sheet: Worksheet, row: int) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=COMMERCIAL_COLUMN_COUNT,
    )
    cell = sheet.cell(
        row=row,
        column=1,
        value=(
            "Примечание: итог включает только выбранные складские позиции. Позиции "
            "из блока добора, gaps и условия поставки не входят в стоимость, пока "
            "они не добавлены отдельными строками и не подтверждены менеджером."
        ),
    )
    cell.font = MUTED_FONT
    cell.alignment = WRAP
    cell.border = GRID_BORDER
    sheet.row_dimensions[row].height = 44


def _write_selected_stock_rows(
    sheet: Worksheet,
    row: int,
    quote_lines: list[Mapping[str, Any]],
) -> int:
    if not quote_lines:
        return row
    row = _write_section(sheet, row, "Выбранные строки склада")
    headers = [
        "Роль",
        "Наименование",
        "component_candidate_id",
        "stock_row_id",
        "Кол-во",
        "Сумма",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
    row += 1
    for item in quote_lines:
        values = [
            item.get("role"),
            _line_title(item),
            item.get("component_candidate_id"),
            item.get("stock_row_id"),
            item.get("quantity"),
            _price_text(item.get("line_total_value"), item.get("line_total_currency")),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value))
            cell.alignment = WRAP
            cell.border = GRID_BORDER
        row += 1
    return row + 1


def _write_title(sheet: Worksheet, row: int, title: str) -> int:
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = Font(size=14, bold=True)
    row += 2
    return row


def _write_key_value(sheet: Worksheet, row: int, label: str, value: Any) -> int:
    sheet.cell(row=row, column=1, value=label)
    sheet.cell(row=row, column=1).font = BOLD_FONT
    sheet.cell(row=row, column=2, value=_cell_value(value))
    sheet.cell(row=row, column=2).alignment = WRAP
    return row + 1


def _write_section(
    sheet: Worksheet,
    row: int,
    title: str,
    *,
    fill: PatternFill = HEADER_FILL,
    font: Font = HEADER_FONT,
    end_column: int = 8,
) -> int:
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = font
    sheet.cell(row=row, column=1).fill = fill
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    return row + 1


def _write_quote_lines(sheet: Worksheet, row: int, quote_lines: list[Mapping[str, Any]]) -> int:
    if not quote_lines:
        return row
    row = _write_section(sheet, row, "Позиции")
    headers = [
        "Роль",
        "Название",
        "component_candidate_id",
        "stock_row_id",
        "Кол-во",
        "Цена",
        "Сумма",
        "Причина",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    row += 1
    for item in quote_lines:
        values = [
            item.get("role"),
            _line_title(item),
            item.get("component_candidate_id"),
            item.get("stock_row_id"),
            item.get("quantity"),
            _price_text(item.get("unit_price_value"), item.get("unit_price_currency")),
            _price_text(item.get("line_total_value"), item.get("line_total_currency")),
            item.get("reason"),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value))
            cell.alignment = WRAP
            cell.border = THIN_BORDER
        row += 1
    return row + 1


def _write_long_text(sheet: Worksheet, row: int, title: str, value: Any) -> int:
    text = (
        _json_preview(value)
        if isinstance(value, Mapping | Sequence) and not isinstance(value, str)
        else _text(value)
    )
    if not text:
        return row
    row = _write_section(sheet, row, title)
    sheet.cell(row=row, column=1, value=text)
    sheet.cell(row=row, column=1).alignment = WRAP
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    return row + 2


def _write_list(sheet: Worksheet, row: int, title: str, value: Any) -> int:
    items = _string_list(value)
    if not items:
        return row
    row = _write_section(sheet, row, title)
    for item in items:
        sheet.cell(row=row, column=1, value=item)
        sheet.cell(row=row, column=1).alignment = WRAP
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
    return row + 1


def _write_procurement_gaps(sheet: Worksheet, row: int, title: str, value: Any) -> int:
    gaps = _procurement_gap_rows(value)
    if not gaps:
        return row
    row = _write_section(sheet, row, title)
    headers = ["Item", "Quantity", "Reason"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
        cell.alignment = CENTER
    row += 1
    for item, quantity, reason in gaps:
        values = [item, quantity, reason]
        for column, value_text in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if column == 2 else WRAP
            cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 34
        row += 1
    return row + 1


def _write_available_alternatives(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
) -> int:
    alternatives = _available_alternative_rows(value)
    if not alternatives:
        return row
    row = _write_section(sheet, row, title)
    headers = ["Требование", "Позиция", "stock_row_id", "Доступно", "Цена", "Причина"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
        cell.alignment = CENTER
    row += 1
    for requirement, item, stock_row_id, quantity, price, reason in alternatives:
        values = [requirement, item, stock_row_id, quantity, price, reason]
        for column, value_text in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if column in {3, 4, 5} else WRAP
            cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 38
        row += 1
    return row + 1


def _write_gap_considered_candidates(
    sheet: Worksheet,
    row: int,
    title: str,
    value: Any,
) -> int:
    candidates = _gap_considered_candidate_rows(value)
    if not candidates:
        return row
    row = _write_section(sheet, row, title)
    headers = ["Gap", "Позиция", "stock_row_id", "Доступно", "Цена", "Причина"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
        cell.alignment = CENTER
    row += 1
    for gap, item, stock_row_id, quantity, price, reason in candidates:
        values = [gap, item, stock_row_id, quantity, price, reason]
        for column, value_text in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if column in {3, 4, 5} else WRAP
            cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 38
        row += 1
    return row + 1


def _write_quote_integrity_adjustments(
    sheet: Worksheet,
    row: int,
    value: Mapping[str, Any],
) -> int:
    adjustments = _quote_integrity_adjustment_rows(value)
    if not adjustments:
        return row
    row = _write_section(sheet, row, "Корректировки quote_integrity")
    headers = [
        "Type",
        "Section",
        "Index",
        "Component ID",
        "Original stock_row_id",
        "Resolved stock_row_id",
        "Reason",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
        cell.alignment = CENTER
    row += 1
    for values in adjustments:
        for column, value_text in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if column in {3, 5, 6} else WRAP
            cell.border = GRID_BORDER
        sheet.row_dimensions[row].height = 38
        row += 1
    return row + 1


def _write_stock_confirmation_table(
    sheet: Worksheet,
    row: int,
    title: str,
    quote: Mapping[str, Any],
    *,
    commercial: bool = False,
) -> int:
    rows = _stock_confirmation_rows(quote)
    if not rows:
        return row
    end_column = COMMERCIAL_COLUMN_COUNT if commercial else 8
    row = _write_section(
        sheet,
        row,
        title,
        fill=REVIEW_FILL if commercial else NOTE_FILL,
        font=BOLD_FONT,
        end_column=end_column,
    )
    headers = ["Позиция", "В КП", "Остаток склада", "Что сделать"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = GRID_BORDER
        cell.alignment = CENTER
    if end_column > len(headers):
        sheet.merge_cells(
            start_row=row,
            start_column=len(headers),
            end_row=row,
            end_column=end_column,
        )
    row += 1
    for item, included_quantity, stock_quantity, action in rows:
        values = [item, included_quantity, stock_quantity, action]
        for column, value_text in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=_cell_value(value_text))
            cell.alignment = CENTER if column in {2, 3} else WRAP
            cell.border = GRID_BORDER
        if end_column > len(values):
            sheet.merge_cells(
                start_row=row,
                start_column=len(values),
                end_row=row,
                end_column=end_column,
            )
        sheet.row_dimensions[row].height = 38
        row += 1
    return row + 1


def _write_compatibility(
    sheet: Worksheet,
    row: int,
    compatibility: Mapping[str, Any],
) -> int:
    if not compatibility:
        return row
    row = _write_section(sheet, row, "Проверка совместимости")
    row = _write_key_value(
        sheet,
        row,
        "Статус",
        _compatibility_status_label(compatibility.get("status")),
    )
    row = _write_list(sheet, row, "Подтверждено", compatibility.get("checked_facts"))
    row = _write_list(sheet, row, "Блокеры", compatibility.get("blocking_mismatches"))
    row = _write_list(
        sheet,
        row,
        "Конфликты выбранных строк",
        compatibility.get("selected_line_conflicts"),
    )
    return _write_list(sheet, row, "Риски", compatibility.get("unresolved_risks"))


def _finish_sheet(
    sheet: Worksheet,
    *,
    widths: Mapping[int, int],
    freeze_panes: str = "A3",
) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = freeze_panes


def _append_text_block(lines: list[str], title: str, value: Any) -> None:
    text = _text(value)
    if text:
        lines.extend([f"## {title}", "", text, ""])


def _append_list_block(lines: list[str], title: str, value: Any) -> None:
    items = _string_list(value)
    if items:
        lines.extend([f"## {title}", ""])
        lines.extend(f"- {item}" for item in items)
        lines.append("")


def _append_stock_confirmation_block(
    lines: list[str],
    title: str,
    quote: Mapping[str, Any],
) -> None:
    items = _stock_confirmation_rows(quote)
    if not items:
        return
    lines.extend([f"## {title}", ""])
    for item, included_quantity, stock_quantity, action in items:
        lines.append(
            f"- {item}: в КП {included_quantity}; склад показывает {stock_quantity}. {action}"
        )
    lines.append("")


def _append_available_alternatives_block(lines: list[str], title: str, value: Any) -> None:
    items = _available_alternative_rows(value)
    if not items:
        return
    lines.extend([f"## {title}", ""])
    for requirement, item, stock_row_id, quantity, price, reason in items:
        parts = [
            f"требование: {requirement}" if requirement else "",
            item,
            f"доступно: {quantity}" if quantity else "",
            f"цена: {price}" if price else "",
            f"stock_row_id: {stock_row_id}" if stock_row_id else "",
            reason,
        ]
        lines.append(f"- {_display_text('. '.join(part for part in parts if part))}")
    lines.append("")


def _append_list_items(lines: list[str], label: str, value: Any) -> None:
    for item in _string_list(value):
        lines.append(f"- {label}: {item}")


def _commercial_part_number(item: Mapping[str, Any]) -> str:
    return _display_text(
        item.get("part_number")
        or item.get("sku")
        or item.get("article")
        or item.get("articul")
    )


def _commercial_producer(item: Mapping[str, Any]) -> str:
    producer = _display_text(
        item.get("producer")
        or item.get("manufacturer")
        or item.get("vendor")
        or item.get("brand")
    )
    if producer:
        return producer
    segments = _commercial_description_segments(item)
    if len(segments) >= 2 and _looks_like_producer_segment(segments[0]):
        return _display_text(segments[0])
    return ""


def _commercial_name(item: Mapping[str, Any]) -> str:
    segments = _commercial_description_segments(item)
    if not segments:
        return "Позиция из матрицы"
    name_index = 1 if len(segments) >= 2 and _looks_like_producer_segment(segments[0]) else 0
    name = segments[name_index]
    part_number = _commercial_part_number(item)
    if part_number and name.casefold().startswith(part_number.casefold()):
        name = name[len(part_number) :].lstrip(" -/:")
    return _display_text(name or segments[0])


def _commercial_description_segments(item: Mapping[str, Any]) -> list[str]:
    text = _text(
        item.get("title")
        or item.get("item_name")
        or item.get("product_name")
        or item.get("description")
        or item.get("item")
    )
    if not text:
        return []
    raw_segments = re.split(r"\s+/\s+", text)
    segments: list[str] = []
    for segment in raw_segments:
        segment = segment.strip()
        if not segment or _is_internal_catalog_segment(segment):
            continue
        segments.append(segment)
    return segments or [text]


def _is_internal_catalog_segment(segment: str) -> bool:
    normalized = segment.casefold()
    internal_markers = (
        "treolan.",
        "sale:",
        "uchmark:",
        "outoftrade:",
        "ntstatus:",
        "brutto:",
        "freenom",
        "freeptrans",
        "prid:",
        "vendor-id",
        "gtin:",
        "ean:",
    )
    return any(marker in normalized for marker in internal_markers)


def _looks_like_producer_segment(segment: str) -> bool:
    text = segment.strip()
    if not text or len(text) > 80:
        return False
    if _is_internal_catalog_segment(text):
        return False
    words = [word for word in re.split(r"\s+", text) if word]
    if len(words) > 5:
        return False
    return bool(re.search(r"[A-Za-zА-Яа-яЁё]", text))


def _line_title(item: Mapping[str, Any]) -> str:
    parts = [
        _text(item.get("producer")),
        _text(item.get("part_number")),
        _text(item.get("title") or item.get("item_name") or item.get("product_name")),
    ]
    title = " ".join(part for part in parts if part)
    return title or "Позиция из матрицы"


def _line_comment(item: Mapping[str, Any]) -> str:
    parts: list[str] = []
    reconciliation_note = _text(item.get("reconciliation_note"))
    if reconciliation_note:
        return _display_text(reconciliation_note)
    reason = _text(item.get("reason"))
    if reason:
        parts.append(reason)
    included = _included_components_text(item.get("included_components_summary"))
    note_text = " ".join(parts).casefold()
    if (
        included
        and "в составе по матрице" not in note_text
        and included.casefold() not in note_text
    ):
        parts.append(f"В составе по матрице: {included}")
    return _display_text(". ".join(parts))


def _included_components_text(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        return _display_text(value)
    if isinstance(value, Mapping):
        labels = [
            ("cpu", "CPU"),
            ("ram", "RAM"),
            ("raid_or_controller", "RAID/controller"),
            ("storage", "storage"),
            ("network", "network"),
            ("power", "power"),
            ("rails", "rails"),
            ("hba_or_options", "HBA/options"),
        ]
        parts = []
        for key, label in labels:
            text = _text(value.get(key))
            if text and not _is_missing_component_summary(text):
                parts.append(f"{label}: {text}")
        return "; ".join(parts)
    if isinstance(value, Sequence) and not isinstance(value, (bytes, bytearray)):
        return "; ".join(
            text
            for text in (_text(item) for item in value)
            if text and not _is_missing_component_summary(text)
        )
    return _display_text(value)


def _is_missing_component_summary(value: str) -> bool:
    normalized = " ".join(str(value or "").casefold().split())
    return normalized in {"не указано", "not specified", "unknown", "n/a", "none"}


def _price_text(value: Any, currency: Any) -> str:
    amount = _amount_text(value)
    currency_text = _text(currency)
    if amount and currency_text:
        return f"{amount} {currency_text}"
    return amount or currency_text


def _total_price_lines(quote: Mapping[str, Any]) -> list[str]:
    total = _price_text(
        quote.get("total_price_value"),
        quote.get("total_price_currency"),
    )
    if total:
        return [total]
    return _totals_by_currency_lines(quote.get("totals_by_currency"))


def _totals_by_currency_lines(value: Any) -> list[str]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        return []
    result: list[str] = []
    for item in value:
        if not isinstance(item, Mapping):
            continue
        total = _price_text(
            item.get("value") or item.get("total_price_value") or item.get("amount"),
            item.get("currency") or item.get("total_price_currency"),
        )
        if total:
            result.append(total)
    return result


def _group_quote_lines(quote_lines: Sequence[Mapping[str, Any]]) -> list[Mapping[str, Any]]:
    grouped: list[dict[str, Any]] = []
    by_key: dict[tuple[str, str, str, str], int] = {}
    quantity_sums: dict[tuple[str, str, str, str], Decimal] = {}
    total_sums: dict[tuple[str, str, str, str], Decimal] = {}

    for line in quote_lines:
        key = _quote_line_group_key(line)
        quantity = _decimal_value(line.get("quantity")) or Decimal("1")
        line_total = _decimal_value(line.get("line_total_value"))
        if key is None:
            grouped.append(dict(line))
            continue
        if key not in by_key:
            by_key[key] = len(grouped)
            grouped.append(dict(line))
            quantity_sums[key] = quantity
            if line_total is not None:
                total_sums[key] = line_total
            continue

        index = by_key[key]
        quantity_sums[key] += quantity
        if line_total is not None and key in total_sums:
            total_sums[key] += line_total
        merged = grouped[index]
        merged["quantity"] = _decimal_json(quantity_sums[key])
        if key in total_sums:
            merged["line_total_value"] = _decimal_json(total_sums[key])
        else:
            unit_price = _decimal_value(merged.get("unit_price_value"))
            if unit_price is not None:
                merged["line_total_value"] = _decimal_json(unit_price * quantity_sums[key])
        merged["line_total_currency"] = merged.get("unit_price_currency")
    return grouped


def _quote_line_group_key(line: Mapping[str, Any]) -> tuple[str, str, str, str] | None:
    product_key = _text(line.get("component_candidate_id") or line.get("part_number"))
    part_number = _text(line.get("part_number"))
    unit_price = _decimal_value(line.get("unit_price_value"))
    currency = _text(line.get("unit_price_currency"))
    if not product_key or unit_price is None or not currency:
        return None
    return (product_key, part_number, str(unit_price.normalize()), currency)


def _decimal_value(value: Any) -> Decimal | None:
    text = _text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _decimal_json(value: Decimal) -> int | str:
    if value == value.to_integral_value():
        return int(value)
    return format(value.normalize(), "f")


def _amount_text(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError):
        return text
    quantized = amount.quantize(Decimal("0.01"))
    if quantized == quantized.to_integral():
        return f"{int(quantized):,}".replace(",", " ")
    integer, _, fraction = f"{quantized:.2f}".partition(".")
    return f"{int(integer):,}".replace(",", " ") + "," + fraction


def _distributor_label(value: Any) -> str:
    code = _text(value).lower()
    labels = {
        "ocs": "OCS",
        "treolan": "Treolan",
    }
    return labels.get(code, _text(value) or "не указан")


def _profile_label(value: Any) -> str:
    profile = _text(value).lower()
    labels = {
        "server": "Серверы",
        "storage": "СХД/хранилища",
        "network": "Сеть",
        "custom": "Авто",
    }
    return labels.get(profile, _text(value) or "Авто")


def _role_label(value: Any) -> str:
    role = _text(value).lower()
    labels = {
        "platform": "Платформа",
        "server": "Сервер",
        "configured_system": "Готовая система",
        "cpu": "Процессор",
        "processor": "Процессор",
        "ram": "Оперативная память",
        "memory": "Оперативная память",
        "storage": "Накопитель",
        "ssd": "Накопитель",
        "hdd": "Накопитель",
        "drive": "Диск",
        "nas": "СХД/NAS",
        "storage_system": "СХД",
        "network": "Сеть",
        "switch": "Коммутатор",
        "network_adapter": "Сетевой адаптер",
        "power_supply": "Блок питания",
        "cable": "Кабель",
        "license": "Лицензия",
        "support": "Поддержка",
    }
    return labels.get(role, _text(value))


def _state_label(value: Any) -> str:
    state = _text(value)
    labels = {
        "quote_draft_review_required": "черновик КП, нужна инженерная проверка",
        "quote_candidate_customer_ready": "кандидат для КП",
        "no_recommendation": "нет валидной рекомендации",
        "matrix_too_large_for_model": "матрица не поместилась в модель",
        "matrix_empty_after_category_selection": "в выбранных категориях нет складских строк",
        "provider_error": "ошибка LLM-провайдера",
        "provider_not_configured": "LLM не настроена",
        "mechanical_validation_failed": "ответ LLM не прошел проверку склада и цен",
        "schema_validation_failed": "ответ LLM в неверном формате",
        "stock_refresh_failed": "не удалось обновить склад перед КП",
    }
    return labels.get(state, state)


def _category_summary(value: Any) -> str:
    categories = _string_list(value)
    if not categories:
        return "не указаны"
    if len(categories) <= 3:
        return ", ".join(categories)
    return f"{len(categories)} категорий"


def _selection_summary(quote: Mapping[str, Any]) -> str:
    parts = [
        _solution_scope_label(quote.get("solution_scope")),
        _substitution_policy_label(quote.get("substitution_policy")),
        _selection_mode_label(quote.get("selection_mode")),
        _completeness_status_label(quote.get("completeness_status")),
        _operational_status_label(quote.get("operational_status")),
    ]
    return "; ".join(part for part in parts if part)


def _status_badge_text(quote: Mapping[str, Any]) -> str:
    selection_mode = _text(quote.get("selection_mode")).lower()
    completeness_status = _text(quote.get("completeness_status")).lower()
    operational_status = _text(quote.get("operational_status")).lower()
    if selection_mode == "exact_complete":
        return "Полное соответствие"
    if selection_mode == "equivalent_complete":
        return "Функциональный эквивалент"
    if selection_mode == "degraded_complete":
        return "Ближайшая складская конфигурация"
    if selection_mode in {"partial_with_anchor", "partial_build"}:
        return "Требуется доукомплектация"
    if selection_mode == "partial_without_anchor" or completeness_status == "partial":
        return "Частичный складской комплект"
    if "requires_completion" in operational_status:
        return "Требуется доукомплектация"
    return "Ближайшая складская конфигурация"


def _solution_scope_label(value: Any) -> str:
    labels = {
        "complete_system": "объем: готовая система",
        "configured_system": "объем: конфигурируемая система",
        "standalone_product": "объем: отдельный товар",
        "replacement_component": "объем: замена компонента",
        "expansion_or_upgrade": "объем: расширение или апгрейд",
        "accessory": "объем: аксессуар",
        "multi_product_solution": "объем: комплект из нескольких товаров",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _substitution_policy_label(value: Any) -> str:
    labels = {
        "forbidden": "аналоги запрещены",
        "allowed_no_downgrade": "аналоги разрешены без ухудшения",
        "allowed_with_disclosed_downgrade": "аналоги разрешены с раскрытием отклонений",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _selection_mode_label(value: Any) -> str:
    labels = {
        "exact": "режим: точное соответствие",
        "exact_complete": "режим: полное точное соответствие",
        "equivalent_or_better": "режим: равноценно или лучше",
        "equivalent_complete": "режим: полный функциональный аналог",
        "analog_with_downgrade": "режим: аналог с отклонением",
        "downgraded_complete": "режим: полный аналог с отклонениями",
        "partial_build": "режим: частичная складская сборка",
        "partial_with_anchor": "режим: частичная сборка со складской базой",
        "partial_without_anchor": "режим: частичный складской комплект",
        "anchor_only": "режим: только складская база",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _completeness_status_label(value: Any) -> str:
    labels = {
        "complete": "комплектность: закрыто полностью",
        "partial": "комплектность: частично",
        "anchor_only": "комплектность: только база",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _operational_status_label(value: Any) -> str:
    labels = {
        "ready": "работоспособность: готово после проверки",
        "ready_after_customer_environment_check": (
            "работоспособность: нужна проверка среды заказчика"
        ),
        "incomplete_needs_procurement": "работоспособность: нужно добрать позиции",
        "operational": "работоспособность: готово после проверки",
        "operational_with_deviations": "работоспособность: есть отклонения",
        "requires_completion": "работоспособность: нужно добрать позиции",
        "not_claimed": "работоспособность: не заявляется",
        "not_applicable": "работоспособность: не применимо",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _no_recommendation_summary(no_recommendation: Mapping[str, Any]) -> str:
    summary = _display_text(no_recommendation.get("summary"))
    if not summary:
        return "КП не сформировано: не удалось закрыть требования по складской матрице."
    if _looks_english(summary):
        failed = _no_recommendation_failed_items(no_recommendation)
        if failed and any(marker in failed[0].lower() for marker in ("cpu", "процессор")):
            return "КП не сформировано: указанная модель процессора не найдена на выбранном складе."
        return "КП не сформировано: не удалось закрыть требования по складской матрице."
    return summary


def _no_recommendation_details(no_recommendation: Mapping[str, Any]) -> str:
    details = _display_text(no_recommendation.get("details"))
    return "" if _looks_english(details) else details


def _no_recommendation_failed_items(no_recommendation: Mapping[str, Any]) -> list[str]:
    result: list[str] = []
    for item in _string_list(no_recommendation.get("failed_requirements")):
        text = _known_missing_requirement(_display_text(item))
        if text:
            result.append(text)
    return result


def _no_recommendation_next_actions(no_recommendation: Mapping[str, Any]) -> list[str]:
    actions = [_display_text(item) for item in _string_list(
        no_recommendation.get("recommended_next_actions")
    )]
    if any(_looks_english(item) for item in actions):
        return [
            "Разрешить ближайший технически подходящий аналог по отсутствующей позиции.",
            "Если нужна строго указанная модель, проверить другой склад или поставку под заказ.",
        ]
    return [item for item in actions if item]


def _known_missing_requirement(text: str) -> str:
    clean = _display_text(text)
    if not clean:
        return ""
    cpu_match = re.match(
        r"(?i)^CPU model:\s*(.+?)\s+is missing from (?:the )?matrix\.?$",
        clean,
    )
    if cpu_match:
        model = cpu_match.group(1).strip()
        model = _russianize_core_count(model)
        return f"Не найдено в матрице склада: процессор {model}."
    missing_match = re.match(
        r"(?i)^(.+?)\s+is missing from (?:the )?matrix\.?$",
        clean,
    )
    if missing_match:
        return f"Не найдено в матрице склада: {missing_match.group(1).strip()}."
    return clean


def _compatibility_status_label(value: Any) -> str:
    status = _text(value).lower()
    labels = {
        "compatible": "совместимо",
        "compatible_selected_lines": "выбранные строки совместимы",
        "anchor_only": "выбрана только складская база",
        "confirmed_selected_set": "выбранные строки подтверждены по матрице",
        "review_required_selected_set": "выбранные строки требуют инженерной проверки",
        "independent_partial_set": "частичный независимый складской комплект",
        "incompatible": "несовместимо",
        "insufficient_facts": "недостаточно данных",
    }
    return labels.get(status, _text(value))


def _russianize_core_count(text: str) -> str:
    def replace(match: re.Match[str]) -> str:
        count = int(match.group(1))
        last_two = count % 100
        last = count % 10
        if 11 <= last_two <= 14:
            word = "ядер"
        elif last == 1:
            word = "ядро"
        elif 2 <= last <= 4:
            word = "ядра"
        else:
            word = "ядер"
        return f"{count} {word}"

    return re.sub(
        r"\b(\d+)\s+(?:cores?|ядро|ядра|ядер)\b",
        replace,
        text,
        flags=re.IGNORECASE,
    )


def _display_text(value: Any) -> str:
    return sanitize_user_facing_text(value)


def _looks_english(text: str) -> bool:
    clean = str(text or "")
    if not clean:
        return False
    latin = len(re.findall(r"[A-Za-z]", clean))
    cyrillic = len(re.findall(r"[А-Яа-яЁё]", clean))
    if latin >= 12 and cyrillic == 0:
        return True
    return latin >= 40 and latin > cyrillic * 2


def _cell_value(value: Any) -> Any:
    if isinstance(value, int | float):
        return value
    return (
        _json_preview(value)
        if isinstance(value, Mapping | Sequence) and not isinstance(value, str)
        else _text(value)
    )


def _json_preview(value: Any, *, limit: int = 4000) -> str:
    if value is None:
        return ""
    text = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return text[:limit]


def _mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _list_of_mappings(value: Any) -> list[Mapping[str, Any]]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        return []
    return [item for item in value if isinstance(item, Mapping)]


def _string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, Mapping):
        return [text] if (text := _display_list_item(value)) else []
    if isinstance(value, str):
        return [value] if value else []
    if not isinstance(value, Sequence) or isinstance(value, (bytes, bytearray)):
        return []
    return [text for item in value if (text := _display_list_item(item))]


def _display_list_item(value: Any) -> str:
    if not isinstance(value, Mapping):
        return _text(value)

    if "item" in value or "quantity" in value:
        item, quantity, reason = _procurement_gap_cells(value)
        headline = " - ".join(part for part in (item, _quantity_with_unit(quantity)) if part)
        if headline and reason:
            return _display_text(f"{headline}. {reason}")
        return _display_text(headline or reason)

    if "code" in value and "message_ru" in value:
        code = _text(value.get("code"))
        message = _text(value.get("message_ru"))
        path = _text(value.get("path"))
        stage = _text(value.get("stage"))
        parts = [
            f"код: {code}" if code else "",
            f"этап: {stage}" if stage else "",
            f"поле: {path}" if path else "",
            message,
        ]
        return ". ".join(part for part in parts if part)

    if "status" in value and ("required_for" in value or "next_action" in value):
        role = _text(value.get("role"))
        requested = _text(value.get("requested"))
        status = _gap_status_label(value.get("status"))
        required_for = _gap_required_for_label(value.get("required_for"))
        impact = _text(value.get("impact"))
        next_action = _text(value.get("next_action"))
        parts = [
            f"роль: {role}" if role else "",
            f"просили: {requested}" if requested else "",
            status,
            required_for,
            f"влияние: {impact}" if impact else "",
            f"дальше: {next_action}" if next_action else "",
        ]
        return _display_text(". ".join(part for part in parts if part))

    if "target_label" in value or "anchor_status" in value:
        label = _text(value.get("target_label") or value.get("label") or value.get("target_id"))
        status = _target_anchor_status_label(value.get("anchor_status"))
        line_id = _text(value.get("anchor_line_id"))
        reason = _text(value.get("reason"))
        parts = [
            label,
            status,
            f"строка {line_id}" if line_id else "",
            reason,
        ]
        return _display_text(" - ".join(part for part in parts if part))

    if "requested" in value or "offered" in value:
        parts: list[str] = []
        requirement_id = _text(value.get("requirement_id"))
        requirement = _text(value.get("requirement"))
        if requirement_id or requirement:
            parts.append(
                "Требование: "
                + " ".join(part for part in (requirement_id, requirement) if part)
            )
        requested = _text(value.get("requested"))
        offered = _text(value.get("offered"))
        if requested or offered:
            parts.append(
                f"просили: {requested or 'не указано'}; "
                f"предложено: {offered or 'не указано'}"
            )
        direction = _deviation_direction_label(value.get("direction"))
        severity = _deviation_severity_label(value.get("severity"))
        if direction or severity:
            parts.append("; ".join(part for part in (direction, severity) if part))
        impact = _text(value.get("impact"))
        if impact:
            parts.append(f"влияние: {impact}")
        reason = _text(value.get("reason"))
        if reason:
            parts.append(f"почему: {reason}")
        return _display_text(". ".join(parts))

    if "outcome" in value or "priority" in value:
        requirement_id = _text(value.get("requirement_id"))
        requirement = _text(value.get("requirement"))
        priority = _requirement_priority_label(value.get("priority"))
        outcome = _requirement_outcome_label(value.get("outcome"))
        requested = _text(value.get("requested"))
        offered = _text(value.get("offered"))
        impact = _text(value.get("impact"))
        parts = [
            part
            for part in (
                " ".join(part for part in (requirement_id, requirement) if part),
                priority,
                outcome,
                f"просили: {requested}" if requested else "",
                f"предложено: {offered}" if offered else "",
                impact,
            )
            if part
        ]
        return _display_text(". ".join(parts))

    if "result" in value:
        result = _text(value.get("result"))
        evidence = "; ".join(_string_list(value.get("evidence"))[:2])
        return _display_text("; ".join(part for part in (result, evidence) if part))

    if any(key in value for key in ("description", "comment", "summary", "needed_action")):
        requirement_id = _text(value.get("requirement_id"))
        requested = _text(value.get("requested"))
        action = _text(value.get("needed_action"))
        description = _text(
            value.get("description") or value.get("comment") or value.get("summary")
        )
        reason = _text(value.get("reason"))
        parts = [
            requirement_id,
            f"просили: {requested}" if requested else "",
            description,
            f"действие: {action}" if action else "",
            f"причина: {reason}" if reason else "",
        ]
        return _display_text(". ".join(part for part in parts if part))

    if "requirement" in value or "reason" in value:
        requirement_id = _text(value.get("requirement_id"))
        requirement = _text(value.get("requirement"))
        reason = _text(value.get("reason"))
        parts = []
        label = " ".join(part for part in (requirement_id, requirement) if part)
        if label:
            parts.append(f"Требование: {label}")
        if reason:
            parts.append(f"причина: {reason}")
        return _display_text(". ".join(parts))

    return _json_preview(value)


def _procurement_gap_rows(value: Any) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    for item in _iter_list_items(value):
        item_text, quantity, reason = _procurement_gap_cells(item)
        if item_text or quantity or reason:
            rows.append((item_text, quantity, reason))
    return rows


def _procurement_gap_cells(value: Any) -> tuple[str, str, str]:
    if not isinstance(value, Mapping):
        return (_text(value), "", "")

    if "item" in value or "quantity" in value:
        item = _text(
            value.get("item")
            or value.get("requested")
            or value.get("requirement")
            or value.get("role")
        )
        quantity = _quantity_text(value.get("quantity") or value.get("qty"))
        reason = _text(
            value.get("reason")
            or value.get("comment")
            or value.get("summary")
            or value.get("impact")
            or value.get("next_action")
        )
        return (item, quantity, reason)

    item = _text(value.get("requested") or value.get("requirement") or value.get("role"))
    reason = _display_list_item(value)
    return (item or reason, _quantity_text(value.get("quantity") or value.get("qty")), reason)


def _available_alternative_rows(value: Any) -> list[tuple[str, str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str, str]] = []
    for item in _iter_list_items(value):
        cells = _available_alternative_cells(item)
        if any(cells):
            rows.append(cells)
    return rows


def _gap_considered_candidate_rows(
    value: Any,
) -> list[tuple[str, str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str, str]] = []
    for gap in _iter_list_items(value):
        if not isinstance(gap, Mapping):
            continue
        gap_item, gap_quantity, _gap_reason = _procurement_gap_cells(gap)
        gap_label = " - ".join(
            part for part in (gap_item, _quantity_with_unit(gap_quantity)) if part
        )
        for candidate in _iter_list_items(gap.get("considered_candidates")):
            if not isinstance(candidate, Mapping):
                text = _display_text(candidate)
                if text:
                    rows.append((gap_label, text, "", "", "", ""))
                continue
            item = _text(
                candidate.get("item")
                or candidate.get("item_name")
                or candidate.get("description")
                or candidate.get("title")
                or candidate.get("part_number")
            )
            quantity = _quantity_text(
                candidate.get("available_quantity")
                or candidate.get("stock_quantity")
                or candidate.get("quantity")
                or candidate.get("qty"),
                lower_bound=bool(candidate.get("quantity_is_greater_than")),
            )
            price = _price_text(
                candidate.get("unit_price_value")
                or candidate.get("price_value")
                or candidate.get("price"),
                candidate.get("unit_price_currency")
                or candidate.get("price_currency")
                or candidate.get("currency"),
            )
            reason = _display_text(
                candidate.get("reason")
                or candidate.get("comment")
                or candidate.get("summary")
                or candidate.get("difference")
            )
            rows.append(
                (
                    gap_label,
                    item,
                    _text(candidate.get("stock_row_id")),
                    quantity,
                    price,
                    reason,
                )
            )
    return [row for row in rows if any(row)]


def _commercial_available_alternative_rows(
    value: Any,
) -> list[tuple[str, str, str, str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str, str, str, str]] = []
    for item in _iter_list_items(value):
        if not isinstance(item, Mapping):
            text = _display_text(item)
            if text:
                rows.append(("", "", "", text, "", "", "", ""))
            continue
        requirement = _display_text(
            item.get("requirement_id")
            or item.get("requirement")
            or item.get("role")
            or item.get("requested")
        )
        quantity = _quantity_text(
            item.get("available_quantity")
            or item.get("stock_quantity")
            or item.get("quantity")
            or item.get("qty"),
            lower_bound=bool(item.get("quantity_is_greater_than")),
        )
        price = _amount_text(
            item.get("unit_price_value")
            or item.get("price_value")
            or item.get("price")
        )
        currency = _text(
            item.get("unit_price_currency")
            or item.get("price_currency")
            or item.get("currency")
        )
        reason = _display_text(
            item.get("reason")
            or item.get("comment")
            or item.get("summary")
            or item.get("difference")
        )
        rows.append(
            (
                requirement,
                _commercial_part_number(item),
                _commercial_producer(item),
                _commercial_name(item),
                quantity,
                price,
                currency,
                reason,
            )
        )
    return [row for row in rows if any(row)]


def _available_alternative_cells(value: Any) -> tuple[str, str, str, str, str, str]:
    if not isinstance(value, Mapping):
        return ("", _text(value), "", "", "", "")

    requirement = _text(
        value.get("requirement_id")
        or value.get("requirement")
        or value.get("requested")
        or value.get("role")
    )
    item = _text(
        value.get("item")
        or value.get("item_name")
        or value.get("description")
        or value.get("title")
        or value.get("part_number")
        or value.get("role")
    )
    stock_row_id = _text(value.get("stock_row_id"))
    quantity = _quantity_text(
        value.get("available_quantity")
        or value.get("stock_quantity")
        or value.get("quantity")
        or value.get("qty"),
        lower_bound=bool(value.get("quantity_is_greater_than")),
    )
    price = _price_text(
        value.get("unit_price_value") or value.get("price_value") or value.get("price"),
        value.get("unit_price_currency") or value.get("price_currency") or value.get("currency"),
    )
    reason = _text(
        value.get("reason")
        or value.get("comment")
        or value.get("summary")
        or value.get("difference")
    )
    return (requirement, item, stock_row_id, quantity, price, reason)


def _quote_integrity_adjustment_rows(
    value: Mapping[str, Any],
) -> list[tuple[str, str, str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str, str, str]] = []
    for item in _iter_list_items(value.get("adjustments")):
        if not isinstance(item, Mapping):
            text = _display_text(item)
            if text:
                rows.append(("", "", "", "", "", "", text))
            continue
        rows.append(
            (
                _text(item.get("type")),
                _text(item.get("section")),
                _text(item.get("index")),
                _text(item.get("component_candidate_id")),
                _text(item.get("original_stock_row_id")),
                _text(item.get("resolved_stock_row_id") or item.get("stock_row_id")),
                _text(item.get("reason") or item.get("resolution")),
            )
        )
    return [row for row in rows if any(row)]


def _stock_confirmation_rows(
    quote: Mapping[str, Any],
) -> list[tuple[str, str, str, str]]:
    quote_lines = _list_of_mappings(quote.get("lines"))
    rows: list[tuple[str, str, str, str]] = []
    seen_keys: set[tuple[str, str, str, str]] = set()

    for line in quote_lines:
        if not line.get("stock_confirmation_required"):
            continue
        row = _stock_confirmation_row_from_line(line)
        key = (
            _text(line.get("component_candidate_id")),
            _text(line.get("stock_row_id")),
            row[1],
            row[2],
        )
        if key not in seen_keys:
            rows.append(row)
            seen_keys.add(key)

    integrity = _mapping(quote.get("quote_integrity"))
    for adjustment in _iter_list_items(integrity.get("adjustments")):
        if not isinstance(adjustment, Mapping):
            continue
        if _text(adjustment.get("type")) != "stock_lower_bound_quantity_confirm":
            continue
        component_id = _text(adjustment.get("component_candidate_id"))
        stock_row_id = _text(adjustment.get("stock_row_id"))
        included_quantity = _quantity_with_unit(
            _quantity_text(adjustment.get("included_quantity"))
        )
        stock_quantity = _quantity_with_unit(
            _quantity_text(
                adjustment.get("displayed_available_quantity"),
                lower_bound=True,
            )
        )
        key = (component_id, stock_row_id, included_quantity, stock_quantity)
        if key in seen_keys:
            continue
        line = _quote_line_for_stock_confirmation(
            quote_lines,
            component_id=component_id,
            stock_row_id=stock_row_id,
        )
        item = _line_title(line) if line else component_id or stock_row_id
        rows.append(
            _stock_confirmation_row(
                item=item,
                included_quantity=included_quantity,
                stock_quantity=stock_quantity,
            )
        )
        seen_keys.add(key)

    return rows


def _quote_line_for_stock_confirmation(
    quote_lines: Sequence[Mapping[str, Any]],
    *,
    component_id: str,
    stock_row_id: str,
) -> Mapping[str, Any] | None:
    if stock_row_id:
        for line in quote_lines:
            if _text(line.get("stock_row_id")) == stock_row_id:
                return line
    if component_id:
        for line in quote_lines:
            if _text(line.get("component_candidate_id")) == component_id:
                return line
    return None


def _stock_confirmation_row_from_line(
    line: Mapping[str, Any],
) -> tuple[str, str, str, str]:
    return _stock_confirmation_row(
        item=_line_title(line),
        included_quantity=_quantity_with_unit(_quantity_text(line.get("quantity"))),
        stock_quantity=_quantity_with_unit(
            _quantity_text(line.get("quantity_value"), lower_bound=True)
        ),
    )


def _stock_confirmation_row(
    *,
    item: str,
    included_quantity: str,
    stock_quantity: str,
) -> tuple[str, str, str, str]:
    action = (
        "Перед отправкой КП нужно подтвердить доступность выбранного количества."
    )
    return (
        item or "Позиция из матрицы",
        included_quantity or "не указано",
        stock_quantity or "нижняя граница остатка",
        action,
    )


def _iter_list_items(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value] if value else []
    if isinstance(value, Mapping):
        return [value]
    if not isinstance(value, Sequence) or isinstance(value, (bytes, bytearray)):
        return [value]
    return [item for item in value if item not in (None, "")]


def _quantity_text(value: Any, *, lower_bound: bool = False) -> str:
    text = _text(value)
    if lower_bound and text and not text.endswith("+"):
        return f"{text}+"
    return text


def _quantity_with_unit(value: str) -> str:
    if not value:
        return ""
    if re.fullmatch(r"\d+(?:[.,]\d+)?\+?", value):
        return f"{value} шт."
    return value


def _deviation_direction_label(value: Any) -> str:
    labels = {
        "upgrade": "лучше исходного требования",
        "downgrade": "хуже исходного требования",
        "different": "другое отличие",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _deviation_severity_label(value: Any) -> str:
    labels = {
        "minor": "незначимое отклонение",
        "material": "существенное отклонение",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _gap_status_label(value: Any) -> str:
    labels = {
        "not_in_matrix": "нет в переданной матрице",
        "out_of_stock": "нет доступного остатка",
        "quantity_shortage": "не хватает количества",
        "no_compatible_item_proven": "совместимый вариант не подтвержден по матрице",
        "not_included": "не входит в выбранную позицию",
        "unknown": "нужно уточнение",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _gap_required_for_label(value: Any) -> str:
    labels = {
        "operational_readiness": "нужно для работоспособности",
        "requested_spec": "нужно для полного закрытия ТЗ",
        "optional_preference": "опциональное пожелание",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _target_anchor_status_label(value: Any) -> str:
    labels = {
        "selected": "выбран",
        "covered": "закрыт выбранной строкой",
        "covered_by_selected_line": "закрыт выбранной строкой",
        "not_found": "не найден в матрице",
        "not_required": "отдельная якорная строка не нужна",
        "alternative": "доступная альтернатива",
        "rejected": "не выбран",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _requirement_priority_label(value: Any) -> str:
    labels = {
        "non_negotiable": "обязательное требование",
        "locked": "зафиксированное требование",
        "core": "основная задача",
        "important": "важное требование",
        "target": "целевое требование",
        "preference": "предпочтение",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _requirement_outcome_label(value: Any) -> str:
    labels = {
        "met": "закрыто",
        "exceeded": "закрыто с запасом",
        "equivalent": "закрыто аналогом",
        "degraded": "закрыто с ухудшением",
        "partially_met": "закрыто частично",
        "missing": "не закрыто",
        "unknown": "нужно уточнение",
        "substituted": "заменено аналогом",
        "not_met": "не закрыто",
        "not_applicable": "не применимо",
    }
    return labels.get(str(value or "").strip().lower(), _text(value))


def _text(value: Any) -> str:
    return str(value or "").strip()
