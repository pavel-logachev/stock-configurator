from __future__ import annotations

import asyncio
import json
from collections.abc import AsyncGenerator, Generator
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

import app.cli.preview_llm_configurator_package as preview_llm_configurator_package_cli
from app.api.routes import match as match_routes
from app.core.config import (
    LlmSettings,
    WebEvidenceSettings,
    get_llm_settings,
    get_settings,
    get_web_evidence_settings,
)
from app.core.database import Base, get_session
from app.db.models import DistributorProduct, DistributorStockPrice, MatchCandidate, MatchRun
from app.llm import configuration_composer as composer_module
from app.llm.configuration_composer import (
    build_llm_configurator_package,
    compose_llm_configurations,
)
from app.matching import match_engine as match_engine_module
from app.matching.match_engine import MatchResult
from app.matching.spec_schema import StockSpec
from app.matching.v3_full_category_quote_service import V3FullCategoryQuoteResult
from app.user_facing_text import contains_cjk_text

SERVER_REQUEST = "Need 2 servers 2U, 2 CPU, 512 GB RAM, SSD, 2 PSU, Moscow"
NETWORK_MATCH_71_TEXT = (
    "Нужен 1 коммутатор 48 портов 1G PoE+, 4 uplink 10G SFP+, "
    "L3, stacking, склад Москва, один самый дешевый вариант для КП"
)
NETWORK_MATCH_74_PLAIN_POE_TEXT = (
    "Нужен 1 коммутатор 48 портов 1G PoE, 4 uplink 10G SFP+, "
    "L3, stacking желательно, склад Москва, один самый дешевый вариант для КП"
)
COMPLEX_SERVER_78_TEXT = """
Execution: 1U
Sockets: 2
CPU: Intel 6th generation, 2 pcs, at least 24 cores
RAM: 256 GB DDR5 RDIMM
Disks: 6 x SSD 1920 GB SATA, 2 x SSD 480 GB SATA
Controller: LSI Logic 9400-8i / LSI 9500-8i
Network adapter: Intel X710-DA2 2x10GbE SFP+
Power: 2 x 2000W hot-swap Platinum, C13-C14 cables, C13-Schuko cables
Cooling: 8 fans N+1
Interfaces: USB 3.0, serial RJ-45, VGA, remote management RJ-45
""".strip()


class AsyncSessionAdapter:
    def __init__(self, session: Session) -> None:
        self._session = session

    def add(self, instance: object) -> None:
        self._session.add(instance)

    async def flush(self) -> None:
        self._session.flush()

    async def commit(self) -> None:
        self._session.commit()

    async def rollback(self) -> None:
        self._session.rollback()

    async def execute(self, statement: Any) -> Any:
        return self._session.execute(statement)


class SessionContext:
    def __init__(self, adapter: AsyncSessionAdapter) -> None:
        self._adapter = adapter

    async def __aenter__(self) -> AsyncSessionAdapter:
        return self._adapter

    async def __aexit__(self, exc_type: object, exc: object, tb: object) -> None:
        if exc is not None:
            await self._adapter.rollback()


@pytest.fixture()
def db_session() -> Generator[Session]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)

    with Session(engine, expire_on_commit=False) as session:
        yield session

    engine.dispose()


@pytest.fixture()
def client(
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> Generator[TestClient]:
    monkeypatch.setenv("DATABASE_URL", "postgresql+asyncpg://test:test@localhost:5432/test")
    monkeypatch.setenv("ENVIRONMENT", "test")
    monkeypatch.setenv("LLM_PROVIDER", "disabled")
    get_settings.cache_clear()
    get_llm_settings.cache_clear()
    get_web_evidence_settings.cache_clear()

    from app.api.main import create_app

    app = create_app()

    async def override_get_session() -> AsyncGenerator[AsyncSessionAdapter, None]:
        yield AsyncSessionAdapter(db_session)

    app.dependency_overrides[get_session] = override_get_session
    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()
    get_settings.cache_clear()
    get_llm_settings.cache_clear()
    get_web_evidence_settings.cache_clear()


def test_create_match_with_text_uses_fallback_extractor_and_saves_run(
    client: TestClient,
    db_session: Session,
) -> None:
    _seed_nerpa_products(db_session)

    response = client.post("/api/v1/match", json={"text": SERVER_REQUEST})

    assert response.status_code == 201
    body = response.json()
    assert body["status"] == "partial_stock_matched"
    assert body["engineer_review_required"] is True
    assert body["total_candidates"] == 2
    assert body["matched_items"] == 0
    assert body["confirmation_text"]
    assert body["candidates"][0]["part_number"] == "D5720-181125SA04"
    assert body["candidates"][0]["available_quantity"] == 3
    assert body["candidates"][0]["price_value"] == "6900"
    assert body["candidates"][0]["price_currency"] == "USD"
    assert [candidate["part_number"] for candidate in body["ready_stock_candidates"]] == [
        "D5720-181125SA04",
        "D5720-181125SA05",
    ]
    assert body["build_candidates"] == []
    assert "Match Engine V0 Report" in body["report_markdown"]
    assert body["report_url"] == f"/api/v1/match/{body['match_run_id']}/report.md"
    assert body["report_xlsx_url"] == f"/api/v1/match/{body['match_run_id']}/report.xlsx"

    match_run = db_session.get(MatchRun, body["match_run_id"])
    assert match_run is not None
    assert match_run.source == "text"
    assert match_run.source_text == SERVER_REQUEST


def test_create_v3_full_category_quote_saves_run_and_reports(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_run_v3_full_category_quote(**_kwargs: Any) -> V3FullCategoryQuoteResult:
        return V3FullCategoryQuoteResult(
            profile="server",
            category_ids=["V1100"],
            distributor_code="ocs",
            result_state="quote_draft_review_required",
            report_json={
                "pipeline_version": "v3_full_category_matrix",
                "llm_configurator_used": True,
                "primary_recommendation_status": "valid",
                "final_status_source": "v3_full_category_quote_validated",
                "v3_result_state": "quote_draft_review_required",
                "v3_profile": "server",
                "category_ids": ["V1100"],
                "distributor_code": "ocs",
                "resolved_request": {
                    "objective": "cheapest_minimum_viable",
                    "customer_task_summary": "Need v3 server",
                },
                "validated_quote": {
                    "engineering_review_required": True,
                    "total_price_value": "100.0000",
                    "total_price_currency": "USD",
                    "why_selected": "Lowest technically workable draft.",
                    "price_audit": ["No cheaper compatible matrix row was selected."],
                    "lines": [
                        {
                            "role": "platform",
                            "component_candidate_id": "ocs:p1",
                            "stock_row_id": "ocs:p1:1",
                            "quantity": 1,
                            "unit_price_value": "100.0000",
                            "unit_price_currency": "USD",
                            "line_total_value": "100.0000",
                            "line_total_currency": "USD",
                            "reason": "Matrix row fits the request.",
                        }
                    ],
                    "compatibility_check": {
                        "status": "compatible",
                        "checked_facts": ["ocs:p1:1 is selected."],
                        "blocking_mismatches": [],
                        "unresolved_risks": [],
                    },
                    "engineer_checks": ["Confirm final compatibility."],
                },
                "diagnostics": {
                    "matrix_row_count": 1,
                    "matrix_component_count": 1,
                    "model": "qwen/qwen3.7-plus",
                },
                "v3_validation_errors": [],
                "v3_validation_warnings": [],
            },
        )

    monkeypatch.setattr(
        match_routes,
        "run_v3_full_category_quote",
        fake_run_v3_full_category_quote,
    )

    response = client.post(
        "/api/v1/match/v3/full-category",
        json={"text": "Need v3 server"},
    )

    assert response.status_code == 200
    body = response.json()
    match_run_id = body["match_run_id"]
    assert match_run_id > 0
    assert body["report_url"] == f"/api/v1/match/{match_run_id}/report.md"
    assert body["report_xlsx_url"] == f"/api/v1/match/{match_run_id}/report.xlsx"
    assert body["report_json"]["match_run_id"] == match_run_id

    match_run = db_session.get(MatchRun, match_run_id)
    assert match_run is not None
    assert match_run.source == "v3_full_category_text"
    assert match_run.source_text == "Need v3 server"
    assert match_run.status == "quote_draft_review_required"
    assert match_run.report_json["pipeline_version"] == "v3_full_category_matrix"
    assert match_run.report_json["report_xlsx_url"] == body["report_xlsx_url"]

    markdown_response = client.get(f"/api/v1/match/{match_run_id}/report.md")
    assert markdown_response.status_code == 200
    assert "КП draft" in markdown_response.text
    assert "Спецификация для КП" in markdown_response.text

    excel_response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")
    assert excel_response.status_code == 200
    workbook = load_workbook(BytesIO(excel_response.content))
    assert workbook.sheetnames == ["КП", "Инженерная проверка", "Склад и диагностика"]
    quote_sheet = workbook["КП"]
    quote_sheet_text = "\n".join(
        str(cell.value)
        for row in quote_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    diagnostics_sheet_text = "\n".join(
        str(cell.value)
        for row in workbook["Склад и диагностика"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert quote_sheet["A1"].value == "Коммерческое предложение - черновик"
    assert "Спецификация для КП" in quote_sheet_text
    assert "Итого по включённым складским позициям" in quote_sheet_text
    assert "Проверка цены" in quote_sheet_text
    assert "component_candidate_id" not in quote_sheet_text
    assert "ocs:p1" not in quote_sheet_text
    assert "component_candidate_id" in diagnostics_sheet_text
    assert "ocs:p1" in diagnostics_sheet_text


def test_create_match_normalizes_v2_no_recommendation_before_response_and_save(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_orchestrator(*args: Any, **kwargs: Any) -> Any:
        spec = StockSpec(source_text="Need future stock config", items=[])
        match_result = MatchResult(
            spec=spec,
            status="partial_stock_matched",
            engineer_review_required=True,
            total_candidates=0,
            matched_items=0,
            missing_requirements=[],
            risk_flags=[],
            candidates=[],
            primary_recommendation_status="no_recommendation",
            llm_fallback_reason="composer_structured_no_recommendation",
        )
        return SimpleNamespace(
            match_result=match_result,
            report_json={
                "pipeline_version": "v2_composer_first",
                "composer_mode": "multi_pass",
                "product_group": "future_group",
                "primary_object": "future_object",
                "primary_recommendation_status": "no_recommendation",
                "final_status_source": "composer_no_recommendation",
                "llm_fallback_reason": "composer_structured_no_recommendation",
                "no_recommendation_reason": {
                    "summary": None,
                    "fallback_reason": None,
                    "failed_requirements": [],
                    "role_failures": [],
                    "recommended_next_actions": [],
                    "diagnostic_notes": None,
                },
                "final_bom_after_repair": {
                    "no_recommendation": {
                        "reason": "Repair pass found no safe complete BOM.",
                        "role_level_reasons": [
                            {"role": "compute_node", "reason": "Socket mismatch."},
                            {"role": "fabric_adapter", "reason": "Stock shortage."},
                        ],
                    }
                },
                "completeness_critic_result": {
                    "recommended_repair_actions": [
                        "Request compatible replacements."
                    ]
                },
            },
        )

    from app.api.routes import match as match_routes

    monkeypatch.setattr(match_routes, "run_ai_match_orchestrator", fake_orchestrator)

    response = client.post(
        "/api/v1/match",
        json={"text": "Need future stock config", "pipeline_v2": True},
    )

    assert response.status_code == 201
    body = response.json()
    reason = body["no_recommendation_reason"]
    assert reason["summary"] == "Repair pass found no safe complete BOM."
    assert reason["fallback_reason"] == "composer_structured_no_recommendation"
    assert [row["role"] for row in reason["role_failures"]] == [
        "compute_node",
        "fabric_adapter",
    ]
    assert reason["recommended_next_actions"] == ["Request compatible replacements."]

    match_run = db_session.get(MatchRun, body["match_run_id"])
    assert match_run is not None
    persisted_reason = match_run.report_json["no_recommendation_reason"]
    assert persisted_reason["summary"] == "Repair pass found no safe complete BOM."
    assert persisted_reason["role_failures"] == reason["role_failures"]


def test_create_match_network_text_uses_semantic_planning_and_persists_details(
    client: TestClient,
    db_session: Session,
) -> None:
    _seed_match_71_network_products(db_session)

    response = client.post("/api/v1/match", json={"text": NETWORK_MATCH_71_TEXT})

    assert response.status_code == 201
    body = response.json()
    matrix = body["component_candidate_matrix"]
    switch_candidates = matrix["switch_candidates"]
    part_numbers = [candidate["part_number"] for candidate in switch_candidates]

    assert body["product_group"] == "network"
    assert any(
        capability["role"] == "switch"
        and "48" in str(capability["capability_id"])
        for capability in body["required_capabilities"]
    )
    assert body["required_roles"] == ["switch"]
    assert body["category_plan"]["switch"] == ["V120100"]
    assert body["category_plan_entries"]
    assert body["role_coverage_summary"]["switch"]["missing"] is False
    assert switch_candidates[0]["part_number"] == "SW-48P-4SFP"
    assert {"SW-5P", "SW-8P", "SW-16P", "SW-24P"}.isdisjoint(part_numbers)
    assert matrix["ready_server_candidates"] == []
    assert body["shortlist_for_llm"]
    assert body["llm_fallback_reason"] == "llm_configurator_disabled"

    match_run_id = body["match_run_id"]
    details = client.get(f"/api/v1/match/{match_run_id}").json()

    assert details["product_group"] == body["product_group"]
    assert details["required_capabilities"] == body["required_capabilities"]
    assert details["required_roles"] == body["required_roles"]
    assert details["category_plan"] == body["category_plan"]
    assert details["component_candidate_matrix"]["switch_candidates"]
    assert details["shortlist_for_llm"] == body["shortlist_for_llm"]


def test_create_match_network_plain_poe_text_preserves_planning_and_matrix(
    client: TestClient,
    db_session: Session,
) -> None:
    _seed_match_71_network_products(db_session)

    response = client.post(
        "/api/v1/match",
        json={"text": NETWORK_MATCH_74_PLAIN_POE_TEXT},
    )

    assert response.status_code == 201
    body = response.json()
    switch_capability = next(
        capability
        for capability in body["required_capabilities"]
        if capability["role"] == "switch"
    )
    switch_requirements = switch_capability["parsed_requirements"]
    optional_text = json.dumps(body["optional_capabilities"], ensure_ascii=False)

    assert body["product_group"] == "network"
    assert body["required_roles"] == ["switch"]
    assert switch_requirements["port_count"] == 48
    assert switch_requirements["poe_required"] is True
    assert switch_requirements["poe_standard"] == "PoE"
    assert switch_requirements["poe_standard"] != "PoE+"
    assert switch_requirements["uplink_count"] == 4
    assert switch_requirements["l3_required"] is True
    assert "stacking_required" not in switch_requirements
    assert "stacking_required" in optional_text
    assert body["category_plan"]["switch"] == ["V120100"]
    assert body["component_candidate_matrix"]["switch_candidates"]
    assert body["shortlist_for_llm"]

    details = client.get(f"/api/v1/match/{body['match_run_id']}").json()
    assert details["required_capabilities"] == body["required_capabilities"]
    assert details["optional_capabilities"] == body["optional_capabilities"]
    assert details["category_plan"] == body["category_plan"]


def test_create_match_complex_server_78_preserves_semantic_fields(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    role_plan = _semantic_server_78_role_plan()
    role_plan.update(
        {
            "requirement_classifier_status": "incomplete_repair",
            "requirement_classifier_incomplete_reason": (
                "repair_source_coverage_below_threshold"
            ),
            "requirement_source_coverage": [
                {
                    "source_text": "C13-C14 cables",
                    "covered": False,
                    "source": "request_text",
                }
            ],
            "requirement_source_coverage_percent": 75.0,
            "unclassified_source_fragments": ["C13-C14 cables"],
            "synthetic_requirement_count": 1,
            "source_backed_requirement_count": 2,
            "requirement_classifier_repair_quality": "incomplete_source_coverage",
            "requirement_classifier_repair_accepted": False,
        }
    )

    async def fake_match_stock_spec(
        spec: StockSpec,
        session: Any,
    ) -> MatchResult:
        return MatchResult(
            spec=spec,
            status="no_stock_match",
            engineer_review_required=True,
            total_candidates=0,
            matched_items=0,
            missing_requirements=[],
            risk_flags=[],
            candidates=[],
            component_candidate_matrix={
                "product_group": "server",
                "role_plan": role_plan,
                "matrix_blueprint": role_plan["matrix_blueprint"],
                "matrix_blueprint_roles": role_plan["matrix_blueprint_roles"],
                "required_capabilities": role_plan["required_capabilities"],
                "required_roles": role_plan["required_roles"],
                "category_plan": {
                    "server_platform": ["platform"],
                    "network_adapter": ["nic"],
                    "cable": ["power-cable"],
                },
                "server_platform_candidates": [],
                "network_adapter_candidates": [],
                "cable_candidates": [],
            },
            product_group="server",
            role_plan=role_plan,
            category_plan={
                "server_platform": ["platform"],
                "network_adapter": ["nic"],
                "cable": ["power-cable"],
            },
            required_capabilities=role_plan["required_capabilities"],
            required_roles=role_plan["required_roles"],
        )

    from app.api.routes import match as match_routes

    monkeypatch.setattr(match_routes, "match_stock_spec", fake_match_stock_spec)

    response = client.post("/api/v1/match", json={"text": COMPLEX_SERVER_78_TEXT})

    assert response.status_code == 201
    body = response.json()
    assert body["product_group"] == "server"
    assert body["primary_object"] == "server"
    assert body["semantic_planner_source"] == "llm"
    assert body["matrix_blueprint_roles"] == role_plan["matrix_blueprint_roles"]
    assert body["embedded_requirements"] == role_plan["embedded_requirements"]
    assert body["not_primary_product_groups"] == role_plan["not_primary_product_groups"]
    assert "network_adapter" in body["required_roles"]
    assert "cable" in body["required_roles"]
    assert body["requirement_classifier_status"] == "incomplete_repair"
    assert body["requirement_source_coverage_percent"] == 75.0
    assert body["unclassified_source_fragments"] == ["C13-C14 cables"]
    assert body["requirement_classifier_repair_accepted"] is False

    details = client.get(f"/api/v1/match/{body['match_run_id']}").json()

    assert details["product_group"] == "server"
    assert details["primary_object"] == "server"
    assert details["matrix_blueprint_roles"] == body["matrix_blueprint_roles"]
    assert details["required_capabilities"] == body["required_capabilities"]
    assert details["category_plan"] == body["category_plan"]
    assert details["requirement_source_coverage"] == body["requirement_source_coverage"]
    assert details["requirement_classifier_incomplete_reason"] == (
        "repair_source_coverage_below_threshold"
    )


def test_preview_and_api_text_paths_share_fake_llm_semantic_plan(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    _seed_complete_extended_component_set(db_session)
    _seed_component_product(
        db_session,
        item_id="nic-1",
        part_number="NIC-X710",
        producer="Intel",
        category_id="V120116",
        item_name="Intel X710-DA2 dual-port 10GbE SFP+ server adapter",
        quantity=2,
        price=Decimal("200"),
    )
    adapter = AsyncSessionAdapter(db_session)

    def session_factory() -> SessionContext:
        return SessionContext(adapter)

    monkeypatch.setattr(
        preview_llm_configurator_package_cli,
        "get_session_factory",
        lambda: session_factory,
    )
    monkeypatch.setattr(
        match_engine_module,
        "OpenAICompatibleLlmClient",
        _ApiSemanticPlannerClient,
    )
    monkeypatch.setenv("LLM_PROVIDER", "openai-compatible")
    monkeypatch.setenv("LLM_BASE_URL", "https://llm.example.test/v1")
    monkeypatch.setenv("LLM_API_KEY", "api-preview-secret")
    monkeypatch.setenv("LLM_MODEL", "qwen-test")
    get_llm_settings.cache_clear()

    preview_exit = asyncio.run(
        preview_llm_configurator_package_cli.run(
            ["--text", COMPLEX_SERVER_78_TEXT, "--json"]
        )
    )
    preview_capture = capsys.readouterr()
    preview = json.loads(preview_capture.out)
    response = client.post("/api/v1/match", json={"text": COMPLEX_SERVER_78_TEXT})

    assert preview_exit == 0
    assert response.status_code == 201
    body = response.json()
    assert preview["semantic_planner_source"] == "llm"
    assert body["semantic_planner_source"] == preview["semantic_planner_source"]
    assert body["product_group"] == preview["product_group"] == "server"
    assert body["primary_object"] == preview["primary_object"] == "server"
    assert body["category_planner_source"] == preview["category_planner_source"]
    assert body["matrix_blueprint_roles"] == preview["matrix_blueprint_roles"]
    assert body["package_budget"]["over_budget"] == preview["package_budget"][
        "over_budget"
    ]
    assert body["package_skipped_reason"] == preview["package_skipped_reason"]
    assert body["count_by_role"] == preview["count_by_role"]
    assert body["package_strategy_decision"] == preview["package_strategy_decision"]
    assert preview["package_strategy_decision"]["decision"] == "use_full_broad_package"
    assert preview["package_strategy_decision"]["full_matrix_required"] is False
    assert "package_strategy_decision" in preview_capture.err
    assert "full_matrix_start" not in preview_capture.err
    assert body["match_trace"]
    assert preview["match_trace"]
    assert "network_adapter" in body["required_roles"]
    assert "api-preview-secret" not in json.dumps(body, ensure_ascii=False)
    get_llm_settings.cache_clear()


def test_create_match_text_server_78_ready_package_calls_online_composer(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = LlmSettings(
        llm_provider="openai-compatible",
        llm_base_url="https://llm.example.test/v1",
        llm_api_key="test-key",
        llm_model="test-model",
        llm_configurator_enabled=True,
        llm_configurator_mode="composer",
        llm_configurator_output_mode="single_best_cost_valid",
        llm_configurator_max_package_chars=200000,
    )
    web_settings = WebEvidenceSettings(
        web_evidence_enabled=True,
        web_evidence_provider="routerai",
        web_evidence_mode="online_composer",
        web_evidence_model="online-model",
    )

    async def fake_match_stock_spec(
        spec: StockSpec,
        session: Any,
    ) -> MatchResult:
        matrix = _api_server_78_like_matrix()
        normalized_requirements = [_api_server_78_requirements()]
        composer_client = _ApiDirectComposerClient(_api_server_78_online_response)
        outcome = compose_llm_configurations(
            user_request=spec.source_text,
            normalized_requirements=normalized_requirements,
            ready_stock_candidates=[],
            component_candidate_matrix=matrix,
            rule_based_build_candidates=[],
            settings=settings,
            llm_client=composer_client,
            web_evidence_settings=web_settings,
        )
        return MatchResult(
            spec=spec,
            status="partial_stock_matched",
            engineer_review_required=True,
            total_candidates=0,
            matched_items=0,
            missing_requirements=[],
            risk_flags=[],
            candidates=[],
            component_candidate_matrix=matrix,
            normalized_requirements=normalized_requirements,
            llm_configurator_enabled=outcome.enabled,
            llm_configurator_used=outcome.used,
            output_mode=outcome.output_mode,
            llm_recommended_build_candidates=outcome.recommended_builds,
            primary_recommendation=outcome.primary_recommendation,
            primary_recommendation_status=outcome.primary_recommendation_status,
            no_recommendation_reason=outcome.no_recommendation_reason,
            commercial_summary=outcome.commercial_summary,
            llm_fallback_reason=outcome.fallback_reason,
            llm_error_type=outcome.error_type,
            llm_http_status=outcome.http_status,
            llm_parse_diagnostics=outcome.parse_diagnostics,
            llm_internal_warnings=outcome.internal_warnings,
            llm_proposals_count=outcome.proposal_count,
            valid_proposals_count=outcome.valid_proposals_count,
            validation_rejected_count=outcome.validation_rejected_count,
            selection_skipped_count=outcome.selection_skipped_count,
            rejected_ai_recommendations_count=outcome.rejected_recommendations_count,
            ai_recommendations_validation_warnings=outcome.validation_warnings,
            ai_validation_summary=outcome.validation_summary,
            rejected_reasons_top=outcome.rejected_reasons_top,
            rejected_ai_recommendations_debug_safe=(
                outcome.rejected_recommendations_debug_safe
            ),
            web_evidence_pack=outcome.evidence_pack,
            llm_evidence_review=outcome.evidence_review,
            llm_package_diagnostics=outcome.package_diagnostics,
            product_group="server",
            role_plan=matrix["role_plan"],
            category_plan=matrix["category_plan"],
            category_plan_entries=matrix["category_plan_entries"],
            category_catalog_summary=matrix["category_catalog_summary"],
            category_planner_source="ai_category_planner",
            category_plan_source="llm",
            required_capabilities=matrix["required_capabilities"],
            required_roles=matrix["required_roles"],
            role_coverage_summary=matrix["role_coverage_summary"],
        )

    from app.api.routes import match as match_routes

    monkeypatch.setattr(match_routes, "match_stock_spec", fake_match_stock_spec)

    response = client.post("/api/v1/match", json={"text": COMPLEX_SERVER_78_TEXT})

    assert response.status_code == 201
    body = response.json()
    assert body["product_group"] == "server"
    assert body["llm_configurator_enabled"] is True
    assert body["online_composer_used"] is True
    assert body["llm_configurator_used"] is True
    assert body["primary_recommendation_status"] == "valid"
    assert body["llm_proposals_count"] == 1
    assert body["package_budget"]["over_budget"] is False
    assert body["package_skipped_reason"] is None
    assert body["package_strategy_decision"]["strategy"] in {
        "full_broad_package_direct_to_composer",
        "full_matrix_reduced_package",
    }
    assert body["match_trace"]
    assert body["category_planner_source"] == "ai_category_planner"
    assert body["count_by_role"]["cpu"] == 1
    assert body["package_approximate_size"]["chars"] == body["package_budget"][
        "final_chars"
    ]


def test_create_match_text_server_81_ready_package_attempts_online_composer_once(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _seed_server_81_component_set(db_session)
    _ApiServer81OpenAIClient.reset()
    monkeypatch.setenv("LLM_PROVIDER", "openai-compatible")
    monkeypatch.setenv("LLM_BASE_URL", "https://llm.example.test/v1")
    monkeypatch.setenv("LLM_API_KEY", "test-key")
    monkeypatch.setenv("LLM_MODEL", "test-model")
    monkeypatch.setenv("LLM_CONFIGURATOR_ENABLED", "true")
    monkeypatch.setenv("LLM_CONFIGURATOR_MODE", "composer")
    monkeypatch.setenv("LLM_CONFIGURATOR_OUTPUT_MODE", "single_best_cost_valid")
    monkeypatch.setenv("LLM_CONFIGURATOR_MAX_PACKAGE_CHARS", "200000")
    monkeypatch.setenv("WEB_EVIDENCE_ENABLED", "true")
    monkeypatch.setenv("WEB_EVIDENCE_PROVIDER", "routerai")
    monkeypatch.setenv("WEB_EVIDENCE_MODE", "online_composer")
    monkeypatch.setenv("WEB_EVIDENCE_BASE_URL", "https://routerai.example.test/v1")
    monkeypatch.setenv("WEB_EVIDENCE_API_KEY", "routerai-key")
    monkeypatch.setenv("WEB_EVIDENCE_MODEL", "online-model")
    get_llm_settings.cache_clear()
    get_web_evidence_settings.cache_clear()

    from app.llm import configuration_composer

    monkeypatch.setattr(
        match_engine_module,
        "OpenAICompatibleLlmClient",
        _ApiServer81OpenAIClient,
    )
    monkeypatch.setattr(
        configuration_composer,
        "OpenAICompatibleLlmClient",
        _ApiServer81OpenAIClient,
    )

    response = client.post("/api/v1/match", json={"text": COMPLEX_SERVER_78_TEXT})

    assert response.status_code == 201
    body = response.json()
    assert _ApiServer81OpenAIClient.composer_calls == 1
    assert body["product_group"] == "server"
    assert body["primary_object"] == "server"
    assert body["semantic_planner_source"] == "llm"
    assert body["category_planner_source"] == "ai_category_planner"
    assert body["package_budget"]["over_budget"] is False
    assert body["package_skipped_reason"] is None
    assert body["llm_configurator_enabled"] is True
    assert body["online_composer_used"] is True
    assert body["llm_configurator_used"] is True
    assert body["llm_proposals_count"] == 1
    decision = body["composer_attempt_decision"]
    assert decision["should_attempt"] is True
    assert decision["blocked_by"] == []
    assert decision["candidate_count_total"] >= 8
    assert decision["provider_configured"] is True

    details = client.get(f"/api/v1/match/{body['match_run_id']}").json()
    assert details["composer_attempt_decision"] == decision
    assert details["package_strategy_decision"] == body["package_strategy_decision"]
    assert details["match_trace"] == body["match_trace"]
    get_llm_settings.cache_clear()
    get_web_evidence_settings.cache_clear()


def test_create_match_with_spec_saves_run_without_confirmation_text(
    client: TestClient,
    db_session: Session,
) -> None:
    _seed_nerpa_products(db_session)

    response = client.post("/api/v1/match", json={"spec": _server_spec_payload()})

    assert response.status_code == 201
    body = response.json()
    assert body["status"] == "partial_stock_matched"
    assert body["total_candidates"] == 2
    assert len(body["candidates"]) == 2
    assert body["confirmation_text"] is None

    match_run = db_session.get(MatchRun, body["match_run_id"])
    assert match_run is not None
    assert match_run.source == "spec"
    assert match_run.source_text == "Spec from stock-bot draft"


def test_get_match_returns_detail_with_candidates(client: TestClient, db_session: Session) -> None:
    match_run_id = _seed_match_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}")

    assert response.status_code == 200
    body = response.json()
    assert body["id"] == match_run_id
    assert body["status"] == "partial_stock_matched"
    assert body["spec_json"]["items"] == []
    assert body["report_json"]["status"] == "partial_stock_matched"
    assert body["risk_flags"] == ["engineer_review_required"]
    assert body["missing_requirements"] == ["RAM below requirement"]
    assert body["candidates"][0]["part_number"] == "D5720-181125SA04"
    assert body["candidates"][0]["price_value"] == "6900"


def test_get_match_report_markdown(client: TestClient, db_session: Session) -> None:
    match_run_id = _seed_match_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.md")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/markdown")
    assert "Match Engine V0 Report" in response.text


def test_get_match_report_excel_contains_expected_sheets(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run_id = _seed_match_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == f'attachment; filename="stock_match_{match_run_id}.xlsx"'
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "AI-рекомендации",
        "Матрица компонентов",
    ]
    merged_ranges = [
        str(merged_range)
        for merged_range in workbook["AI-рекомендации"].merged_cells.ranges
    ]
    assert "A1:H1" in merged_ranges
    assert "A3:H3" in merged_ranges
    assert "A4:B4" in merged_ranges
    assert "C4:D4" in merged_ranges
    assert "E4:F4" in merged_ranges
    assert "G4:H4" in merged_ranges
    assert "A6:H6" in merged_ranges
    assert "A7:H7" in merged_ranges
    assert "A8:H8" in merged_ranges
    assert workbook["AI-рекомендации"]["A1"].value == "AI-рекомендации по складскому подбору"
    assert workbook["AI-рекомендации"]["A3"].value == "Служебные данные"
    assert workbook["AI-рекомендации"]["A4"].value == "Номер подбора"
    assert workbook["AI-рекомендации"]["C4"].value == match_run_id
    assert workbook["AI-рекомендации"]["A6"].alignment.wrap_text is True
    assert workbook["AI-рекомендации"]["A6"].alignment.vertical == "top"
    assert workbook["AI-рекомендации"]["A7"].alignment.wrap_text is True
    assert workbook["AI-рекомендации"]["A7"].alignment.vertical == "top"
    assert workbook["AI-рекомендации"].row_dimensions[6].height > 20
    assert workbook["AI-рекомендации"].row_dimensions[7].height > 20
    assert workbook["AI-рекомендации"].freeze_panes == "A10"
    assert all(
        workbook["AI-рекомендации"].column_dimensions[column].width <= 24
        for column in "ABCDEFGH"
    )
    assert workbook["AI-рекомендации"].max_column <= 8
    matrix_sheet = workbook["Матрица компонентов"]
    assert matrix_sheet["A1"].value == "Охват матрицы"
    assert matrix_sheet.freeze_panes.startswith("A")


def test_saved_network_report_excel_filters_server_artifacts(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run_id = _seed_saved_network_report_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    ai_text = _worksheet_text(workbook["AI-рекомендации"])
    matrix_text = _worksheet_text(workbook["Матрица компонентов"])

    assert workbook.sheetnames == ["AI-рекомендации", "Матрица компонентов"]
    assert "Сетевое оборудование" in ai_text
    assert "product_group=network" in ai_text
    assert "Количество сетевого оборудования: 1" in ai_text
    assert "Коммутатор" in matrix_text
    assert "PoE" in ai_text
    assert "L2/L3" in ai_text or "L3" in ai_text
    assert "OS3254P/370W/A1A" in matrix_text
    assert "Сильное соответствие" in matrix_text
    assert "483.84" in matrix_text
    assert "17" in matrix_text

    for forbidden in (
        "Количество серверов",
        "CPU support",
        "QVL",
        "DIMM",
        "NVMe backplane",
    ):
        assert forbidden not in ai_text
    for forbidden in (
        "Готовые серверы",
        "готовый сервер",
        "Платформы",
        "\nCPU\n",
        "\nRAM\n",
        "CPU-LEGACY",
    ):
        assert forbidden not in matrix_text


def test_saved_storage_report_excel_uses_storage_requirements(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run_id = _seed_saved_storage_report_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    ai_text = _worksheet_text(workbook["AI-рекомендации"])
    matrix_text = _worksheet_text(workbook["Матрица компонентов"])

    assert "product_group=storage" in ai_text
    assert "Количество СХД: 1" in ai_text
    assert "Емкость: usable 100 ТБ; raw 120 ТБ; RAID6" in ai_text
    assert "Контроллеры: 2" in ai_text
    assert "Протокол/порты: FC; 4 x 32G SFP28" in ai_text
    assert "Количество серверов" not in ai_text
    assert "СХД" in matrix_text
    assert "Контроллеры СХД" in matrix_text
    assert "Готовые серверы" not in matrix_text
    assert "готовый сервер" not in matrix_text
    assert "CPU-LEGACY" not in matrix_text


def test_saved_server_report_excel_keeps_server_matrix_behavior(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run_id = _seed_saved_server_report_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    ai_text = _worksheet_text(workbook["AI-рекомендации"])
    matrix_text = _worksheet_text(workbook["Матрица компонентов"])

    assert "Количество серверов: 2" in ai_text
    assert "CPU" in ai_text
    assert "RAM" in ai_text
    assert "Готовые серверы" in matrix_text
    assert "готовый сервер" in matrix_text
    assert "READY-SERVER" in matrix_text


def test_get_match_report_excel_formats_build_amounts_and_cpu_columns(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run_id = _seed_build_match_run(db_session)

    response = client.get(f"/api/v1/match/{match_run_id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["AI-рекомендации"]
    workbook_text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "LLM Composer не использовался: слой выключен." in workbook_text
    assert sheet.max_column <= 8


def test_api_and_excel_hide_incompatible_cpu_kits_from_build_candidates(
    client: TestClient,
    db_session: Session,
) -> None:
    _seed_component_product(
        db_session,
        item_id="asus-platform",
        part_number="90SF03A1-M00070",
        producer="ASUS",
        category_id="V110100",
        item_name="ASUS 2U dual socket DDR5 server platform 2x PSU",
        quantity=2,
        price=Decimal("2000"),
    )
    _seed_component_product(
        db_session,
        item_id="hpe-cpu-kit",
        part_number="P49616-B21",
        producer="HPE",
        category_id="V110103",
        item_name="HPE Intel Xeon Gold 5416S processor kit",
        quantity=4,
        price=Decimal("900"),
    )
    _seed_component_product(
        db_session,
        item_id="ram-ddr5",
        part_number="RAM-DDR5-64G",
        producer="Samsung",
        category_id="V110104",
        item_name="DDR5 RDIMM 64GB server memory module",
        quantity=16,
        price=Decimal("150"),
    )
    _seed_component_product(
        db_session,
        item_id="ssd-1",
        part_number="SSD-960G",
        producer="Samsung",
        category_id="V110106",
        item_name="Server SSD 960GB SATA",
        quantity=2,
        price=Decimal("200"),
    )

    response = client.post("/api/v1/match", json={"spec": _server_spec_payload()})

    assert response.status_code == 201
    body = response.json()
    assert body["ready_stock_candidates"] == []
    assert len(body["build_candidates"]) == 1
    build_candidate = body["build_candidates"][0]
    assert build_candidate["platform"]["part_number"] == "90SF03A1-M00070"
    assert build_candidate["completeness_status"] == "incomplete"
    assert build_candidate["missing_component_roles"] == ["cpu"]
    assert build_candidate["excluded_from_total_roles"] == ["cpu"]
    assert build_candidate["total_price_value"] == "6800"
    assert build_candidate["score"] is not None
    assert build_candidate["rank_reason"]
    assert body["component_candidate_matrix"]
    assert "P49616-B21" not in str(build_candidate)

    report_response = client.get(f"/api/v1/match/{body['match_run_id']}/report.xlsx")

    assert report_response.status_code == 200
    workbook = load_workbook(BytesIO(report_response.content))
    assert workbook.sheetnames == ["AI-рекомендации", "Матрица компонентов"]
    component_sheet = workbook["Матрица компонентов"]
    component_values = "\n".join(
        str(cell.value)
        for row in component_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "P49616-B21" not in component_values


def test_excel_ai_recommendations_use_validated_components_not_llm_display(
    client: TestClient,
    db_session: Session,
) -> None:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера, 512 ГБ RAM DDR5",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={
            "items": [{"item_type": "server", "quantity": 2, "name": "server"}],
            "source_text": "Нужно 2 сервера, 512 ГБ RAM DDR5",
        },
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "llm_proposals_count": 10,
            "ai_recommendations_count": 1,
            "rejected_ai_recommendations_count": 9,
            "ai_validation_summary": {
                "accepted": 1,
                "rejected_fatal": 2,
                "rejected_missing_required": 3,
                "rejected_stock": 1,
                "rejected_role_mismatch": 0,
                "rejected_unknown_component": 0,
                "rejected_right_size": 1,
                "rejected_duplicate": 2,
                "rejected_other": 0,
                "rejected": 9,
            },
            "ai_recommendations": [
                {
                    "source_type": "partial_build",
                    "decision": "recommend_with_checks",
                    "title": "Частичная сборка",
                    "display_name": "ASUS + Intel + Micron 512GB RAM",
                    "quantity_required": 2,
                    "total_price_value": "6400",
                    "total_price_currency": "USD",
                    "total_price_note": "без RAM",
                    "completeness_status": "incomplete",
                    "missing_component_roles": ["ram"],
                    "why_selected": "Платформа и CPU есть на складе.",
                    "critical_checks": [
                        "Llm_rec_1: web evidence not found for Платформа; keep engineer проверить"
                    ],
                    "evidence_summary": {
                        "status_text": "Явных конфликтов по найденным источникам не выявлено.",
                        "confidence": "high",
                        "sources_count": 2,
                        "confirmed": ["socket: LGA4677", "тип памяти: DDR5"],
                        "missing": ["CPU support list"],
                        "engineering_checks": ["Проверить CPU support list платформы."],
                    },
                    "components": [
                        {
                            "role": "server_platform",
                            "producer": "ASUS",
                            "part_number": "PLATFORM",
                            "quantity_required": 2,
                        },
                        {
                            "role": "cpu",
                            "producer": "Intel",
                            "part_number": "CPU",
                            "quantity_required": 4,
                        },
                        {
                            "role": "ssd",
                            "producer": "Samsung",
                            "part_number": "SSD",
                            "quantity_required": 4,
                        },
                    ],
                }
            ],
            "component_candidate_matrix": {
                "platform_candidates": [
                    {
                        "component_candidate_id": "platform-evidence",
                        "role": "server_platform",
                        "producer": "ASUS",
                        "part_number": "PLATFORM",
                        "name": "ASUS PLATFORM",
                    }
                ]
            },
            "web_evidence_pack": {
                "enabled": True,
                "provider": "fake",
                "total_tasks": 1,
                "completed_tasks": 1,
                "components": [
                    {
                        "component_candidate_id": "platform-evidence",
                        "role": "server_platform",
                        "part_number": "PLATFORM",
                        "name": "ASUS PLATFORM",
                        "evidence_status": "found",
                        "confidence": "high",
                        "facts": {
                            "socket_family": "LGA4677",
                            "memory_type": "DDR5",
                        },
                        "sources": [
                            {
                                "domain": "servers.asus.com",
                                "url": "https://servers.asus.com/platform",
                                "title": "ASUS PLATFORM",
                                "snippet": "LGA4677 DDR5",
                            }
                        ],
                    }
                ],
            },
        },
        report_markdown="# report",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["AI-рекомендации"]
    workbook_text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Платформа: ASUS PLATFORM - 1 шт." in workbook_text
    assert "CPU: Intel CPU - 2 шт." in workbook_text
    assert "SSD: Samsung SSD - 2 шт." in workbook_text
    assert "Платформа: 2 шт." in workbook_text
    assert "CPU: 4 шт." in workbook_text
    assert "SSD: 4 шт." in workbook_text
    assert "Micron" not in workbook_text
    assert "RAM:" not in workbook_text
    assert "1. Частичная сборка" in workbook_text
    assert "AI проверил 10 вариантов, к показу выбрано 1." in workbook_text
    assert "Часть вариантов была скрыта как дубли или уступающие по цене/рискам." in workbook_text
    assert "Валидатор отклонил: 2 из-за совместимости" in workbook_text
    assert "4 из-за неполной комплектации" in workbook_text
    assert "Скрыто при выборе: 2 как дубли" in workbook_text
    assert "1 как уступающие по цене/рискам" in workbook_text
    assert "только после доукомплектования и инженерной проверки" in workbook_text
    assert "Доказательная проверка" in workbook_text
    assert "Источники: 2" in workbook_text
    assert "Проверить инженеру" in workbook_text
    assert "список поддерживаемых CPU" in workbook_text
    assert "Llm_rec_" not in workbook_text
    assert "web evidence not found" not in workbook_text
    assert "keep engineer" not in workbook_text
    matrix_text = "\n".join(
        str(cell.value)
        for row in workbook["Матрица компонентов"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Доказательства" in matrix_text
    assert "Уверенность источников" in matrix_text
    assert "servers.asus.com" in matrix_text


def test_excel_ai_recommendations_use_grouped_presales_output(
    client: TestClient,
    db_session: Session,
) -> None:
    group = _grouped_excel_fixture_group()
    group["group_title"] = "Intel LGA4677 / DDR5 / NVMe с高密度"
    group["architecture_summary"] = "architecture_summary: с高密度 NVMe"
    group["why_group_matters"] = "why_group_matters: с高密度 NVMe"
    group["engineer_checks"] = [
        "Проверить совместимость CPU с платформой.",
        "Проверить список поддерживаемых CPU платформы.",
    ]
    group["platform_options"][0]["why_this_platform"] = (
        "why_this_platform: с高密度 NVMe; raw JSON; "
        "component_candidate_id: platform-secret-id"
    )
    group["platform_options"][0]["engineer_checks"] = [
        "Проверить совместимость CPU с платформой.",
        "Проверить список поддерживаемых CPU платформы.",
        "Проверить QVL памяти.",
        "Проверить совместимость RAM.",
    ]
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера, 512 ГБ RAM DDR5, 2 SSD NVMe",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={
            "items": [{"item_type": "server", "quantity": 2, "name": "server"}],
            "source_text": "Нужно 2 сервера, 512 ГБ RAM DDR5, 2 SSD NVMe",
        },
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "grouped_presales_mode_used": True,
            "configuration_groups_count": 1,
            "configuration_groups": [group],
            "quote_recommendation": {
                "for_cheapest_quote": "ASUS PLATFORM-CHEAP - 8 600 USD",
                "for_database_preferred": "Supermicro SYS-621C-TN12R - 10 200 USD",
                "for_engineering_clarity": "Supermicro SYS-621C-TN12R - 10 200 USD",
                "summary": "quote_recommendation: Minimal cost build meeting all core specs.",
            },
            "ai_recommendations": [
                {
                    "title": "Legacy card should not be rendered",
                    "components": [],
                }
            ],
            "component_candidate_matrix": {},
        },
        report_markdown="# report",
        created_at=datetime(2026, 5, 10, 12, 0, tzinfo=UTC),
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["AI-рекомендации"]
    sheet_text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Предварительная спецификация для КП" in sheet_text
    assert "Сервер в сборе" in sheet_text
    assert "Платформа" in sheet_text
    assert "CPU" in sheet_text
    assert "RAM" in sheet_text
    assert "SSD" in sheet_text
    assert "Цена за весь запрос" in sheet_text
    assert "Альтернатива спокойнее для инженеров" in sheet_text
    assert "Конфигурационные семейства" in sheet_text
    assert "Компонентная база" in sheet_text
    assert "Варианты платформ" in sheet_text
    assert "Что проверить инженеру" in sheet_text
    assert "Сумма за весь запрос" in sheet_text
    assert "Инженерный статус" in sheet_text
    assert not contains_cjk_text(sheet_text)
    assert "высокоплот" in sheet_text
    assert "Minimal cost" not in sheet_text
    assert "Proven" not in sheet_text
    assert "Premium platform" not in sheet_text
    assert "preliminary_requires_engineer_review" not in sheet_text
    assert "why_this_platform:" not in sheet_text
    assert "quote_recommendation:" not in sheet_text
    assert "component_candidate_id" not in sheet_text
    assert "platform-secret-id" not in sheet_text
    assert "raw JSON" not in sheet_text
    assert "llm_rec" not in sheet_text
    assert "Legacy card should not be rendered" not in sheet_text


def test_excel_ai_recommendations_use_primary_recommendation_top_block(
    client: TestClient,
    db_session: Session,
) -> None:
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера, 512 ГБ RAM DDR5, 2 SSD NVMe",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={
            "items": [{"item_type": "server", "quantity": 2, "name": "server"}],
            "source_text": "Нужно 2 сервера, 512 ГБ RAM DDR5, 2 SSD NVMe",
        },
        report_json={
            "match_run_id": 501,
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "output_mode": "single_best_cost_valid",
            "primary_recommendation_status": "valid",
            "primary_recommendation": _primary_excel_recommendation(),
            "commercial_summary": {"mode": "single_best_cost_valid"},
            "grouped_presales_mode_used": False,
            "configuration_groups_count": 1,
            "configuration_groups": [_grouped_excel_fixture_group()],
            "component_candidate_matrix": {},
        },
        report_markdown="# report",
        created_at=datetime(2026, 5, 10, 12, 0, tzinfo=UTC),
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["AI-рекомендации"]
    assert sheet["A1"].value == "Предварительная спецификация для КП"
    assert sheet["A4"].value == "Позиция"
    assert sheet["D4"].value == "На 1 сервер"
    assert sheet["F4"].value == "Всего к заказу"
    assert sheet["G4"].value == "Остаток"
    assert sheet["H4"].value == "Примечание"
    assert sheet["A23"].value == "Служебные данные"
    source_row = next(
        row_index
        for row_index in range(1, sheet.max_row + 1)
        if str(sheet.cell(row=row_index, column=1).value or "").startswith("Исходный запрос")
    )
    normalized_row = next(
        row_index
        for row_index in range(1, sheet.max_row + 1)
        if str(sheet.cell(row=row_index, column=1).value or "").startswith(
            "Нормализованные требования"
        )
    )
    assert source_row > 20
    assert normalized_row > source_row
    sheet_text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Предварительная спецификация для КП" in sheet_text
    assert "Рекомендуемый вариант для самого дешевого КП" not in sheet_text
    assert "Сервер в сборе" in sheet_text
    assert "Позиция" in sheet_text
    assert "На 1 сервер" in sheet_text
    assert "Всего к заказу" in sheet_text
    assert "Остаток" in sheet_text
    assert "Примечание" in sheet_text
    assert "Ориентировочно за 2 сервера: 8 600 USD" in sheet_text
    assert "Проверить перед КП" in sheet_text
    assert "component_candidate_id" not in sheet_text
    assert '"facts"' not in sheet_text
    assert '"evidence"' not in sheet_text
    assert "preliminary_requires_engineer_review" not in sheet_text
    assert "llm_rec" not in sheet_text
    assert "Конфигурационная база" not in sheet_text
    assert "Варианты платформ" not in sheet_text


def test_excel_ai_recommendations_show_online_composer_evidence_passport(
    client: TestClient,
    db_session: Session,
) -> None:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [{"item_type": "server", "quantity": 2, "name": "server"}]},
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "evidence_mode": "online_composer",
            "online_composer_used": True,
            "evidence_sources_count": 0,
            "web_evidence_diagnostics": {
                "evidence_mode": "online_composer",
                "online_composer_used": True,
                "evidence_sources_count": 0,
            },
            "ai_recommendations": [
                {
                    "source_type": "build_from_parts",
                    "title": "Online build",
                    "quantity_required": 2,
                    "total_price_value": "8800",
                    "total_price_currency": "USD",
                    "why_selected_short": "Закрывает требования по складу.",
                    "evidence_summary": {
                        "status": "not_confirmed",
                        "sources_count": 0,
                        "not_confirmed": ["support list"],
                        "source_domains": [],
                    },
                    "components": [
                        {
                            "role": "server_platform",
                            "producer": "ASUS",
                            "part_number": "PLATFORM",
                            "quantity_required": 2,
                            "available_quantity": 2,
                        }
                    ],
                }
            ],
        },
        report_markdown="# report",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["AI-рекомендации"]
    workbook_text = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert sheet["A5"].value == "Режим проверки"
    assert sheet["C5"].value == "online composer"
    assert sheet["E5"].value == "Источники"
    assert sheet["G5"].value == 0
    assert "Источники: 0" in workbook_text
    assert "web evidence not found" not in workbook_text
    assert "keep engineer" not in workbook_text


def test_excel_ai_recommendations_split_confidence_without_evidence(
    client: TestClient,
    db_session: Session,
) -> None:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text=SERVER_REQUEST,
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=["engineer_review_required"],
        spec_json={"items": []},
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "ai_recommendations_count": 1,
            "evidence_used": False,
            "evidence_mode": "separate",
            "ai_recommendations": [
                {
                    "source_type": "build_from_parts",
                    "decision": "recommend",
                    "title": "Предварительная сборка",
                    "display_name": "Gooxi platform",
                    "quantity_required": 1,
                    "total_price_value": "1000",
                    "total_price_currency": "USD",
                    "confidence": "high",
                    "commercial_fit_confidence": "high",
                    "evidence_summary": {
                        "status": "disabled",
                        "sources_count": 0,
                        "confidence": "unknown",
                    },
                    "why_selected_short": "Коммерчески подходит по цене и наличию.",
                    "components": [
                        {
                            "role": "server_platform",
                            "producer": "Gooxi",
                            "part_number": "PLATFORM",
                            "quantity_required": 1,
                            "available_quantity": 1,
                        }
                    ],
                }
            ],
        },
        report_markdown="# report",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    workbook_text = "\n".join(
        str(cell.value)
        for row in workbook["AI-рекомендации"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Коммерческое соответствие: высокое" in workbook_text
    assert "инженерная подтвержденность: предварительно, требуется проверка" in workbook_text
    assert "инженерная подтвержденность: проверено" not in workbook_text


def test_excel_ai_recommendations_hide_fatal_invalid_recommendations(
    client: TestClient,
    db_session: Session,
) -> None:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [{"item_type": "server", "quantity": 2, "name": "server"}]},
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": False,
            "llm_fallback_reason": "llm_configurator_all_recommendations_rejected",
            "ai_recommendations": [
                {
                    "source_type": "build_from_parts",
                    "decision": "recommend",
                    "title": "Fatal build",
                    "display_name": "ASUS RS521A-E12-RS24U + Intel Xeon Gold 5220R",
                    "compatibility_warnings": [
                        "fatal socket mismatch: SP5 platform cannot use LGA3647 CPU"
                    ],
                }
            ],
            "component_candidate_matrix": {},
        },
        report_markdown="# report",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet_text = "\n".join(
        str(cell.value)
        for row in workbook["AI-рекомендации"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Fatal build" not in sheet_text
    assert "Intel Xeon Gold 5220R" not in sheet_text
    assert "llm_configurator_all_recommendations_rejected" in sheet_text


def test_excel_ai_recommendations_show_selected_duplicate_pool_recommendation(
    client: TestClient,
    db_session: Session,
) -> None:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text="Нужно 2 сервера",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [{"item_type": "server", "quantity": 2, "name": "server"}]},
        report_json={
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "llm_proposals_count": 5,
            "valid_proposals_count": 5,
            "ai_recommendations_count": 1,
            "rejected_ai_recommendations_count": 4,
            "validation_rejected_count": 0,
            "selection_skipped_count": 4,
            "ai_validation_summary": {
                "accepted": 1,
                "accepted_after_validation": 5,
                "validation_rejected_count": 0,
                "selection_skipped_count": 4,
                "selection_skipped_duplicate": 4,
                "rejected": 4,
            },
            "ai_recommendations": [
                {
                    "source_type": "build_from_parts",
                    "decision": "recommend",
                    "title": "Оптимальный по цене вариант",
                    "display_name": "ASUS + Intel",
                    "quantity_required": 2,
                    "total_price_value": "8800",
                    "total_price_currency": "USD",
                    "right_size_note": "Подбор: минимально подходящий по требованиям",
                    "why_selected": "Лучший представитель одинаковых BOM.",
                    "components": [
                        {
                            "role": "server_platform",
                            "producer": "ASUS",
                            "part_number": "PLATFORM",
                            "quantity_required": 2,
                        },
                        {
                            "role": "cpu",
                            "producer": "Intel",
                            "part_number": "CPU",
                            "quantity_required": 4,
                        },
                    ],
                }
            ],
            "component_candidate_matrix": {},
        },
        report_markdown="# report",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()

    response = client.get(f"/api/v1/match/{match_run.id}/report.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet_text = "\n".join(
        str(cell.value)
        for row in workbook["AI-рекомендации"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "1. Оптимальный по цене вариант" in sheet_text
    assert "ASUS PLATFORM" in sheet_text
    assert "AI проверил 5 вариантов, к показу выбрано 1." in sheet_text
    assert "Часть вариантов была скрыта как дубли или уступающие по цене/рискам." in sheet_text
    assert "Скрыто при выборе: 4 как дубли." in sheet_text
    assert "показано 0 безопасных рекомендаций" not in sheet_text
    assert "AI не смог сформировать безопасные рекомендации" not in sheet_text


def test_create_match_with_llm_composer_returns_api_and_excel_recommendations(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _seed_complete_component_set(db_session)
    monkeypatch.setenv("LLM_PROVIDER", "openai-compatible")
    monkeypatch.setenv("LLM_BASE_URL", "https://llm.example.test/v1")
    monkeypatch.setenv("LLM_API_KEY", "test-key")
    monkeypatch.setenv("LLM_MODEL", "test-model")
    monkeypatch.setenv("LLM_CONFIGURATOR_ENABLED", "true")
    monkeypatch.setenv("LLM_CONFIGURATOR_MODE", "composer")
    monkeypatch.setenv("LLM_CONFIGURATOR_TIMEOUT_SECONDS", "120")
    monkeypatch.setenv("LLM_CONFIGURATOR_READ_TIMEOUT_SECONDS", "900")
    monkeypatch.setenv("LLM_CONFIGURATOR_MAX_OUTPUT_TOKENS", "16384")
    get_llm_settings.cache_clear()

    from app.llm import configuration_composer

    monkeypatch.setattr(
        configuration_composer,
        "OpenAICompatibleLlmClient",
        _FakeOpenAIComposerClient,
    )

    response = client.post("/api/v1/match", json={"spec": _server_spec_payload()})

    assert response.status_code == 201
    body = response.json()
    assert body["llm_configurator_enabled"] is True
    assert body["llm_configurator_used"] is True
    assert body["ai_recommendation_mode"] == "ai_success"
    assert body["ai_recommendations_count"] == 1
    assert body["llm_proposals_count"] == 1
    assert body["valid_proposals_count"] == 1
    assert body["validation_rejected_count"] == 0
    assert body["selection_skipped_count"] == 0
    assert body["llm_fallback_reason"] is None
    assert body["llm_error_type"] is None
    assert body["llm_http_status"] is None
    assert len(body["llm_recommended_build_candidates"]) == 1
    llm_build = body["llm_recommended_build_candidates"][0]
    assert llm_build["total_price_value"] == "8800"
    assert llm_build["total_price_currency"] == "USD"
    assert llm_build["rank_reason"] == ["Balanced stocked configuration."]
    assert llm_build["optimization_mode"] == "cost_minimal_fit"
    assert llm_build["right_size_note"] == "Подбор: минимально подходящий по требованиям"
    assert body["build_candidates"]
    assert body["ai_validation_summary"]["accepted"] == 1
    assert body["ai_validation_summary"]["accepted_after_validation"] == 1
    assert body["ai_validation_summary"]["rejected"] == 0
    assert body["rejected_reasons_top"] == []
    assert body["output_mode"] == "single_best_cost_valid"
    assert body["primary_recommendation_status"] == "valid"
    assert body["primary_recommendation"]["candidate_type"] == "build_from_parts"
    assert body["primary_recommendation"]["engineering_confidence_code"] == (
        "preliminary_requires_engineer_review"
    )
    assert body["primary_recommendation"]["engineering_confidence"] == (
        "предварительно, нужна инженерная проверка"
    )
    assert body["commercial_summary"]
    assert "Сервер в сборе" in body["commercial_summary"]["copy_paste_text"]
    assert "Платформа" in body["commercial_summary"]["copy_paste_text"]
    assert "CPU" in body["commercial_summary"]["copy_paste_text"]
    assert "RAM" in body["commercial_summary"]["copy_paste_text"]
    assert "SSD" in body["commercial_summary"]["copy_paste_text"]
    assert "components" not in body["commercial_summary"]
    commercial_summary_text = json.dumps(body["commercial_summary"], ensure_ascii=False)
    for forbidden in (
        "component_candidate_id",
        '"facts"',
        '"evidence"',
        "raw JSON",
        "llm_rec",
        "why_this_platform:",
        "quote_recommendation:",
        "Engineering confidence",
        "Tradeoff",
        "Minimal cost",
        "Proven",
        "Premium platform",
        "高密度",
        "preliminary_requires_engineer_review",
    ):
        assert forbidden not in commercial_summary_text
    assert body["grouped_presales_mode_used"] is False
    assert body["configuration_groups_count"] == 0
    assert body["configuration_groups"] == []
    assert body["quote_recommendation"] == {}
    assert body["commercial_summary"]["title"] == "Предварительная спецификация для КП"
    assert body["commercial_summary"]["server_line"].startswith("Сервер в сборе")
    assert "per_server_lines" in body["commercial_summary"]
    assert "total_order_lines" in body["commercial_summary"]
    assert "comment_lines" in body["commercial_summary"]
    assert "engineer_checks" in body["commercial_summary"]
    assert "bom_rows" in body["commercial_summary"]

    report_response = client.get(f"/api/v1/match/{body['match_run_id']}/report.xlsx")
    workbook = load_workbook(BytesIO(report_response.content))
    assert workbook.sheetnames == ["AI-рекомендации", "Матрица компонентов"]
    llm_sheet = workbook["AI-рекомендации"]
    assert llm_sheet["A1"].value == "Предварительная спецификация для КП"
    assert llm_sheet["A4"].value == "Позиция"
    assert llm_sheet["D4"].value == "На 1 сервер"
    assert llm_sheet["F4"].value == "Всего к заказу"
    assert llm_sheet["G4"].value == "Остаток"
    assert llm_sheet["H4"].value == "Примечание"
    assert llm_sheet["A23"].value == "Служебные данные"
    assert "Конфигурационные семейства" not in [
        cell.value
        for row in llm_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert str(llm_sheet.freeze_panes).startswith("A")
    assert llm_sheet.column_dimensions["H"].width == 18
    llm_sheet_text = "\n".join(
        str(cell.value)
        for row in llm_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Компонентная база" not in llm_sheet_text
    assert "Варианты платформ" not in llm_sheet_text
    assert "Предварительная спецификация для КП" in llm_sheet_text
    assert "Сервер в сборе" in llm_sheet_text
    assert "Проверить перед КП" in llm_sheet_text
    assert any(height.height and height.height > 40 for height in llm_sheet.row_dimensions.values())
    assert "Ориентировочно за 2 сервера: 8 800 USD" in llm_sheet_text
    assert "8 800" in llm_sheet_text
    assert "USD" in llm_sheet_text
    assert "Рекомендуемый вариант для самого дешевого КП" not in llm_sheet_text
    component_sheet = workbook["Матрица компонентов"]
    component_header_row = int(str(component_sheet.freeze_panes).replace("A", "")) - 1
    component_headers = [cell.value for cell in component_sheet[component_header_row]]
    assert "Оценка соответствия" in component_headers
    assert "Почему подходит" in component_headers
    assert "Превышение требования" in component_headers
    assert "Fit label" not in component_headers
    assert component_sheet["A1"].value == "Охват матрицы"
    assert component_sheet.auto_filter.ref.startswith("A")
    assert component_sheet.column_dimensions["D"].width == 55
    assert component_sheet["H3"].alignment.wrap_text is True
    workbook_values = [
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    workbook_text = "\n".join(workbook_values)
    assert "AI-рекомендации сформированы автоматически по складским данным" in workbook_text
    assert "AI проверил 1 вариант, к показу выбрано 1." in workbook_text
    assert "обязательна инженерная проверка совместимости и комплектации" in workbook_text
    assert "на заказ" not in workbook_text
    assert "Fit label" not in workbook_text
    assert "overfit" not in workbook_text
    assert "cores" not in workbook_text
    assert "component_candidate_id" not in workbook_text
    assert "test-key" not in workbook_text
    assert _FakeOpenAIComposerClient.kwargs["timeout_seconds"] == 120
    assert _FakeOpenAIComposerClient.kwargs["read_timeout_seconds"] == 900
    assert _FakeOpenAIComposerClient.kwargs["max_output_tokens"] == 16384
    assert _FakeOpenAIComposerClient.kwargs["use_response_format"] is False


def test_excel_repair_notice_hides_raw_diagnostics(
    client: TestClient,
    db_session: Session,
) -> None:
    run_id = _seed_match_run(db_session)
    match_run = db_session.get(MatchRun, run_id)
    assert match_run is not None
    match_run.report_json = {
        "status": "partial_stock_matched",
        "llm_configurator_enabled": True,
        "llm_configurator_used": True,
        "grouped_presales_mode_used": True,
        "configuration_groups_count": 1,
        "configuration_groups": [_grouped_excel_fixture_group()],
        "quote_recommendation": {
            "for_cheapest_quote": "Gooxi GOOXI-CHEAP - 48 800 USD",
            "summary": "safe summary",
        },
        "llm_proposals_count": 2,
        "valid_proposals_count": 2,
        "ai_recommendations_count": 1,
        "llm_repair_used": True,
        "llm_repair_success": True,
        "llm_repair_critique_summary": [
            "component_candidate_id ram-secret raw JSON headers Authorization"
        ],
        "llm_repair_fallback_reason": "should-not-show",
    }
    db_session.commit()

    response = client.get(f"/api/v1/match/{run_id}/report.xlsx")
    workbook = load_workbook(BytesIO(response.content))
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "AI перепроверил цены по матрице" in workbook_text
    assert "component_candidate_id" not in workbook_text
    assert "ram-secret" not in workbook_text
    assert "raw JSON" not in workbook_text
    assert "Authorization" not in workbook_text
    assert "should-not-show" not in workbook_text


def test_list_match_runs_returns_recent_brief_source_text(
    client: TestClient,
    db_session: Session,
) -> None:
    older_id = _seed_match_run(
        db_session,
        source_text="older request",
        created_at=datetime(2026, 5, 9, 12, 0, tzinfo=UTC),
    )
    newer_id = _seed_match_run(
        db_session,
        source_text="newer request " + "x" * 160,
        created_at=datetime(2026, 5, 9, 13, 0, tzinfo=UTC),
    )

    response = client.get("/api/v1/match?limit=10")

    assert response.status_code == 200
    items = response.json()["items"]
    assert [item["id"] for item in items] == [newer_id, older_id]
    assert items[0]["source_text"].startswith("newer request")
    assert items[0]["source_text"].endswith("...")
    assert len(items[0]["source_text"]) == 120


def test_get_unknown_match_returns_404(client: TestClient) -> None:
    response = client.get("/api/v1/match/404")

    assert response.status_code == 404
    assert response.json()["detail"] == "Match run not found."


def test_create_match_empty_request_returns_422(client: TestClient) -> None:
    response = client.post("/api/v1/match", json={})

    assert response.status_code == 422
    assert response.json()["detail"] == "Request body must include non-empty 'text' or 'spec'."


def _server_spec_payload() -> dict[str, Any]:
    return {
        "items": [
            {
                "item_type": "server",
                "quantity": 2,
                "name": "server",
                "requirements": {
                    "form_factor": "2U",
                    "cpu": {"sockets": 2},
                    "ram": {"min_gb": 512},
                    "storage": {"type": "SSD"},
                    "power": {"psu_count": 2, "redundant_psu": True},
                },
            }
        ],
        "shipment_city": "Moscow",
        "requirements": {},
        "source_text": "Spec from stock-bot draft",
    }


def _worksheet_text(sheet: Any) -> str:
    return "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _seed_nerpa_products(db_session: Session) -> None:
    _seed_nerpa_product(
        db_session,
        item_id="1000841882",
        part_number="D5720-181125SA04",
        item_name="Server NERPA D5720 2U 2x CPU SSD 2x PSU 2x DDR4 64GB",
        quantity=3,
    )
    _seed_nerpa_product(
        db_session,
        item_id="1000841883",
        part_number="D5720-181125SA05",
        item_name="Server NERPA D5720 2U 2x CPU SSD 2x PSU 2x DDR4 32GB",
        quantity=1,
    )


def _seed_nerpa_product(
    db_session: Session,
    *,
    item_id: str,
    part_number: str,
    item_name: str,
    quantity: int,
) -> None:
    synced_at = datetime(2026, 5, 9, 12, 0, tzinfo=UTC)
    now = datetime(2026, 5, 9, 12, 5, tzinfo=UTC)
    db_session.add(
        DistributorProduct(
            distributor_code="ocs",
            item_id=item_id,
            product_key=item_id,
            part_number=part_number,
            producer="NERPA",
            category_id="V1100",
            item_name=item_name,
            item_name_rus="NERPA server",
            product_name=f"NERPA {part_number}",
            product_description=None,
            product_notes=None,
            hscode="8471490000",
            ean="04600000000012",
            is_in_mpt_registry=False,
            is_project_item=False,
            traceable=False,
            condition="Regular",
            warranty="Distributor warranty 12 months",
            original_country_iso_code="RU",
            vat_percent=Decimal("20"),
            serial_number_availability=None,
            catalog_path_json=[{"category_id": "V1100", "name": "Servers in assembly"}],
            package_json={"weight": 25.0},
            raw_json={"product": {"itemId": item_id, "partNumber": part_number}},
            synced_at=synced_at,
            created_at=now,
            updated_at=now,
        )
    )
    db_session.add(
        DistributorStockPrice(
            distributor_code="ocs",
            item_id=item_id,
            product_key=item_id,
            shipment_city="Moscow",
            location="MSK",
            location_description="Moscow",
            location_type="ShipmentCity",
            quantity_value=quantity,
            quantity_is_greater_than=False,
            can_reserve=True,
            departure_date=None,
            arrival_date=None,
            delivery_date=None,
            price_order_value=Decimal("6900"),
            price_order_currency="USD",
            price_list_value=Decimal("6900"),
            price_list_currency="USD",
            end_user_value=Decimal("7100"),
            end_user_currency="USD",
            raw_json={"productKey": item_id, "location": "MSK"},
            synced_at=synced_at,
            created_at=now,
        )
    )
    db_session.commit()


def _seed_component_product(
    db_session: Session,
    *,
    item_id: str,
    part_number: str,
    producer: str,
    category_id: str,
    item_name: str,
    quantity: int,
    price: Decimal,
) -> None:
    synced_at = datetime(2026, 5, 9, 12, 0, tzinfo=UTC)
    now = datetime(2026, 5, 9, 12, 5, tzinfo=UTC)
    db_session.add(
        DistributorProduct(
            distributor_code="ocs",
            item_id=item_id,
            product_key=item_id,
            part_number=part_number,
            producer=producer,
            category_id=category_id,
            item_name=item_name,
            item_name_rus=item_name,
            product_name=f"{producer} {part_number}",
            product_description=None,
            product_notes=None,
            hscode="8473308000",
            ean=None,
            is_in_mpt_registry=False,
            is_project_item=False,
            traceable=False,
            condition="Regular",
            warranty="Distributor warranty 12 months",
            original_country_iso_code="CN",
            vat_percent=Decimal("20"),
            serial_number_availability=None,
            catalog_path_json=[{"category_id": category_id, "name": "Server component"}],
            package_json={},
            raw_json={"product": {"itemId": item_id, "partNumber": part_number}},
            synced_at=synced_at,
            created_at=now,
            updated_at=now,
        )
    )
    db_session.add(
        DistributorStockPrice(
            distributor_code="ocs",
            item_id=item_id,
            product_key=item_id,
            shipment_city="Moscow",
            location="MSK",
            location_description="Moscow",
            location_type="ShipmentCity",
            quantity_value=quantity,
            quantity_is_greater_than=False,
            can_reserve=True,
            departure_date=None,
            arrival_date=None,
            delivery_date=None,
            price_order_value=price,
            price_order_currency="USD",
            price_list_value=price,
            price_list_currency="USD",
            end_user_value=price,
            end_user_currency="USD",
            raw_json={"productKey": item_id, "location": "MSK"},
            synced_at=synced_at,
            created_at=now,
        )
    )
    db_session.commit()


def _seed_match_71_network_products(db_session: Session) -> None:
    rows = [
        (
            "switch-good",
            "SW-48P-4SFP",
            "48-port 1G RJ45 PoE+ switch 740W 4 uplink 10G SFP+ L3 stacking",
            "1200",
        ),
        ("switch-5p", "SW-5P", "5-port desktop switch", "25"),
        ("switch-8p", "SW-8P", "8-port 1G PoE switch", "100"),
        ("switch-16p", "SW-16P", "16-port 1G PoE switch", "180"),
        ("switch-24p", "SW-24P", "24-port 1G PoE+ switch 2 uplink 10G SFP+", "260"),
    ]
    for item_id, part_number, name, price in rows:
        _seed_component_product(
            db_session,
            item_id=item_id,
            part_number=part_number,
            producer="NetVendor",
            category_id="V120100",
            item_name=name,
            quantity=10,
            price=Decimal(price),
        )


def _network_77_requirements() -> dict[str, Any]:
    return {
        "product_group": "network",
        "server_qty": 1,
        "device_qty": 1,
        "network_device_role": "switch",
        "location": "Moscow",
        "required_roles": ["switch"],
        "required_capabilities": [
            {
                "capability_id": "switch.48x1g.poe.4x10g.l3",
                "role": "switch",
                "hard": True,
                "parsed_requirements": {
                    "device_count": 1,
                    "port_count": 48,
                    "port_speed": "1GbE",
                    "port_media": "RJ45",
                    "uplink_count": 4,
                    "uplink_speed": "10GbE",
                    "uplink_media": "SFP+",
                    "poe_required": True,
                    "poe_standard": "PoE+",
                    "l3_required": True,
                },
            }
        ],
        "optional_capabilities": [
            {
                "capability_id": "switch.stacking",
                "role": "switch",
                "hard": False,
                "parsed_requirements": {"stacking_required": True},
            }
        ],
    }


def _seed_saved_network_report_run(db_session: Session) -> int:
    now = datetime(2026, 5, 16, 12, 0, tzinfo=UTC)
    requirements = _network_77_requirements()
    switch_candidate = {
        "candidate_id": "switch-origo",
        "component_candidate_id": "switch-origo",
        "role": "switch",
        "role_label": "готовый сервер",
        "producer": "Origo",
        "part_number": "OS3254P/370W/A1A",
        "name": "Origo OS3254P/370W/A1A 48x1G PoE+ 4x10G SFP+ L3",
        "available_quantity": 17,
        "price_value": "483.84",
        "price_currency": "USD",
        "quantity_required": 1,
        "extracted_facts": {
            "port_count": 48,
            "port_speed": "1GbE",
            "port_media": "RJ45",
            "uplink_count": 4,
            "uplink_speed": "10GbE",
            "uplink_media": "SFP+",
            "poe_supported": True,
            "poe_budget_w": 370,
            "poe_standard": "PoE+",
            "l3_supported": True,
            "stacking_supported": True,
        },
        "fit_label": "exact_or_close_fit",
        "fit_tier": "strong_fit",
        "fit_reason": "Закрывает 48 портов, PoE, uplink SFP+ и L3.",
    }
    match_run = MatchRun(
        source="text",
        source_text=NETWORK_MATCH_74_PLAIN_POE_TEXT,
        status="stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=1,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [], "shipment_city": "Moscow"},
        report_json={
            "status": "stock_matched",
            "product_group": "network",
            "primary_recommendation_status": "valid",
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "normalized_requirements": requirements,
            "primary_recommendation": {
                "product_group": "network",
                "candidate_type": "build_from_parts",
                "decision": "recommend",
                "components": [{**switch_candidate, "server_quantity": 1}],
                "total_price_value": "483.84",
                "total_price_currency": "USD",
            },
            "component_matrix_coverage_summary": {
                "total_products_by_role": {
                    "ready_server": 0,
                    "platform": 0,
                    "cpu": 0,
                    "ram": 0,
                    "switch": 3,
                },
                "eligible_products_by_role": {
                    "ready_server": 0,
                    "platform": 0,
                    "cpu": 0,
                    "ram": 0,
                    "switch": 1,
                },
                "sent_to_llm_by_role": {
                    "ready_server": 0,
                    "platform": 0,
                    "cpu": 0,
                    "ram": 0,
                    "switch": 1,
                },
                "omitted_by_role": {"switch": 0},
                "bucket_summary_by_role": {"switch": {"strong_fit": 1}},
                "limit_per_role": 100,
                "selection_strategy": "bucketed_broad_matrix_v3",
            },
            "component_candidate_matrix": {
                "product_group": "network",
                "normalized_requirements": requirements,
                "switch_candidates": [switch_candidate],
                "ready_server_candidates": [
                    {
                        "role": "ready_server",
                        "part_number": "READY-LEGACY",
                        "name": "Legacy ready server must be hidden",
                    }
                ],
                "cpu_candidates": [
                    {
                        "role": "cpu",
                        "part_number": "CPU-LEGACY",
                        "name": "Legacy CPU row must be hidden",
                    }
                ],
            },
            "ready_stock_candidates": [
                {
                    "candidate_id": "legacy-ready",
                    "part_number": "READY-LEGACY",
                    "item_name": "Legacy ready server must be hidden",
                    "matched_requirements": ["legacy"],
                }
            ],
        },
        report_markdown="# Network report\n",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()
    return match_run.id


def _seed_saved_storage_report_run(db_session: Session) -> int:
    now = datetime(2026, 5, 16, 12, 5, tzinfo=UTC)
    requirements = {
        "product_group": "storage",
        "server_qty": 1,
        "system_qty": 1,
        "location": "Moscow",
        "storage_min_capacity": "100 TB",
        "raw_capacity_tb": 120,
        "usable_capacity_tb": 100,
        "redundancy_level": "RAID6",
        "controller_count": 2,
        "drive_count": 24,
        "drive_capacity_tb": 7.68,
        "drive_type": "SSD",
        "drive_interface": "SAS",
        "host_protocol": "FC",
        "host_port_count": 4,
        "host_port_speed": "32G",
        "host_port_media": "SFP28",
        "support_required": True,
        "license_required": True,
        "warranty_months": 36,
        "required_roles": ["storage_system", "controller"],
    }
    match_run = MatchRun(
        source="text",
        source_text="Нужна СХД 100 ТБ usable, FC 32G, два контроллера",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [], "shipment_city": "Moscow"},
        report_json={
            "status": "partial_stock_matched",
            "product_group": "storage",
            "normalized_requirements": requirements,
            "component_matrix_coverage_summary": {
                "total_products_by_role": {
                    "ready_server": 0,
                    "cpu": 0,
                    "storage_system": 1,
                    "controller": 1,
                },
                "eligible_products_by_role": {"storage_system": 1, "controller": 1},
                "sent_to_llm_by_role": {"storage_system": 1, "controller": 1},
                "omitted_by_role": {"storage_system": 0, "controller": 0},
                "bucket_summary_by_role": {"storage_system": {"possible_fit": 1}},
                "limit_per_role": 100,
                "selection_strategy": "bucketed_broad_matrix_v3",
            },
            "component_candidate_matrix": {
                "product_group": "storage",
                "normalized_requirements": requirements,
                "storage_system_candidates": [
                    {
                        "role": "storage_system",
                        "producer": "StorageVendor",
                        "part_number": "ARR-100U",
                        "name": "StorageVendor ARR-100U 100TB FC 32G",
                        "available_quantity": 2,
                        "price_value": "5000",
                        "price_currency": "USD",
                        "extracted_facts": {
                            "usable_capacity_tb": 100,
                            "raw_capacity_tb": 120,
                            "host_protocol": "FC",
                            "host_port_count": 4,
                            "host_port_speed": "32G",
                        },
                        "fit_tier": "possible_fit",
                    }
                ],
                "controller_candidates": [
                    {
                        "role": "controller",
                        "producer": "StorageVendor",
                        "part_number": "CTRL-2",
                        "name": "Dual controller option",
                        "available_quantity": 2,
                    }
                ],
                "cpu_candidates": [
                    {
                        "role": "cpu",
                        "part_number": "CPU-LEGACY",
                        "name": "Legacy CPU row must be hidden",
                    }
                ],
            },
        },
        report_markdown="# Storage report\n",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()
    return match_run.id


def _seed_saved_server_report_run(db_session: Session) -> int:
    now = datetime(2026, 5, 16, 12, 10, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text=SERVER_REQUEST,
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"items": [], "shipment_city": "Moscow"},
        report_json={
            "status": "partial_stock_matched",
            "product_group": "server",
            "normalized_requirements": {
                "product_group": "server",
                "server_qty": 2,
                "form_factor": "2U",
                "cpu_per_server": 2,
                "total_cpu_required": 4,
                "ram_gb_per_server": 512,
                "storage_type_preference": "SSD",
                "storage_qty_per_server": 2,
                "location": "Moscow",
            },
            "component_candidate_matrix": {
                "product_group": "server",
                "ready_server_candidates": [
                    {
                        "role": "ready_server",
                        "role_label": "готовый сервер",
                        "producer": "NERPA",
                        "part_number": "READY-SERVER",
                        "name": "NERPA READY-SERVER",
                        "available_quantity": 2,
                        "price_value": "6900",
                        "price_currency": "USD",
                    }
                ],
            },
        },
        report_markdown="# Server report\n",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.commit()
    return match_run.id


def _seed_complete_component_set(db_session: Session) -> None:
    _seed_component_product(
        db_session,
        item_id="platform-1",
        part_number="PLATFORM-2U-2S",
        producer="ASUS",
        category_id="V110100",
        item_name="ASUS 2U dual socket DDR5 server platform 2x PSU",
        quantity=2,
        price=Decimal("2000"),
    )
    _seed_component_product(
        db_session,
        item_id="cpu-1",
        part_number="CPU-SERVER",
        producer="Intel",
        category_id="V110103",
        item_name="Intel Xeon server CPU",
        quantity=4,
        price=Decimal("500"),
    )
    _seed_component_product(
        db_session,
        item_id="ram-64",
        part_number="RAM-64G",
        producer="Samsung",
        category_id="V110104",
        item_name="DDR5 RDIMM 64GB server memory module",
        quantity=16,
        price=Decimal("150"),
    )
    _seed_component_product(
        db_session,
        item_id="ssd-1",
        part_number="SSD-960G",
        producer="Samsung",
        category_id="V110106",
        item_name="Server SSD 960GB SATA",
        quantity=2,
        price=Decimal("200"),
    )


def _seed_complete_extended_component_set(db_session: Session) -> None:
    _seed_complete_component_set(db_session)
    _seed_component_product(
        db_session,
        item_id="platform-1u",
        part_number="PLATFORM-1U-2S",
        producer="ASUS",
        category_id="V110100",
        item_name="ASUS 1U dual socket DDR5 8xSFF server platform 2x PSU",
        quantity=2,
        price=Decimal("2500"),
    )
    _seed_component_product(
        db_session,
        item_id="ssd-1920",
        part_number="SSD-1920-SATA",
        producer="KIOXIA",
        category_id="V110106",
        item_name="Server SSD 1920GB SATA mixed use drive",
        quantity=16,
        price=Decimal("220"),
    )
    _seed_component_product(
        db_session,
        item_id="controller-1",
        part_number="LSI-9400-8I",
        producer="LSI",
        category_id="V110107",
        item_name="LSI Logic 9400-8i SAS SATA storage controller",
        quantity=2,
        price=Decimal("400"),
    )
    _seed_component_product(
        db_session,
        item_id="psu-1",
        part_number="PSU-2000W",
        producer="Delta",
        category_id="V110108",
        item_name="2000W hot-swap Platinum redundant server power supply",
        quantity=4,
        price=Decimal("300"),
    )
    _seed_component_product(
        db_session,
        item_id="cable-1",
        part_number="C13-C14",
        producer="CableCo",
        category_id="V110109",
        item_name="C13-C14 server power cable",
        quantity=8,
        price=Decimal("20"),
    )


def _seed_server_81_component_set(db_session: Session) -> None:
    rows = [
        (
            "platform-81",
            "PLATFORM-1U-2S",
            "ASUS",
            "V110100",
            "ASUS 1U dual socket Intel Xeon DDR5 8xSFF server platform 2x2000W PSU",
            2,
            "2500",
        ),
        (
            "cpu-81",
            "CPU-6GEN-24C",
            "Intel",
            "V110103",
            "Intel Xeon 6th generation 24 core server CPU",
            4,
            "700",
        ),
        (
            "ram-81",
            "RAM-DDR5-64G",
            "Samsung",
            "V110104",
            "DDR5 RDIMM 64GB server memory module",
            16,
            "150",
        ),
        (
            "ssd-81",
            "SSD-1920-SATA",
            "KIOXIA",
            "V110106",
            "Server SSD 1920GB SATA mixed use drive",
            16,
            "220",
        ),
        (
            "controller-81",
            "LSI-9400-8I",
            "LSI",
            "V110107",
            "LSI Logic 9400-8i SAS SATA storage controller",
            2,
            "400",
        ),
        (
            "nic-81",
            "X710-DA2",
            "Intel",
            "V120116",
            "Intel X710-DA2 dual port 10GbE SFP+ server network adapter",
            2,
            "200",
        ),
        (
            "psu-81",
            "PSU-2000W",
            "Delta",
            "V110108",
            "2000W hot-swap Platinum redundant server power supply",
            4,
            "300",
        ),
        (
            "cable-81",
            "C13-C14",
            "CableCo",
            "V110109",
            "C13-C14 server power cable",
            8,
            "20",
        ),
    ]
    for item_id, part_number, producer, category_id, item_name, quantity, price in rows:
        _seed_component_product(
            db_session,
            item_id=item_id,
            part_number=part_number,
            producer=producer,
            category_id=category_id,
            item_name=item_name,
            quantity=quantity,
            price=Decimal(price),
        )


class _ApiSemanticPlannerClient:
    def __init__(self, settings: object, **kwargs: Any) -> None:
        self.settings = settings
        self.kwargs = kwargs

    def close(self) -> None:
        return None

    def generate_json(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        if "AI Semantic Matrix Planner V2" in system_prompt:
            json.loads(user_prompt)
            return _semantic_server_78_payload()
        if "Distributor Category Planner" in system_prompt:
            return {
                "category_plan": [
                    {
                        "role": "server_platform",
                        "selected_category_ids": ["V110100"],
                        "purpose": "base_device",
                        "capability_ids": ["server_platform.1u.2s.8sff"],
                        "hard_optional_relation": "hard",
                        "reason": "server platform category",
                        "confidence": "high",
                    },
                    {
                        "role": "cpu",
                        "selected_category_ids": ["V110103"],
                        "purpose": "component",
                        "capability_ids": ["cpu.intel"],
                        "hard_optional_relation": "hard",
                        "reason": "server CPU category",
                        "confidence": "high",
                    },
                    {
                        "role": "ram",
                        "selected_category_ids": ["V110104"],
                        "purpose": "component",
                        "capability_ids": ["ram.ddr5"],
                        "hard_optional_relation": "hard",
                        "reason": "server RAM category",
                        "confidence": "high",
                    },
                    {
                        "role": "storage",
                        "selected_category_ids": ["V110106"],
                        "purpose": "drive",
                        "capability_ids": ["storage.sata"],
                        "hard_optional_relation": "hard",
                        "reason": "server storage category",
                        "confidence": "high",
                    },
                    {
                        "role": "network_adapter",
                        "selected_category_ids": ["V120116"],
                        "purpose": "component",
                        "capability_ids": ["network_adapter.10gbe.sfpplus.x710_da2"],
                        "hard_optional_relation": "hard",
                        "reason": "server NIC category",
                        "confidence": "high",
                    },
                    {
                        "role": "storage_controller",
                        "selected_category_ids": ["V110107"],
                        "purpose": "component",
                        "capability_ids": ["storage_controller.hba"],
                        "hard_optional_relation": "hard",
                        "reason": "server HBA category",
                        "confidence": "high",
                    },
                    {
                        "role": "power_supply",
                        "selected_category_ids": ["V110108"],
                        "purpose": "component",
                        "capability_ids": ["power_supply.2000w"],
                        "hard_optional_relation": "hard",
                        "reason": "server PSU category",
                        "confidence": "high",
                    },
                    {
                        "role": "cable",
                        "selected_category_ids": ["V110109"],
                        "purpose": "cable",
                        "capability_ids": ["cable.c13_c14"],
                        "hard_optional_relation": "hard",
                        "reason": "server power cable category",
                        "confidence": "high",
                    },
                ],
                "missing_category_roles": [],
                "category_plan_warnings": [],
            }
        raise AssertionError(f"unexpected LLM prompt: {system_prompt[:80]}")


class _ApiServer81OpenAIClient:
    composer_calls = 0
    composer_packages: list[dict[str, Any]] = []

    def __init__(self, settings: object, **kwargs: Any) -> None:
        self.settings = settings
        self.kwargs = kwargs

    @classmethod
    def reset(cls) -> None:
        cls.composer_calls = 0
        cls.composer_packages = []

    def close(self) -> None:
        return None

    def generate_json(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        if "AI Semantic Matrix Planner V2" in system_prompt:
            json.loads(user_prompt)
            return _api_server_81_semantic_payload()
        if "Distributor Category Planner" in system_prompt:
            return _api_server_81_category_response()
        if "Online Composer V1" in system_prompt or "LLM Configuration Composer" in system_prompt:
            package = json.loads(user_prompt)
            type(self).composer_calls += 1
            type(self).composer_packages.append(package)
            return _api_server_81_online_response(package)
        raise AssertionError(f"unexpected LLM prompt: {system_prompt[:80]}")


def _api_server_81_semantic_payload() -> dict[str, Any]:
    roles = [
        "server_platform",
        "cpu",
        "ram",
        "storage",
        "storage_controller",
        "network_adapter",
        "power_supply",
        "cable",
    ]
    return {
        "primary_product_group": "server",
        "primary_object": "server",
        "confidence": "high",
        "classification_reason": "Server BOM with component categories.",
        "matrix_blueprint": {
            "roles": [
                {
                    "role": role,
                    "required": True,
                    "source_text": role,
                    "characteristics_to_match": {},
                    "hard_capability_ids": [],
                }
                for role in roles
            ]
        },
        "required_capabilities": [],
        "optional_capabilities": [],
        "embedded_requirements": [],
        "not_primary_product_groups": [],
        "logistics_constraints": {},
        "commercial_instructions": [],
        "response_instructions": [],
        "engineer_review_instructions": [],
        "unsupported_or_unmapped_requirements": [],
        "planner_warnings": [],
    }


def _api_server_81_category_response() -> dict[str, Any]:
    role_categories = {
        "server_platform": "V110100",
        "cpu": "V110103",
        "ram": "V110104",
        "storage": "V110106",
        "storage_controller": "V110107",
        "network_adapter": "V120116",
        "power_supply": "V110108",
        "cable": "V110109",
    }
    return {
        "category_plan": [
            {
                "role": role,
                "selected_category_ids": [category_id],
                "purpose": "component",
                "capability_ids": [],
                "hard_optional_relation": "hard",
                "reason": f"{role} category",
                "confidence": "high",
            }
            for role, category_id in role_categories.items()
        ],
        "missing_category_roles": [],
        "category_plan_warnings": [],
    }


def _semantic_server_78_payload() -> dict[str, Any]:
    role_plan = _semantic_server_78_role_plan()
    return {
        "primary_product_group": "server",
        "primary_object": "server",
        "confidence": "high",
        "classification_reason": "The request is a server BOM with embedded NIC/cables.",
        "matrix_blueprint": role_plan["matrix_blueprint"],
        "required_capabilities": [],
        "optional_capabilities": [],
        "embedded_requirements": role_plan["embedded_requirements"],
        "not_primary_product_groups": role_plan["not_primary_product_groups"],
        "logistics_constraints": {},
        "commercial_instructions": [],
        "response_instructions": [],
        "engineer_review_instructions": [],
        "unsupported_or_unmapped_requirements": [],
        "planner_warnings": [],
    }


def _semantic_server_78_role_plan() -> dict[str, Any]:
    required_capabilities = [
        {
            "capability_id": "server_platform.1u.2s.8sff",
            "role": "server_platform",
            "source_text": "1U, 2 sockets, 8 SFF slots",
            "hard": True,
            "parsed_requirements": {
                "form_factor": "1U",
                "socket_count": 2,
                "front_sff_slots_min": 8,
            },
        },
        {
            "capability_id": "network_adapter.10gbe.sfpplus.x710_da2",
            "role": "network_adapter",
            "source_text": "Intel X710-DA2 2x10GbE SFP+",
            "hard": True,
            "parsed_requirements": {
                "model_hint": "Intel X710-DA2",
                "min_ports_per_server": 2,
                "speed": "10GbE",
                "media": "SFP+",
            },
        },
        {
            "capability_id": "power_cable.c13",
            "role": "cable",
            "original_role": "power_cable",
            "source_text": "C13-C14 and C13-Schuko power cables",
            "hard": True,
            "parsed_requirements": {
                "cable_type": "power",
                "connector_types": ["C13-C14", "C13-Schuko"],
            },
        },
    ]
    matrix_blueprint = {
        "roles": [
            {
                "role": "server_platform",
                "required": True,
                "source_text": "1U, 2 sockets, 8 SFF slots",
                "characteristics_to_match": {
                    "form_factor": "1U",
                    "socket_count": 2,
                },
                "hard_capability_ids": ["server_platform.1u.2s.8sff"],
            },
            {
                "role": "network_adapter",
                "required": True,
                "source_text": "Intel X710-DA2 2x10GbE SFP+",
                "characteristics_to_match": {
                    "speed": "10GbE",
                    "media": "SFP+",
                },
                "hard_capability_ids": ["network_adapter.10gbe.sfpplus.x710_da2"],
            },
            {
                "role": "cable",
                "required": True,
                "source_text": "C13-C14 and C13-Schuko power cables",
                "characteristics_to_match": {
                    "cable_type": "power",
                },
                "hard_capability_ids": ["power_cable.c13"],
                "original_role": "power_cable",
            },
        ]
    }
    return {
        "product_group": "server",
        "primary_product_group": "server",
        "primary_object": "server",
        "semantic_planner_source": "llm",
        "semantic_planner_confidence": "high",
        "selected_product_group_reason": "The request is a server BOM.",
        "deterministic_product_group_hint": "network",
        "semantic_planner_disagreement": True,
        "matrix_blueprint": matrix_blueprint,
        "matrix_blueprint_roles": ["server_platform", "network_adapter", "cable"],
        "embedded_requirements": [
            {
                "product_group": "network",
                "role": "network_adapter",
                "reason": "SFP+ belongs to the server NIC.",
            }
        ],
        "not_primary_product_groups": [
            {
                "product_group": "network",
                "reason": "Not a standalone switch/cable request.",
            }
        ],
        "requirements": [],
        "required_capabilities": required_capabilities,
        "optional_capabilities": [],
        "required_roles": ["server_platform", "network_adapter", "cable"],
        "optional_roles": [],
        "requirements_by_role": {
            row["role"]: row["parsed_requirements"] | {"required": True}
            for row in required_capabilities
        },
        "unsupported_or_unmapped_requirements": [],
        "planner_warnings": [],
        "role_catalog": ["server_platform", "network_adapter", "cable"],
    }


class _ApiDirectComposerClient:
    def __init__(self, responder: Any) -> None:
        self._responder = responder
        self.package: dict[str, Any] = {}

    def generate_json(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        assert "Do not invent products" in system_prompt
        self.package = json.loads(user_prompt)
        return self._responder(self.package)


def _api_server_78_requirements() -> dict[str, Any]:
    return {
        "product_group": "server",
        "server_qty": 2,
        "cpu_per_server": 2,
        "total_cpu_required": 4,
        "ram_gb_per_server": 512,
        "ram_type_preference": "DDR5",
        "storage_required": True,
        "storage_type_preference": "SSD",
        "storage_qty_per_server": 2,
        "network_required": True,
        "network_min_ports_per_server": 2,
        "network_speed": "10GbE",
        "network_media": "SFP+",
        "psu_count_per_server": 2,
        "required_roles": [
            "server_platform",
            "cpu",
            "ram",
            "ssd",
        ],
        "required_capabilities": [],
    }


def _api_server_78_like_matrix() -> dict[str, Any]:
    role_specs = [
        ("platform_candidates", "server_platform", "platform", "ASUS", "V110100"),
        ("cpu_candidates", "cpu", "cpu", "Intel", "V110103"),
        ("ram_candidates", "ram", "ram", "Samsung", "V110104"),
        ("ssd_candidates", "ssd", "storage", "KIOXIA", "V110106"),
        (
            "storage_controller_candidates",
            "storage_controller",
            "storage_controller",
            "LSI",
            "V110107",
        ),
        (
            "network_adapter_candidates",
            "network_adapter",
            "network_adapter",
            "Intel",
            "V120116",
        ),
        (
            "power_supply_candidates",
            "power_supply",
            "power_supply",
            "Delta",
            "V110108",
        ),
        ("cable_candidates", "cable", "cable", "CableCo", "V110109"),
    ]
    category_plan = {
        "server_platform": ["V110100"],
        "cpu": ["V110103"],
        "ram": ["V110104"],
        "storage": ["V110106"],
        "storage_controller": ["V110107"],
        "network_adapter": ["V120116"],
        "power_supply": ["V110108"],
        "cable": ["V110109"],
    }
    matrix: dict[str, Any] = {
        "product_group": "server",
        "primary_object": "server",
        "semantic_planner_source": "llm",
        "semantic_planner_used": True,
        "category_planner_source": "ai_category_planner",
        "category_plan_source": "llm",
        "category_plan": category_plan,
        "category_plan_entries": [
            {"role": role, "category_ids": category_ids}
            for role, category_ids in category_plan.items()
        ],
        "category_catalog_summary": {"distributor_code": "ocs"},
        "role_plan": _api_server_78_requirements(),
        "required_capabilities": [],
        "required_roles": _api_server_78_requirements()["required_roles"],
        "role_coverage_summary": {
            role: {"required": True, "missing": False, "sent_to_llm_count": 1}
            for _, role, _, _, _ in role_specs
        },
        "matrix_distiller_used": False,
        "matrix_distiller_source": "skipped",
        "matrix_distiller_diagnostics": {
            "reason": "package_within_budget_or_not_distillable",
            "package_skipped_reason": None,
        },
        "broad_count_by_role": {
            role: 1
            for _, role, _, _, _ in role_specs
        },
    }
    for matrix_key, role, output_role, producer, category_id in role_specs:
        matrix[matrix_key] = [
            _api_component_candidate(
                role=role,
                output_role=output_role,
                producer=producer,
                category_id=category_id,
            )
        ]
    return matrix


def _api_component_candidate(
    *,
    role: str,
    output_role: str,
    producer: str,
    category_id: str,
) -> dict[str, Any]:
    facts_by_role = {
        "server_platform": {"socket_count": 2, "memory_type": "DDR5"},
        "cpu": {"cpu_cores": 24, "normalized_vendor": "Intel"},
        "ram": {"ram_capacity_gb": 64, "ram_type": "DDR5"},
        "ssd": {"storage_capacity_tb": 1.92, "storage_interface": "SATA"},
        "storage_controller": {"controller_count": 1, "storage_interface": "SAS"},
        "network_adapter": {
            "network_ports_count": 2,
            "network_speed": "10GbE",
            "network_speed_gbps": 10,
            "network_media": "SFP+",
            "network_interface": "SFP+",
        },
        "power_supply": {"power_w": 2000, "redundant_psu": True},
        "cable": {"cable_type": "C13-C14"},
    }
    return {
        "component_candidate_id": f"{role}-1",
        "role": role,
        "category_id": category_id,
        "category_name": f"{role} category",
        "distributor_code": "ocs",
        "item_id": f"{role}-item-1",
        "product_key": f"ocs:{role}-item-1",
        "producer": producer,
        "part_number": f"{output_role.upper()}-1",
        "item_name": f"{producer} {role} candidate",
        "name": f"{producer} {role} candidate",
        "available_quantity": 100,
        "price_value": str(Decimal("100")),
        "price_currency": "USD",
        "fit_tier": "strong_fit",
        "score": 100,
        "extracted_facts": facts_by_role.get(role, {}),
    }


def test_composer_package_keeps_87_cooling_feature_non_blocking() -> None:
    matrix = _api_server_78_like_matrix()
    matrix["platform_candidates"][0]["package_json"] = {
        "contents": ["platform chassis", "accessory kit"],
        "weight": 25,
    }
    matrix["platform_candidates"][0]["ocs_content_properties"] = [
        {"name": "Комплект поставки", "value": "platform chassis, accessory kit"}
    ]
    cooling_requirement = {
        "requirement_id": "req_cooling",
        "source_text": "Cooling: 8 fans N+1",
        "classification": "primary_object_feature",
        "product_group": "server",
        "target_role": "server_platform",
        "target_primary_object": "server",
        "hard_or_optional": "hard",
        "reason": "Cooling N+1 is a server platform feature.",
        "confidence": "high",
        "should_block_before_composer": False,
        "should_appear_in_composer_brief": True,
        "should_be_validated_after_composer": True,
        "category_needed": False,
        "parsed_requirements": {"fan_count": 8, "redundancy": "N+1"},
    }
    role_plan = {
        **matrix["role_plan"],
        "classified_requirements": [cooling_requirement],
        "primary_object_feature_requirements": [cooling_requirement],
        "unmapped_requirements_blocking": [],
        "unmapped_requirements_non_blocking": [],
        "required_capabilities": [
            {
                "capability_id": "server_platform.cooling_nplus1",
                "role": "server_platform",
                "source_text": "Cooling: 8 fans N+1",
                "requirement_text": "Cooling: 8 fans N+1",
                "hard": True,
                "parsed_requirements": {"fan_count": 8, "redundancy": "N+1"},
                "requirement_classification": "primary_object_feature",
                "category_needed": False,
                "should_block_before_composer": False,
                "should_be_validated_after_composer": True,
            }
        ],
    }
    matrix["role_plan"] = role_plan
    matrix["classified_requirements"] = [cooling_requirement]
    matrix["primary_object_feature_requirements"] = [cooling_requirement]
    matrix["required_capabilities"] = role_plan["required_capabilities"]
    matrix["required_roles"] = ["server_platform", "cpu", "ram", "ssd"]
    matrix["missing_required_roles"] = ["unmapped"]
    matrix["missing_required_roles_before_llm"] = ["unmapped"]
    matrix["missing_required_capabilities"] = [
        {
            "capability_id": "cooling.8fans.nplus1",
            "role": "unmapped",
            "status": "missing_category",
            "source_text": "Cooling: 8 fans N+1",
            "requirement_classification": "primary_object_feature",
            "category_needed": False,
        }
    ]

    package = build_llm_configurator_package(
        user_request=COMPLEX_SERVER_78_TEXT,
        normalized_requirements=_api_server_78_requirements(),
        ready_stock_candidates=[],
        component_candidate_matrix=matrix,
        rule_based_build_candidates=[],
        max_package_chars=1_500_000,
    )

    assert "unmapped" not in package["required_roles"]
    assert package["missing_required_roles_before_llm"] == []
    assert package["missing_required_capabilities_before_llm"] == []
    assert package["classified_requirements"][0]["classification"] == "primary_object_feature"
    assert package["classified_requirements"][0]["fulfillment_mode"] == "included_in_primary_object"
    assert package["primary_object_feature_requirements"][0]["target_role"] == "server_platform"
    assert package["component_candidate_matrix"]["platform"][0]["package_json"]["weight"] == 25
    assert package["component_candidate_matrix"]["platform"][0]["content_properties"]


def test_platform_feature_unknown_is_unverified_not_satisfied() -> None:
    capability = {
        "capability_id": "server_platform.cooling_nplus1",
        "role": "server_platform",
        "source_text": "Cooling: 8 fans N+1",
        "hard": True,
        "parsed_requirements": {"fan_count": 8, "redundancy": "N+1"},
        "requirement_classification": "primary_object_feature",
    }
    selected = {
        "server_platform": composer_module._IndexedComponentCandidate(
            component_candidate_id="platform-1",
            prompt_role="platform",
            internal_role="server_platform",
            row={"component_candidate_id": "platform-1", "extracted_facts": {}},
            source={},
        )
    }

    rows = composer_module._hard_capability_validation(
        selected=selected,
        quantities={"server_platform": 1},
        normalized_requirements={
            "product_group": "server",
            "required_capabilities": [capability],
        },
    )

    assert rows[0]["status"] == "unverified_hard_requirement"
    assert rows[0]["requirement_classification"] == "primary_object_feature"


def test_bundle_fulfillment_without_evidence_is_unverified_not_satisfied() -> None:
    selected = {
        "server_platform": composer_module._IndexedComponentCandidate(
            component_candidate_id="platform-1",
            prompt_role="platform",
            internal_role="server_platform",
            row={"component_candidate_id": "platform-1", "available_quantity": 1},
            source={},
        )
    }
    classified = {
        "requirement_id": "bundle_accessory",
        "source_text": "required accessory is in the bundle",
        "classification": "accessory_or_consumable",
        "target_role": "other_accessory",
        "fulfillment_target_role": "server_platform",
        "hard_or_optional": "hard",
        "fulfillment_mode": "included_in_bundle_or_kit",
        "should_create_bom_role": False,
        "should_validate_after_composer": True,
        "engineer_check_ru": "Подтвердить комплектность в карточке товара.",
    }

    rows = composer_module._hard_capability_validation(
        selected=selected,
        quantities={"server_platform": 1},
        normalized_requirements={
            "product_group": "server",
            "classified_requirements": [classified],
        },
    )

    assert rows[0]["status"] == "unverified_hard_requirement"
    assert rows[0]["fulfillment_mode"] == "included_in_bundle_or_kit"
    assert rows[0]["satisfied_by"] is None


def test_bundle_fulfillment_with_evidence_can_be_satisfied_by_selected_target() -> None:
    selected = {
        "server_platform": composer_module._IndexedComponentCandidate(
            component_candidate_id="platform-1",
            prompt_role="platform",
            internal_role="server_platform",
            row={"component_candidate_id": "platform-1", "available_quantity": 1},
            source={},
        )
    }
    classified = {
        "requirement_id": "bundle_accessory",
        "source_text": "required accessory is in the bundle",
        "classification": "accessory_or_consumable",
        "target_role": "other_accessory",
        "fulfillment_target_role": "server_platform",
        "hard_or_optional": "hard",
        "fulfillment_mode": "included_in_bundle_or_kit",
        "evidence_source": "package_json",
        "evidence_text": "Package contents include the requested accessory.",
        "should_create_bom_role": False,
        "should_validate_after_composer": True,
    }

    rows = composer_module._hard_capability_validation(
        selected=selected,
        quantities={"server_platform": 1},
        normalized_requirements={
            "product_group": "server",
            "classified_requirements": [classified],
        },
    )

    assert rows[0]["status"] == "satisfied"
    assert rows[0]["satisfied_by"] == "bundle_or_kit"
    assert rows[0]["evidence_text"] == "Package contents include the requested accessory."


def test_platform_feature_contradiction_is_hard_mismatch() -> None:
    capability = {
        "capability_id": "server_platform.cooling_nplus1",
        "role": "server_platform",
        "source_text": "Cooling: 8 fans N+1",
        "hard": True,
        "parsed_requirements": {"fan_count": 8, "redundancy": "N+1"},
        "requirement_classification": "primary_object_feature",
    }
    selected = {
        "server_platform": composer_module._IndexedComponentCandidate(
            component_candidate_id="platform-1",
            prompt_role="platform",
            internal_role="server_platform",
            row={
                "component_candidate_id": "platform-1",
                "extracted_facts": {"fan_count": 4, "redundancy": "N+1"},
            },
            source={},
        )
    }

    rows = composer_module._hard_capability_validation(
        selected=selected,
        quantities={"server_platform": 1},
        normalized_requirements={
            "product_group": "server",
            "required_capabilities": [capability],
        },
    )

    assert rows[0]["status"] == "hard_mismatch"
    assert rows[0]["component_candidate_id"] == "platform-1"


def _api_server_78_online_response(package: dict[str, Any]) -> dict[str, Any]:
    matrix = package["component_candidate_matrix"]
    component_ids = {
        "platform": matrix["platform"][0]["component_candidate_id"],
        "cpu": matrix["cpu"][0]["component_candidate_id"],
        "ram": matrix["ram"][0]["component_candidate_id"],
        "storage": matrix["ssd"][0]["component_candidate_id"],
        "network_adapter": matrix["network_adapter"][0]["component_candidate_id"],
    }
    return {
        "primary_recommendation": {
            "candidate_type": "build_from_parts",
            "title": "Server #78 test build",
            "component_candidate_ids": component_ids,
            "quantities": {
                "platform": 2,
                "cpu": 4,
                "ram": 16,
                "storage": 4,
                "network_adapter": 2,
            },
            "why_selected": "Complete stocked server #78-like build.",
            "engineer_checks": ["Check platform support list."],
            "evidence_summary": {
                "status": "confirmed",
                "sources_count": 1,
                "confirmed_facts": ["Selected from matrix."],
                "source_domains": ["example.test"],
            },
        },
        "general_notes": [],
    }


def _api_server_81_online_response(package: dict[str, Any]) -> dict[str, Any]:
    matrix = package["component_candidate_matrix"]
    component_ids = {
        "platform": matrix["platform"][0]["component_candidate_id"],
        "cpu": matrix["cpu"][0]["component_candidate_id"],
        "ram": matrix["ram"][0]["component_candidate_id"],
        "storage": matrix["ssd"][0]["component_candidate_id"],
        "storage_controller": matrix["storage_controller"][0]["component_candidate_id"],
        "network_adapter": matrix["network_adapter"][0]["component_candidate_id"],
        "power_supply": matrix["power_supply"][0]["component_candidate_id"],
        "cable": matrix["cable"][0]["component_candidate_id"],
    }
    return {
        "primary_recommendation": {
            "candidate_type": "build_from_parts",
            "title": "Server #81 test build",
            "component_candidate_ids": component_ids,
            "quantities": {
                "platform": 2,
                "cpu": 4,
                "ram": 8,
                "storage": 8,
                "storage_controller": 2,
                "network_adapter": 2,
                "power_supply": 4,
                "cable": 4,
            },
            "why_selected": "Complete stocked server #81-like build.",
            "engineer_checks": ["Check platform support list."],
            "evidence_summary": {
                "status": "confirmed",
                "sources_count": 1,
                "confirmed_facts": ["Selected from matrix."],
                "source_domains": ["example.test"],
            },
        },
        "general_notes": [],
    }


class _FakeOpenAIComposerClient:
    kwargs: dict[str, Any] = {}

    def __init__(self, settings: object, **kwargs: Any) -> None:
        self.settings = settings
        type(self).kwargs = kwargs

    def close(self) -> None:
        return None

    def generate_json(self, system_prompt: str, user_prompt: str) -> dict[str, Any]:
        assert "Do not invent products" in system_prompt
        package = json.loads(user_prompt)
        source_candidate = package["rule_based_build_candidates"][0]
        return {
            "recommendations": [
                {
                    "recommendation_id": "llm_build_1",
                    "source_type": "build_from_parts",
                    "source_candidate_id": source_candidate["candidate_id"],
                    "decision": "recommend",
                    "title": "Recommended build 1",
                    "display_name": "Recommended build 1",
                    "components": {
                        "platform": source_candidate["platform"]["part_number"],
                        "cpu": "CPU",
                        "ram": "RAM",
                        "storage": "SSD",
                    },
                    "quantities": {"platform": 2, "cpu": 4, "ram": 16, "ssd": 2},
                    "total_price_value": None,
                    "total_price_currency": None,
                    "price_note": "за весь запрос",
                    "why_selected": "Balanced stocked configuration.",
                    "why_selected_short": "Balanced stocked configuration.",
                    "right_size_note": "Подбор: минимально подходящий по требованиям",
                    "what_is_missing": [],
                    "critical_checks": ["Check platform support list."],
                    "engineering_review_required": True,
                    "confidence": "medium",
                }
            ],
            "general_notes": ["CPU choice is preliminary."],
        }


def _seed_match_run(
    db_session: Session,
    *,
    source_text: str = SERVER_REQUEST,
    created_at: datetime | None = None,
) -> int:
    now = created_at or datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    match_run = MatchRun(
        source="text",
        source_text=source_text,
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=["RAM below requirement"],
        risk_flags_json=["engineer_review_required"],
        spec_json={"items": []},
        report_json={"status": "partial_stock_matched"},
        report_markdown="# Match Engine V0 Report\n\nNERPA D5720-181125SA04\n",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.flush()
    db_session.add(
        MatchCandidate(
            match_run_id=match_run.id,
            distributor_code="ocs",
            item_id="1000841882",
            product_key="1000841882",
            part_number="D5720-181125SA04",
            producer="NERPA",
            category_id="V1100",
            item_name="Server NERPA D5720",
            confidence_score=80,
            price_value=Decimal("6900"),
            price_currency="USD",
            available_quantity=3,
            reservable_locations=1,
            matched_requirements_json=["category"],
            missing_requirements_json=["RAM below requirement"],
            risk_flags_json=["engineer_review_required"],
            raw_json={"spec_item_index": 0},
            created_at=now,
        )
    )
    db_session.commit()
    return match_run.id


def _seed_build_match_run(db_session: Session) -> int:
    now = datetime(2026, 5, 9, 12, 30, tzinfo=UTC)
    raw_json = {
        "spec_item_index": 0,
        "quantity_required": 2,
        "candidate_type": "build_from_parts",
        "components": [
            {
                "role": "server_platform",
                "role_ru": "Платформа",
                "producer": "ASUS",
                "part_number": "90SF03A1-M00070",
                "item_name": "ASUS platform",
                "quantity_required": 2,
                "available_quantity": 2,
            },
            {
                "role": "ram",
                "role_ru": "RAM",
                "producer": "Samsung",
                "part_number": "RAM-64G",
                "item_name": "Samsung RAM",
                "quantity_required": 16,
                "available_quantity": 16,
            },
            {
                "role": "ssd",
                "role_ru": "SSD",
                "producer": "Samsung",
                "part_number": "SSD-960G",
                "item_name": "Samsung SSD",
                "quantity_required": 2,
                "available_quantity": 2,
            },
        ],
        "total_price_value": "59130",
        "total_price_currency": "USD",
        "missing_components": ["Неполная сборка - требуется подбор CPU."],
        "compatibility_warnings": [
            "Требуется инженерная проверка совместимости CPU с выбранной платформой.",
            "2 БП не подтверждены по данным платформы; требуется проверить комплектацию.",
            "Форм-фактор 2U не подтвержден по данным платформы.",
        ],
        "engineer_review_required": True,
        "completeness_status": "incomplete",
        "completeness_label": "Неполная сборка - требуется подбор CPU.",
        "included_component_roles": ["server_platform", "ram", "ssd"],
        "missing_component_roles": ["cpu"],
        "excluded_from_total_roles": ["cpu"],
        "cpu_per_server": 2,
        "total_cpu_required": 4,
        "total_price_note": "без CPU",
    }
    match_run = MatchRun(
        source="text",
        source_text=SERVER_REQUEST,
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=1,
        matched_items=0,
        missing_requirements_json=["Неполная сборка - требуется подбор CPU."],
        risk_flags_json=[
            "Требуется инженерная проверка совместимости CPU с выбранной платформой.",
        ],
        spec_json={"items": []},
        report_json={"status": "partial_stock_matched", "build_candidates": [raw_json]},
        report_markdown="# Match Engine V0 Report\n",
        created_at=now,
    )
    db_session.add(match_run)
    db_session.flush()
    db_session.add(
        MatchCandidate(
            match_run_id=match_run.id,
            distributor_code="ocs",
            item_id="build-1-platform",
            product_key="platform",
            part_number="90SF03A1-M00070",
            producer="ASUS",
            category_id="V110100",
            item_name="Неполная сборка на платформе ASUS 90SF03A1-M00070",
            confidence_score=60,
            price_value=Decimal("59130"),
            price_currency="USD",
            available_quantity=1,
            reservable_locations=1,
            matched_requirements_json=["Платформа выбрана"],
            missing_requirements_json=["Неполная сборка - требуется подбор CPU."],
            risk_flags_json=[
                "Требуется инженерная проверка совместимости CPU с выбранной платформой.",
            ],
            raw_json=raw_json,
            created_at=now,
        )
    )
    db_session.commit()
    return match_run.id


def _grouped_excel_fixture_group() -> dict[str, Any]:
    return {
        "group_id": "cfg_group_1",
        "group_title": "Intel LGA4677 / DDR5 / NVMe",
        "architecture_summary": "Одна компонентная база, несколько платформ.",
        "component_base": {
            "cpu": {
                "role": "cpu",
                "producer": "Intel",
                "part_number": "Xeon Gold 6530",
                "item_name": "Intel Xeon Gold 6530",
                "quantity_required": 4,
                "per_server_quantity": 2,
                "available_quantity": 20,
                "price_value": "500",
                "price_currency": "USD",
                "cpu_cores": 32,
            },
            "ram": {
                "role": "ram",
                "producer": "Micron",
                "part_number": "MTC20F1045S1RC48BA2",
                "item_name": "Micron MTC20F1045S1RC48BA2 32GB DDR5 RDIMM",
                "quantity_required": 32,
                "per_server_quantity": 16,
                "server_quantity": 2,
                "available_quantity": 100,
                "price_value": "100",
                "price_currency": "USD",
                "ram_module_capacity_gb": 32,
                "ram_total_gb_per_server": 512,
            },
            "storage": {
                "role": "ssd",
                "producer": "KIOXIA",
                "part_number": "KCD8XRUG3T84",
                "item_name": "KIOXIA CD8-R 3.84TB U.3 NVMe",
                "quantity_required": 4,
                "per_server_quantity": 2,
                "available_quantity": 40,
                "price_value": "300",
                "price_currency": "USD",
                "storage_capacity_tb": 3.84,
                "facts": {"storage_interface": "U.3 NVMe"},
            },
        },
        "platform_options": [
            {
                "option_id": "platform_option_1_1",
                "role": "cheapest_quote",
                "platform": {
                    "role": "server_platform",
                    "producer": "ASUS",
                    "part_number": "PLATFORM-CHEAP",
                    "quantity_required": 2,
                    "available_quantity": 2,
                },
                "total_price_value": "8600",
                "total_price_currency": "USD",
                "stock_status": "достаточно для текущего запроса",
                "why_this_platform": "Minimal cost build meeting all core specs.",
                "tradeoffs": ["Проверить BIOS и CPU support list."],
                "engineer_checks": ["Проверить CPU support list / BIOS."],
                "engineering_confidence": "preliminary_requires_engineer_review",
            },
            {
                "option_id": "platform_option_1_2",
                "role": "preferred_for_database",
                "platform": {
                    "role": "server_platform",
                    "producer": "Supermicro",
                    "part_number": "SYS-621C-TN12R",
                    "quantity_required": 2,
                    "available_quantity": 2,
                },
                "total_price_value": "10200",
                "total_price_currency": "USD",
                "stock_status": "достаточно для текущего запроса",
                "why_this_platform": (
                    "Proven Supermicro architecture with cost-optimized components."
                ),
                "tradeoffs": ["Premium platform and PCIe 5.0 SSD for critical DB workloads."],
                "engineer_checks": ["Проверить QVL RAM."],
                "engineering_confidence": "preliminary_requires_engineer_review",
            },
        ],
        "recommended_option_id": "platform_option_1_1",
        "why_group_matters": "Платформы сравниваются внутри Intel DDR5 NVMe семьи.",
    }


def _primary_excel_recommendation() -> dict[str, Any]:
    group = _grouped_excel_fixture_group()
    base = group["component_base"]
    platform = group["platform_options"][0]["platform"]
    return {
        "candidate_type": "build_from_parts",
        "title": "Cheapest valid complete stock build",
        "component_candidate_ids": {
            "platform": "platform-cheap",
            "cpu": "cpu-xeon",
            "ram": "ram-micron",
            "ssd": "ssd-kioxia",
        },
        "why_selected": "Minimal cost build meeting all core specs.",
        "assumptions": [],
        "engineer_checks": [],
        "components": [
            platform,
            base["cpu"],
            base["ram"],
            base["storage"],
        ],
        "total_price_value": "8600",
        "total_price_currency": "USD",
    }
