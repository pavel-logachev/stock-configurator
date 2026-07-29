from __future__ import annotations

import inspect
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.db.models import MatchRun
from app.reports import composer_result_normalizer
from app.reports.composer_result_normalizer import (
    COMPOSER_NO_SAFE_COMPLETE_BOM,
    COMPOSER_STRUCTURED_NO_RECOMMENDATION,
    normalize_composer_report_json,
    normalize_composer_result,
)
from app.reports.excel_report import build_match_excel_report
from app.telegram_bot.formatting import format_match_summary


def test_generic_structured_no_recommendation_maps_top_level_reason() -> None:
    result = normalize_composer_result(
        product_group="security",
        primary_object="camera_kit",
        original_request_text="Need a complete security kit",
        requirement_contract={"engineer_checks": ["Confirm mounting constraints."]},
        role_evaluation_coverage_by_role={
            "sensor": {
                "candidate_count": 3,
                "considered_count": 3,
                "all_candidates_considered": True,
            }
        },
        bom_composer_output={
            "no_recommendation": {
                "reason": "No safe complete kit can be assembled.",
                "role_failures": [
                    {
                        "role": "sensor",
                        "reason": "No candidate has the required certification.",
                        "suggested_action": "Buy certified sensors.",
                    }
                ],
                "hard_requirements_failed": [
                    {"requirement_text": "certified motion detection"}
                ],
                "hard_requirements_met": [
                    {"requirement_text": "indoor cameras are available"}
                ],
                "recommended_repair_actions": ["Request certified alternatives."],
                "general_notes": ["Every sensor candidate was considered."],
            }
        },
        final_status_source="composer_no_recommendation",
        primary_recommendation_status="no_recommendation",
    )

    reason = result["no_recommendation_reason"]
    assert result["llm_fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION
    assert reason["product_group"] == "security"
    assert reason["primary_object"] == "camera_kit"
    assert reason["role_failures"][0]["role"] == "sensor"
    assert reason["role_failures"][0]["candidate_coverage"]["considered_count"] == 3
    assert reason["hard_requirements_failed"][0]["requirement_text"] == (
        "certified motion detection"
    )
    assert reason["hard_requirements_met"][0]["requirement_text"] == (
        "indoor cameras are available"
    )
    assert "Every sensor candidate was considered." in reason["diagnostic_notes"]


def test_91_like_structured_no_recommendation_maps_roles_generically() -> None:
    result = normalize_composer_result(
        product_group="server",
        primary_object="server",
        bom_composer_output={
            "no_recommendation": {
                "reason": "Repair pass could not build a complete safe BOM.",
                "role_failures": [
                    {
                        "role": "ssd",
                        "reason": "No available candidate closes the requested drive set.",
                    },
                    {
                        "role": "cable",
                        "reason": "Cable requirement remains unverified.",
                    },
                ],
                "general_notes": ["Role coverage was complete before repair."],
            }
        },
        final_status_source="composer_no_recommendation",
        primary_recommendation_status="no_recommendation",
    )

    reason = result["no_recommendation_reason"]
    assert [row["role"] for row in reason["role_failures"]] == ["ssd", "cable"]
    assert reason["fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION
    assert "Role coverage was complete before repair." in reason["diagnostic_notes"]


def test_network_like_no_recommendation_maps_same_way() -> None:
    reason = _normalized_reason_for_roles(
        "network",
        "network switch",
        ["switch", "transceiver", "support"],
    )

    assert [row["role"] for row in reason["role_failures"]] == [
        "switch",
        "transceiver",
        "support",
    ]
    assert reason["recommended_next_actions"] == ["Procure missing roles."]


def test_storage_like_no_recommendation_maps_same_way() -> None:
    reason = _normalized_reason_for_roles(
        "storage",
        "storage array",
        ["storage_system", "drive_shelf", "license"],
    )

    assert [row["role"] for row in reason["role_failures"]] == [
        "storage_system",
        "drive_shelf",
        "license",
    ]
    assert reason["fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION


def test_structured_aliases_and_critic_diagnostics_map_generically() -> None:
    result = normalize_composer_result(
        product_group="future_group",
        primary_object="future_object",
        repair_composer_output={
            "requirement_analysis": {
                "hard_requirements_met": [
                    {"requirement_text": "rack depth is available"}
                ]
            },
            "no_recommendation": {
                "reason": "Critical shortages prevent a safe complete BOM.",
                "role_level_reasons": [
                    {
                        "role": "compute_node",
                        "reason": "No stocked candidate satisfies the socket request.",
                    }
                ],
                "failures_by_role": {
                    "fabric_adapter": "No safe adapter candidate was verified."
                },
                "missing_roles": ["power_module"],
            },
            "general_notes": ["Some non-blocking accessories are available."],
        },
        completeness_critic_result={
            "missing_roles": ["license_pack"],
            "insufficient_quantities": [
                {
                    "role": "memory_bank",
                    "required_quantity": 8,
                    "available_quantity": 4,
                }
            ],
            "unverified_requirements": [
                {"role": "rack_rail", "requirement_text": "tool-less rails"}
            ],
            "hard_mismatch_risks": [
                {"role": "chassis", "reason": "generation mismatch"}
            ],
            "recommended_repair_actions": ["Ask distributor for alternatives."],
        },
        role_evaluation_coverage_by_role={
            "fabric_adapter": {
                "considered_count": 9,
                "candidate_count": 10,
                "all_candidates_considered": False,
                "failed_chunk_count": 0,
            }
        },
        final_status_source="composer_no_recommendation",
        primary_recommendation_status="no_recommendation",
    )

    reason = result["no_recommendation_reason"]
    roles = {row["role"] for row in reason["role_failures"]}
    assert result["llm_fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION
    assert reason["summary"] == "Critical shortages prevent a safe complete BOM."
    assert {"compute_node", "fabric_adapter", "power_module", "license_pack"} <= roles
    assert reason["recommended_next_actions"] == ["Ask distributor for alternatives."]
    assert reason["hard_mismatch_risks"] == [
        {"role": "chassis", "reason": "generation mismatch"}
    ]
    assert reason["unverified_requirements"] == [
        {"role": "rack_rail", "requirement_text": "tool-less rails"}
    ]
    assert any(
        item.get("role") == "memory_bank" for item in reason["failed_requirements"]
    )
    assert reason["hard_requirements_met"] == [
        {"requirement_text": "rack depth is available"}
    ]
    assert any(
        item.get("source") == "general_notes"
        for item in reason["partial_available_components"]
    )
    assert any("fabric_adapter" in note for note in reason["diagnostic_notes"])
    assert reason["coverage_diagnostics"][0]["role"] == "fabric_adapter"


def test_report_json_post_normalizer_persists_v2_multi_pass_no_recommendation() -> None:
    report_json = normalize_composer_report_json(
        {
            "composer_mode": "multi_pass",
            "product_group": "future_group",
            "primary_object": "future_object",
            "primary_recommendation_status": "no_recommendation",
            "final_status_source": "composer_no_recommendation",
            "llm_fallback_reason": "composer_structured_no_recommendation",
            "no_recommendation_reason": {
                "summary": None,
                "fallback_reason": None,
                "failed_requirements": [],
                "role_failures": [],
                "partial_available_components": [],
                "hard_requirements_met": None,
                "hard_requirements_failed": None,
                "recommended_next_actions": [],
                "diagnostic_notes": None,
            },
            "final_bom_after_repair": {
                "requirement_analysis": {
                    "hard_requirements_met": ["baseline accessories available"]
                },
                "no_recommendation": {
                    "reason": "Repair pass found incompatible hard requirements.",
                    "role_level_reasons": [
                        {"role": "compute_node", "reason": "Socket mismatch."},
                        {"role": "fabric_adapter", "reason": "Stock shortage."},
                    ],
                },
                "general_notes": ["Partial platform information is usable."],
            },
            "completeness_critic_result": {
                "hard_mismatch_risks": [
                    {"role": "compute_node", "reason": "socket mismatch"}
                ],
                "recommended_repair_actions": ["Request compatible replacements."],
            },
            "role_evaluation_coverage_by_role": {
                "fabric_adapter": {
                    "considered_count": 113,
                    "candidate_count": 114,
                    "all_candidates_considered": False,
                    "failed_chunk_count": 0,
                }
            },
        }
    )

    reason = report_json["no_recommendation_reason"]
    assert report_json["llm_fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION
    assert reason["summary"] == "Repair pass found incompatible hard requirements."
    assert [row["role"] for row in reason["role_failures"][:2]] == [
        "compute_node",
        "fabric_adapter",
    ]
    assert reason["recommended_next_actions"] == ["Request compatible replacements."]
    assert reason["hard_mismatch_risks"] == [
        {"role": "compute_node", "reason": "socket mismatch"}
    ]
    assert any("113/114" in note for note in reason["diagnostic_notes"])


def test_structured_fallback_reason_is_not_old_no_proposals_value() -> None:
    result = normalize_composer_result(
        product_group="accessories",
        primary_object="kit",
        bom_composer_output={
            "no_recommendation": {
                "reason": "Composer intentionally declined.",
                "role_failures": [{"role": "mount", "reason": "No safe candidate."}],
            }
        },
        final_status_source="composer_no_recommendation",
        primary_recommendation_status="no_recommendation",
        llm_fallback_reason="llm_configurator_no_proposals",
    )

    assert result["llm_fallback_reason"] == COMPOSER_STRUCTURED_NO_RECOMMENDATION
    assert result["llm_fallback_reason"] != "llm_configurator_no_proposals"


def test_excel_and_telegram_use_normalized_role_failures_and_actions() -> None:
    reason = _normalized_reason_for_roles(
        "network",
        "network switch",
        ["switch", "transceiver", "support"],
    )
    summary = {
        "match_run_id": 77,
        "product_group": "network",
        "llm_configurator_enabled": True,
        "llm_configurator_used": False,
        "primary_recommendation_status": "no_recommendation",
        "no_recommendation_reason": reason,
    }

    telegram_text = format_match_summary(summary)

    assert "Проблемы по ролям:" in telegram_text
    assert "transceiver" in telegram_text
    assert "Procure missing roles." in telegram_text
    assert "Не хватает ролей:" not in telegram_text

    workbook_text = _excel_text(
        {
            **summary,
            "status": "partial_stock_matched",
            "ai_recommendations_count": 0,
            "valid_proposals_count": 0,
        }
    )
    assert "Проблемы по ролям:" in workbook_text
    assert "support" in workbook_text
    assert "Procure missing roles." in workbook_text


def test_valid_recommendation_still_renders_normally() -> None:
    text = format_match_summary(
        {
            "match_run_id": 88,
            "llm_configurator_enabled": True,
            "llm_configurator_used": True,
            "primary_recommendation_status": "valid",
            "primary_recommendation": {
                "candidate_type": "build_from_parts",
                "decision": "recommend",
                "title": "Safe validated build",
                "components": [
                    {
                        "role": "main_unit",
                        "producer": "Acme",
                        "part_number": "A-1",
                        "quantity_required": 1,
                        "available_quantity": 2,
                    }
                ],
                "total_price_value": "100",
                "total_price_currency": "USD",
                "why_selected": "Validated by code checks.",
                "engineer_checks": ["Confirm final compatibility."],
            },
        }
    )

    assert "Безопасную складскую рекомендацию дать нельзя" not in text
    assert "Safe validated build" in text or "Acme" in text


def test_validation_rejected_bom_renders_validation_mismatches_generically() -> None:
    result = normalize_composer_result(
        product_group="ups",
        primary_object="ups",
        code_validation_result={
            "validation_hard_mismatches": [
                {
                    "role": "battery_pack",
                    "type": "quantity_mismatch",
                    "message": "Composer requested more units than stock contains.",
                }
            ],
            "validation_summary": {"validation_rejected_count": 1},
        },
        final_status_source="composer_rejected_by_validation",
        primary_recommendation_status="no_recommendation",
    )

    reason = result["no_recommendation_reason"]
    assert result["llm_fallback_reason"] == COMPOSER_NO_SAFE_COMPLETE_BOM
    assert reason["hard_mismatch_risks"][0]["role"] == "battery_pack"

    telegram_text = format_match_summary(
        {
            "match_run_id": 99,
            "product_group": "ups",
            "llm_configurator_enabled": True,
            "llm_configurator_used": False,
            "primary_recommendation_status": "no_recommendation",
            "final_status_source": "composer_rejected_by_validation",
            "no_recommendation_reason": reason,
        }
    )
    assert "Почему нельзя показать BOM как КП-ready:" in telegram_text
    assert "battery_pack" in telegram_text


def test_no_unsafe_bom_is_shown_as_quote_ready() -> None:
    reason = normalize_composer_result(
        product_group="workstation",
        primary_object="workstation",
        code_validation_result={
            "validation_hard_mismatches": [
                {"role": "gpu", "message": "Selected component is not in stock."}
            ],
            "validation_summary": {"validation_rejected_count": 1},
        },
        final_status_source="composer_rejected_by_validation",
        primary_recommendation_status="no_recommendation",
    )["no_recommendation_reason"]

    text = format_match_summary(
        {
            "match_run_id": 100,
            "product_group": "workstation",
            "llm_configurator_enabled": True,
            "llm_configurator_used": False,
            "primary_recommendation_status": "no_recommendation",
            "primary_recommendation": {"title": "Unsafe BOM"},
            "no_recommendation_reason": reason,
        }
    )

    assert "Unsafe BOM" not in text
    assert "Предварительная спецификация для КП" not in text
    assert "Безопасную складскую рекомендацию дать нельзя" in text


def test_normalizer_has_no_role_or_catalog_hardcode() -> None:
    source = inspect.getsource(composer_result_normalizer)

    for forbidden in (
        "server_platform",
        "ssd",
        "cable",
        "switch",
        "transceiver",
        "storage_system",
        "drive_shelf",
        "vendor",
        "sku",
        "category",
        "ocs",
    ):
        assert f'"{forbidden}"' not in source
        assert f"'{forbidden}'" not in source


def _normalized_reason_for_roles(
    product_group: str,
    primary_object: str,
    roles: list[str],
) -> dict[str, Any]:
    result = normalize_composer_result(
        product_group=product_group,
        primary_object=primary_object,
        bom_composer_output={
            "no_recommendation": {
                "reason": "No safe complete BOM can be assembled.",
                "role_failures": [
                    {"role": role, "reason": f"{role} is not closed from stock."}
                    for role in roles
                ],
                "recommended_next_actions": ["Procure missing roles."],
            }
        },
        final_status_source="composer_no_recommendation",
        primary_recommendation_status="no_recommendation",
    )
    return result["no_recommendation_reason"]


def _excel_text(report_json: dict[str, Any]) -> str:
    match_run = MatchRun(
        id=501,
        source="text",
        source_text="Need safe stock recommendation",
        status="partial_stock_matched",
        engineer_review_required=True,
        total_candidates=0,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={"source_text": "Need safe stock recommendation"},
        report_json=report_json,
        report_markdown="# report",
        created_at=datetime(2026, 5, 24, 12, 0, tzinfo=UTC),
    )
    workbook = load_workbook(BytesIO(build_match_excel_report(match_run)))
    return "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
