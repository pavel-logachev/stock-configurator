from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.db.models import MatchRun
from app.reports.v3_full_category_report import (
    build_v3_full_category_excel_report,
    build_v3_full_category_markdown_report,
)


def test_v3_no_recommendation_report_is_russian_and_compact() -> None:
    report_json = _v3_no_recommendation_report_json()

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=170)

    assert "КП не сформировано" in markdown
    assert "указанная модель процессора не найдена" in markdown
    assert "Не найдено в матрице склада: процессор Intel Xeon Silver 4310 (12 ядер)." in markdown
    assert "Разрешить ближайший технически подходящий аналог" in markdown
    assert "107 категорий" in markdown
    assert "V000, V001" not in markdown
    assert "An exact Intel Xeon Silver 4310" not in markdown
    assert "Allow a compatible" not in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])
    diagnostics_text = _sheet_text(workbook["Склад и диагностика"])

    for sheet_text in (quote_text, review_text):
        assert "КП не сформировано" in sheet_text
        assert "указанная модель процессора не найдена" in sheet_text
        assert (
            "Не найдено в матрице склада: процессор Intel Xeon Silver 4310 (12 ядер)."
            in sheet_text
        )
        assert "Разрешить ближайший технически подходящий аналог" in sheet_text
        assert "An exact Intel Xeon Silver 4310" not in sheet_text
        assert "Allow a compatible" not in sheet_text

    assert "107 категорий" not in quote_text
    assert "matrix_row_count" in diagnostics_text
    assert "V000, V001" not in quote_text


def test_v3_no_recommendation_report_formats_structured_failed_requirements() -> None:
    report_json = _v3_no_recommendation_report_json()
    report_json["no_recommendation_reason"]["summary"] = (
        "КП не сформировано: нет валидного аналога."
    )
    report_json["no_recommendation_reason"]["failed_requirements"] = [
        {
            "requirement_id": "R1",
            "requirement": "Сервер HPE DL380 Gen12",
            "reason": "Поколение Gen12 отсутствует на складе.",
        }
    ]

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=192)
    assert "Требование: R1 Сервер HPE DL380 Gen12" in markdown
    assert "причина: Поколение Gen12 отсутствует на складе." in markdown
    assert '"requirement_id"' not in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])

    for sheet_text in (quote_text, review_text):
        assert "Требование: R1 Сервер HPE DL380 Gen12" in sheet_text
        assert "причина: Поколение Gen12 отсутствует на складе." in sheet_text
        assert '"requirement_id"' not in sheet_text


def test_v3_quote_report_surfaces_deviation_notes() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=188)

    assert "## Отклонения от ТЗ" in markdown
    assert "HPE DL380 Gen12" in markdown
    assert "лучший доступный аналог" in markdown
    assert "В составе по матрице: CPU: Xeon 4510" in markdown
    assert markdown.count("В составе по матрице") == 1
    assert "## Целевые объекты" in markdown
    assert "HPE DL380 Gen11 - выбран - строка L1" in markdown
    assert "Intel Xeon Gold 6544Y - 6 шт." in markdown
    assert "Точная позиция не найдена в переданной матрице." in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])

    assert "В составе по матрице: CPU: Xeon 4510" in quote_text
    assert quote_text.count("В составе по матрице") == 1
    assert "Целевые объекты" in quote_text
    assert "HPE DL380 Gen11" in quote_text
    assert "выбран" in quote_text
    assert "L1" in quote_text
    assert "Позиция" in quote_text
    assert "Кол-во" in quote_text
    assert "Причина" in quote_text
    assert "Intel Xeon Gold 6544Y" in quote_text
    assert "6" in quote_text
    assert "Точная позиция не найдена в переданной матрице." in quote_text
    assert "причина: Точная позиция" not in quote_text
    assert "Intel Xeon Gold 6544Y" in review_text
    assert "Точная позиция не найдена в переданной матрице." in review_text
    for sheet_text in (quote_text, review_text):
        assert "Отклонения от ТЗ" in sheet_text
        assert "HPE DL380 Gen12" in sheet_text
        assert "лучший доступный аналог" in sheet_text


def test_v3_quote_report_shows_multi_currency_total_and_groups_commercial_lines() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()
    quote = report_json["validated_quote"]
    quote["total_price_value"] = None
    quote["total_price_currency"] = None
    quote["totals_by_currency"] = [
        {"currency": "USD", "value": "218147.00"},
        {"currency": "RUR", "value": "2867815.05"},
    ]
    quote["lines"] = [
        {
            "component_candidate_id": "ocs:3000048380",
            "role": "heatsink",
            "producer": "HPE",
            "part_number": "P48818-B21",
            "item_name": "HPE High Performance Heat Sink Kit",
            "quantity": 1,
            "unit_price_value": "132.00",
            "unit_price_currency": "USD",
            "line_total_value": "132.00",
            "line_total_currency": "USD",
            "reason": "Первый складской остаток.",
            "stock_row_id": "ocs:3000048380:56724",
        },
        {
            "component_candidate_id": "ocs:3000048380",
            "role": "heatsink",
            "producer": "HPE",
            "part_number": "P48818-B21",
            "item_name": "HPE High Performance Heat Sink Kit",
            "quantity": 1,
            "unit_price_value": "132.00",
            "unit_price_currency": "USD",
            "line_total_value": "132.00",
            "line_total_currency": "USD",
            "reason": "Второй складской остаток.",
            "stock_row_id": "ocs:3000048380:57581",
        },
    ]

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=253)
    assert "218 147 USD" in markdown
    assert "2 867 815,05 RUR" in markdown
    assert markdown.count("HPE P48818-B21 HPE High Performance Heat Sink Kit") == 1
    assert "кол-во 2" in markdown
    assert "сумма 264 USD" in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    diagnostics_text = _sheet_text(workbook["Склад и диагностика"])

    assert "218 147 USD" in quote_text
    assert "2 867 815,05 RUR" in quote_text
    assert quote_text.count("HPE High Performance Heat Sink Kit") == 1
    assert "2" in quote_text
    assert "264" in quote_text
    assert diagnostics_text.count("HPE High Performance Heat Sink Kit") == 2


def test_v3_quote_report_surfaces_available_alternatives() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()
    report_json["validated_quote"]["available_alternatives"] = [
        {
            "requirement_id": "R_HBA",
            "component_candidate_id": "treolan:014002/1",
            "item": "HPE SN1100Q 16Gb Dual Port FC HBA",
            "stock_row_id": "treolan:014002/1:123",
            "available_quantity": 3,
            "quantity_is_greater_than": True,
            "unit_price_value": "1176.00",
            "unit_price_currency": "USD",
            "reason": "Ниже скорость: 16Gb FC вместо запрошенных 32Gb FC.",
        }
    ]
    report_json["validated_quote"]["quote_integrity"] = {
        "version": "quote_integrity_reconciler_v3",
        "status": "mechanically_adjusted",
        "adjustments": [
            {
                "type": "stock_row_id_repaired",
                "section": "available_alternative",
                "index": 1,
                "resolution": "unique_component_candidate",
                "component_candidate_id": "treolan:014002/1",
                "original_stock_row_id": "treolan:014002/1:999",
                "resolved_stock_row_id": "treolan:014002/1:123",
            }
        ],
        "warnings": ["quote_integrity.stock_row_id_repaired"],
    }

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=188)

    assert "## Доступные варианты для согласования - не включены в итог" in markdown
    assert "требование: R_HBA" in markdown
    assert "HPE SN1100Q 16Gb Dual Port FC HBA" in markdown
    assert "доступно: 3+" in markdown
    assert "цена: 1 176 USD" in markdown
    assert "stock_row_id: treolan:014002/1:123" in markdown
    assert "16Gb FC вместо запрошенных 32Gb FC" in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])
    diagnostics_text = _sheet_text(workbook["Склад и диагностика"])

    assert "Доступные варианты для согласования - не включены в итог" in quote_text
    assert "R_HBA" in quote_text
    assert "HPE SN1100Q 16Gb Dual Port FC HBA" in quote_text
    assert "3+" in quote_text
    assert "treolan:014002/1:123" not in quote_text
    assert "treolan:014002/1:999" not in quote_text
    assert "1 176" in quote_text
    assert "USD" in quote_text
    assert "Ниже скорость: 16Gb FC вместо запрошенных 32Gb FC." in quote_text

    assert "Доступные альтернативы, raw" in diagnostics_text
    assert "Корректировки quote_integrity" in diagnostics_text
    assert "stock_row_id_repaired" in diagnostics_text
    assert "unique_component_candidate" in diagnostics_text
    assert "treolan:014002/1:999" in diagnostics_text

    for sheet_text in (review_text, diagnostics_text):
        assert "HPE SN1100Q 16Gb Dual Port FC HBA" in sheet_text
        assert "treolan:014002/1:123" in sheet_text
        assert "R_HBA" in sheet_text
        assert "1 176 USD" in sheet_text
        assert "Ниже скорость: 16Gb FC вместо запрошенных 32Gb FC." in sheet_text
    assert "Доступные варианты для согласования - не включены в итог" in review_text


def test_v3_quote_report_surfaces_lower_bound_stock_confirmation() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()
    line = report_json["validated_quote"]["lines"][0]
    line.update(
        {
            "component_candidate_id": "ocs:3000048406",
            "stock_row_id": "ocs:3000048406:106561",
            "producer": "HPE",
            "part_number": "P71674-425",
            "item_name": "HPE ProLiant DL380 Gen11",
            "quantity": 3,
            "quantity_value": 1,
            "quantity_is_greater_than": True,
            "stock_confirmation_required": True,
        }
    )
    report_json["validated_quote"]["quote_integrity"] = {
        "version": "quote_integrity_reconciler_v9",
        "status": "mechanically_adjusted",
        "adjustments": [
            {
                "type": "stock_lower_bound_quantity_confirm",
                "section": "line",
                "index": 1,
                "component_candidate_id": "ocs:3000048406",
                "stock_row_id": "ocs:3000048406:106561",
                "displayed_available_quantity": 1,
                "included_quantity": 3,
                "resolution": "kept_llm_quantity_with_stock_confirmation",
            }
        ],
        "warnings": ["quote_integrity.stock_lower_bound_requires_confirmation"],
    }

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=291)

    assert "Подтвердить количество по складу" in markdown
    assert "P71674-425" in markdown
    assert "в КП 3 шт." in markdown
    assert "склад показывает 1+ шт." in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])
    diagnostics_text = _sheet_text(workbook["Склад и диагностика"])

    for sheet_text in (quote_text, review_text):
        assert "Подтвердить количество по складу" in sheet_text
        assert "P71674-425" in sheet_text
        assert "3 шт." in sheet_text
        assert "1+ шт." in sheet_text
        assert (
            "Перед отправкой КП нужно подтвердить доступность выбранного количества."
            in sheet_text
        )
    assert "stock_lower_bound_quantity_confirm" in diagnostics_text


def test_v3_quote_report_keeps_commercial_sheet_free_of_raw_stock_fields() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()
    line = report_json["validated_quote"]["lines"][0]
    line.update(
        {
            "producer": "",
            "stock_row_id": "treolan:servers:777",
            "part_number": "P71674-425",
            "item_name": (
                "Hewlett Packard Enterprise / P71674-425 Сервер HP ProLiant DL380 Gen11 "
                "/ treolan.prid:777 / sale:available / GTIN:123456"
            ),
        }
    )

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    diagnostics_text = _sheet_text(workbook["Склад и диагностика"])

    assert "P71674-425" in quote_text
    assert "Hewlett Packard Enterprise" in quote_text
    assert "Сервер HP ProLiant DL380 Gen11" in quote_text
    assert "treolan:servers:777" not in quote_text
    assert "treolan.prid" not in quote_text
    assert "sale:available" not in quote_text
    assert "GTIN:123456" not in quote_text
    assert "treolan:servers:777" in diagnostics_text


def test_v3_quote_report_surfaces_v6_partial_gaps() -> None:
    report_json = _v3_quote_report_json_with_deviation_notes()
    quote = report_json["validated_quote"]
    quote.update(
        {
            "client_status_label": "Ближайший складской вариант требует добора",
            "selection_mode": "partial_build",
            "completeness_status": "partial",
            "operational_status": "incomplete_needs_procurement",
            "client_summary": "Можно поставить базовую платформу со склада.",
            "coverage_summary": "База закрыта, CPU нужно добрать.",
            "key_deviations": [
                {
                    "requirement_id": "R2",
                    "requested": "2 CPU",
                    "offered": "не выбрано",
                    "direction": "different",
                    "severity": "material",
                    "impact": "Без CPU система неполная.",
                    "reason": "Совместимый CPU не подтвержден по матрице.",
                }
            ],
            "procurement_gaps": [
                {
                    "requirement_id": "R2",
                    "role": "cpu",
                    "requested": "2 CPU",
                    "status": "no_compatible_item_proven",
                    "required_for": "operational_readiness",
                    "impact": "Нужно добрать CPU.",
                    "next_action": "Запросить CPU под выбранную базу.",
                }
            ],
        }
    )

    markdown = build_v3_full_category_markdown_report(report_json, match_run_id=193)

    assert "## Тип предложения" in markdown
    assert "Ближайший складской вариант требует добора" in markdown
    assert "## Что нужно добрать или согласовать" in markdown
    assert "совместимый вариант не подтвержден по матрице" in markdown

    workbook = load_workbook(
        BytesIO(build_v3_full_category_excel_report(_match_run(report_json)))
    )
    quote_text = _sheet_text(workbook["КП"])
    review_text = _sheet_text(workbook["Инженерная проверка"])

    for sheet_text in (quote_text, review_text):
        assert "Что нужно добрать или согласовать" in sheet_text
        assert "совместимый вариант не подтвержден по матрице" in sheet_text
        assert "Нужно добрать CPU." in sheet_text


def _match_run(report_json: dict[str, Any]) -> MatchRun:
    return MatchRun(
        id=170,
        source="v3_full_category_text",
        source_text="Сервер 1 x CPU Intel Xeon Silver 4310",
        status="no_recommendation",
        engineer_review_required=True,
        total_candidates=0,
        matched_items=0,
        missing_requirements_json=[],
        risk_flags_json=[],
        spec_json={},
        report_json=report_json,
        report_markdown=None,
        created_at=datetime(2026, 6, 17, 14, 35, tzinfo=UTC),
    )


def _v3_quote_report_json_with_deviation_notes() -> dict[str, Any]:
    return {
        "pipeline_version": "v3_full_category_matrix",
        "v3_result_state": "quote_draft_review_required",
        "v3_profile": "server",
        "distributor_code": "ocs",
        "category_ids": ["V1100"],
        "diagnostics": {
            "matrix_row_count": 538,
            "matrix_component_count": 538,
            "model": "test-model",
        },
        "validated_quote": {
            "title": "Лучший доступный аналог",
            "total_price_value": "8121.3000",
            "total_price_currency": "USD",
            "why_selected": "Выбран лучший доступный аналог из складской матрицы.",
            "target_decisions": [
                {
                    "target_label": "HPE DL380 Gen11",
                    "anchor_status": "selected",
                    "anchor_line_id": "L1",
                    "reason": "Ближайшая складская база.",
                }
            ],
            "deviation_notes": [
                (
                    "Исходное требование: HPE DL380 Gen12; выбрано: "
                    "складской аналог другого поколения; класс: другое; "
                    "влияние: требуется согласование замены."
                )
            ],
            "price_audit": [],
            "assumptions": [],
            "procurement_gaps": [
                {
                    "item": "Intel Xeon Gold 6544Y",
                    "quantity": 6,
                    "reason": "Точная позиция не найдена в переданной матрице.",
                }
            ],
            "engineer_checks": [],
            "compatibility_check": {
                "status": "compatible",
                "checked_facts": ["stock row fits selected analog"],
                "blocking_mismatches": [],
                "unresolved_risks": [],
            },
            "lines": [
                {
                    "role": "platform",
                    "producer": "HPE",
                    "part_number": "DL380-G11",
                    "item_name": "Server platform",
                    "quantity": 1,
                    "unit_price_value": "8121.3000",
                    "unit_price_currency": "USD",
                    "line_total_value": "8121.3000",
                    "line_total_currency": "USD",
                    "reason": (
                        "Ближайшая база. В составе по матрице: "
                        "CPU: Xeon 4510; RAM: 64GB DDR5."
                    ),
                    "included_components_summary": {
                        "cpu": "Xeon 4510",
                        "ram": "64GB DDR5",
                        "raid_or_controller": "MR408i-o",
                        "storage": "2x960GB SATA",
                        "network": "4x1GbE",
                        "power": "2x1000W",
                    },
                }
            ],
            "engineering_review_required": True,
        },
        "v3_validation_errors": [],
        "v3_validation_warnings": [],
    }


def _v3_no_recommendation_report_json() -> dict[str, Any]:
    return {
        "pipeline_version": "v3_full_category_matrix",
        "v3_result_state": "no_recommendation",
        "v3_profile": "server",
        "distributor_code": "treolan",
        "category_ids": [f"V{index:03d}" for index in range(107)],
        "diagnostics": {
            "matrix_row_count": 696,
            "matrix_component_count": 696,
            "model": "test-model",
        },
        "no_recommendation_reason": {
            "summary": (
                "An exact Intel Xeon Silver 4310 CPU is required, but this model is "
                "not present in the supplied matrix."
            ),
            "details": (
                "The matrix contains other CPUs, but the request did not explicitly "
                "allow a substitute, so no safe recommendation can be made."
            ),
            "failed_requirements": [
                "CPU model: Intel Xeon Silver 4310 (12 cores) is missing from the matrix."
            ],
            "recommended_next_actions": [
                "Allow a compatible technical alternative from stock.",
                "Check another distributor if the exact CPU is mandatory.",
            ],
        },
        "v3_validation_errors": [],
        "v3_validation_warnings": [],
    }


def _sheet_text(sheet: Any) -> str:
    return "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
